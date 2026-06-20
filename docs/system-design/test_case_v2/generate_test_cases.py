"""
HomeLodge Test Case Generator
Generates one Excel file per use case, with tabs for each flow type.
Format matches the existing test_case_example.xlsx structure.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

OUTPUT_DIR = r"c:\Users\Admin\Documents\HomeLodge-Doc\docs\system-design\test_case_v2"
CREATED_BY = "Aisyah"
VERSION = "1.0"

# ─────────────────────────────────────────────────────────────
# Colour palette
# ─────────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")  # dark blue – label cells
VALUE_FILL   = PatternFill("solid", fgColor="D6E4F0")  # light blue – value cells
TABLE_HEAD   = PatternFill("solid", fgColor="2E75B6")  # mid blue – table header row
ALT_ROW      = PatternFill("solid", fgColor="EBF3FB")  # very light blue – alternate step rows
WHITE_FILL   = PatternFill("solid", fgColor="FFFFFF")
NORMAL_FILL  = PatternFill("solid", fgColor="E2EFDA")  # green tint for normal flow tabs
ALT_FILL     = PatternFill("solid", fgColor="FFF2CC")  # yellow tint for alternate flow tabs
EXCEP_FILL   = PatternFill("solid", fgColor="FCE4D6")  # orange tint for exception flow tabs

THIN  = Side(style="thin",  color="AAAAAA")
MED   = Side(style="medium", color="1F4E79")
OUTER = Border(left=MED, right=MED, top=MED, bottom=MED)
INNER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WHITE_FONT = Font(name="Calibri", color="FFFFFF", bold=True, size=10)
LABEL_FONT = Font(name="Calibri", color="FFFFFF", bold=True, size=10)
VALUE_FONT = Font(name="Calibri", color="1F4E79", size=10)
BODY_FONT  = Font(name="Calibri", size=10)
HEAD_FONT  = Font(name="Calibri", color="FFFFFF", bold=True, size=10)
TITLE_FONT = Font(name="Calibri", bold=True, size=11, color="1F4E79")


def apply_border(cell, border=INNER):
    cell.border = border


def write_header_row(ws, row, col, label, value, label_fill=HEADER_FILL, value_fill=VALUE_FILL, merge_label=2, merge_value=3):
    """Write a label+value pair and merge cells."""
    lc = ws.cell(row=row, column=col, value=label)
    lc.font = LABEL_FONT
    lc.fill = label_fill
    lc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    apply_border(lc, INNER)
    # merge label cells
def build_sheet(ws, tc_id, tc_desc, prerequisites, test_data, test_scenario, steps, flow_color=None):
    """
    Populate a worksheet with the standard test case template.

    steps: list of (step_details, expected_result)
    prerequisites: list of strings
    test_data: list of strings
    """
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 28
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 22

    # Row 1: Test Case ID (A-B label, C-D value) | Test Case Description (E-F label, G-H value)
    ws["A1"] = "Test Case ID"
    ws["A1"].font = LABEL_FONT; ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center"); ws["A1"].border = INNER
    ws.merge_cells("A1:B1")
    ws["C1"] = tc_id
    ws["C1"].font = VALUE_FONT; ws["C1"].fill = VALUE_FILL
    ws["C1"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True); ws["C1"].border = INNER
    ws.merge_cells("C1:D1")
    ws["E1"] = "Test Case Description"
    ws["E1"].font = LABEL_FONT; ws["E1"].fill = HEADER_FILL
    ws["E1"].alignment = Alignment(horizontal="left", vertical="center"); ws["E1"].border = INNER
    ws["F1"].border = INNER  # placeholder for merged area
    ws.merge_cells("E1:F1")
    ws["G1"] = tc_desc
    ws["G1"].font = VALUE_FONT; ws["G1"].fill = VALUE_FILL
    ws["G1"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True); ws["G1"].border = INNER
    ws.merge_cells("G1:H1")
    ws.row_dimensions[1].height = 25

    # Row 2: Created By | Reviewed By | Version
    ws["A2"] = "Created By"
    ws["A2"].font = LABEL_FONT; ws["A2"].fill = HEADER_FILL
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center"); ws["A2"].border = INNER
    ws["B2"] = CREATED_BY
    ws["B2"].font = VALUE_FONT; ws["B2"].fill = VALUE_FILL
    ws["B2"].alignment = Alignment(horizontal="left", vertical="center"); ws["B2"].border = INNER
    ws.merge_cells("B2:C2")
    ws["D2"] = "Reviewed By"
    ws["D2"].font = LABEL_FONT; ws["D2"].fill = HEADER_FILL
    ws["D2"].alignment = Alignment(horizontal="left", vertical="center"); ws["D2"].border = INNER
    ws["E2"] = ""
    ws["E2"].font = VALUE_FONT; ws["E2"].fill = VALUE_FILL
    ws["E2"].alignment = Alignment(horizontal="left", vertical="center"); ws["E2"].border = INNER
    ws.merge_cells("E2:F2")
    ws["G2"] = "Version"
    ws["G2"].font = LABEL_FONT; ws["G2"].fill = HEADER_FILL
    ws["G2"].alignment = Alignment(horizontal="left", vertical="center"); ws["G2"].border = INNER
    ws["H2"] = VERSION
    ws["H2"].font = VALUE_FONT; ws["H2"].fill = VALUE_FILL
    ws["H2"].alignment = Alignment(horizontal="left", vertical="center"); ws["H2"].border = INNER

    # Row 3: QA Tester's Log
    ws["A3"] = "QA Tester's Log"
    ws["A3"].font = LABEL_FONT; ws["A3"].fill = HEADER_FILL
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center"); ws["A3"].border = INNER
    ws.merge_cells("A3:B3")
    ws["C3"] = ""
    ws["C3"].font = VALUE_FONT; ws["C3"].fill = VALUE_FILL
    ws["C3"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True); ws["C3"].border = INNER
    ws.merge_cells("C3:H3")

    # Row 4: Tester's Name | Date Tested | Pass/Fail
    ws["A4"] = "Tester's Name"
    ws["A4"].font = LABEL_FONT; ws["A4"].fill = HEADER_FILL
    ws["A4"].alignment = Alignment(horizontal="left", vertical="center"); ws["A4"].border = INNER
    ws["B4"] = ""
    ws["B4"].font = VALUE_FONT; ws["B4"].fill = VALUE_FILL
    ws["B4"].alignment = Alignment(horizontal="left", vertical="center"); ws["B4"].border = INNER
    ws.merge_cells("B4:C4")
    ws["D4"] = "Date Tested"
    ws["D4"].font = LABEL_FONT; ws["D4"].fill = HEADER_FILL
    ws["D4"].alignment = Alignment(horizontal="left", vertical="center"); ws["D4"].border = INNER
    ws["E4"] = ""
    ws["E4"].font = VALUE_FONT; ws["E4"].fill = VALUE_FILL
    ws["E4"].alignment = Alignment(horizontal="left", vertical="center"); ws["E4"].border = INNER
    ws.merge_cells("E4:F4")
    ws["G4"] = "Test Case (Pass/Fail/Not Executed)"
    ws["G4"].font = LABEL_FONT; ws["G4"].fill = HEADER_FILL
    ws["G4"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True); ws["G4"].border = INNER
    ws["H4"] = ""
    ws["H4"].font = VALUE_FONT; ws["H4"].fill = VALUE_FILL
    ws["H4"].alignment = Alignment(horizontal="left", vertical="center"); ws["H4"].border = INNER
    ws.row_dimensions[4].height = 30

    # ── Row 5: Prerequisites header / Test Data header ──────
    ws["A5"] = "S #"
    ws["A5"].font = HEAD_FONT; ws["A5"].fill = TABLE_HEAD
    ws["A5"].alignment = Alignment(horizontal="center", vertical="center"); ws["A5"].border = INNER

    ws["B5"] = "Prerequisites:"
    ws["B5"].font = HEAD_FONT; ws["B5"].fill = TABLE_HEAD
    ws["B5"].alignment = Alignment(horizontal="left", vertical="center"); ws["B5"].border = INNER
    ws.merge_cells("B5:D5")

    ws["E5"] = "S #"
    ws["E5"].font = HEAD_FONT; ws["E5"].fill = TABLE_HEAD
    ws["E5"].alignment = Alignment(horizontal="center", vertical="center"); ws["E5"].border = INNER

    ws["F5"] = "Test Data"
    ws["F5"].font = HEAD_FONT; ws["F5"].fill = TABLE_HEAD
    ws["F5"].alignment = Alignment(horizontal="left", vertical="center"); ws["F5"].border = INNER
    ws.merge_cells("F5:H5")

    # ── Rows 6–N: Prerequisites + Test Data ─────────────────
    max_rows = max(len(prerequisites), len(test_data), 3)
    for i in range(max_rows):
        r = 6 + i
        fill = ALT_ROW if i % 2 == 0 else WHITE_FILL

        ws.cell(row=r, column=1, value=i + 1).font = BODY_FONT
        ws.cell(row=r, column=1).fill = fill
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=1).border = INNER

        pr_text = prerequisites[i] if i < len(prerequisites) else None
        ws.cell(row=r, column=2, value=pr_text).font = BODY_FONT
        ws.cell(row=r, column=2).fill = fill
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.cell(row=r, column=2).border = INNER
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)

        ws.cell(row=r, column=5, value=i + 1).font = BODY_FONT
        ws.cell(row=r, column=5).fill = fill
        ws.cell(row=r, column=5).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=5).border = INNER

        td_text = test_data[i] if i < len(test_data) else None
        ws.cell(row=r, column=6, value=td_text).font = BODY_FONT
        ws.cell(row=r, column=6).fill = fill
        ws.cell(row=r, column=6).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.cell(row=r, column=6).border = INNER
        ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=8)

    # ── Test Scenario row ────────────────────────────────────
    scenario_row = 6 + max_rows
    ws.cell(row=scenario_row, column=1, value="Test Scenario").font = LABEL_FONT
    ws.cell(row=scenario_row, column=1).fill = HEADER_FILL
    ws.cell(row=scenario_row, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=scenario_row, column=1).border = INNER
    ws.merge_cells(start_row=scenario_row, start_column=1, end_row=scenario_row, end_column=2)

    ws.cell(row=scenario_row, column=3, value=test_scenario).font = TITLE_FONT
    ws.cell(row=scenario_row, column=3).fill = VALUE_FILL
    ws.cell(row=scenario_row, column=3).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.cell(row=scenario_row, column=3).border = INNER
    ws.merge_cells(start_row=scenario_row, start_column=3, end_row=scenario_row, end_column=8)
    ws.row_dimensions[scenario_row].height = 30

    # ── Steps table header ───────────────────────────────────
    head_row = scenario_row + 1
    ws.cell(row=head_row, column=1, value="Step #").font = HEAD_FONT
    ws.cell(row=head_row, column=1).fill = TABLE_HEAD
    ws.cell(row=head_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=head_row, column=1).border = INNER

    ws.cell(row=head_row, column=2, value="Step Details").font = HEAD_FONT
    ws.cell(row=head_row, column=2).fill = TABLE_HEAD
    ws.cell(row=head_row, column=2).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=head_row, column=2).border = INNER
    ws.merge_cells(start_row=head_row, start_column=2, end_row=head_row, end_column=4)

    ws.cell(row=head_row, column=5, value="Expected Results").font = HEAD_FONT
    ws.cell(row=head_row, column=5).fill = TABLE_HEAD
    ws.cell(row=head_row, column=5).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=head_row, column=5).border = INNER
    ws.merge_cells(start_row=head_row, start_column=5, end_row=head_row, end_column=6)

    ws.cell(row=head_row, column=7, value="Actual Results").font = HEAD_FONT
    ws.cell(row=head_row, column=7).fill = TABLE_HEAD
    ws.cell(row=head_row, column=7).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=head_row, column=7).border = INNER

    ws.cell(row=head_row, column=8, value="Pass / Fail / Not executed / Suspended").font = HEAD_FONT
    ws.cell(row=head_row, column=8).fill = TABLE_HEAD
    ws.cell(row=head_row, column=8).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.cell(row=head_row, column=8).border = INNER
    ws.row_dimensions[head_row].height = 28

    # ── Step rows ────────────────────────────────────────────
    for i, (detail, expected) in enumerate(steps):
        r = head_row + 1 + i
        fill = ALT_ROW if i % 2 == 0 else WHITE_FILL
        s_act.font = BODY_FONT; s_act.fill = WHITE_FILL
        s_act.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        apply_border(s_act, INNER)

        s_pf = ws.cell(row=r, column=8, value="")
        s_pf.font = BODY_FONT; s_pf.fill = WHITE_FILL
        s_pf.alignment = Alignment(horizontal="center", vertical="center")
        apply_border(s_pf, INNER)

    # Tab colour
    ws.sheet_properties.tabColor = flow_color.fgColor if flow_color else "D6E4F0"


def create_workbook(filename, sheets_data):
    """
    sheets_data: list of dicts with keys:
        tab_name, tc_id, tc_desc, prerequisites, test_data,
        test_scenario, steps, flow_color
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    for sd in sheets_data:
        ws = wb.create_sheet(title=sd["tab_name"])
        build_sheet(
            ws=ws,
            tc_id=sd["tc_id"],
            tc_desc=sd["tc_desc"],
            prerequisites=sd["prerequisites"],
            test_data=sd["test_data"],
            test_scenario=sd["test_scenario"],
            steps=sd["steps"],
            flow_color=sd.get("flow_color"),
        )

    path = os.path.join(OUTPUT_DIR, filename)
    wb.save(path)
    print(f"  Saved: {filename}")


# ══════════════════════════════════════════════════════════════
# MODULE 1 – AUTHENTICATION
# ══════════════════════════════════════════════════════════════

def gen_auth01():
    """UC-AUTH-01: Register Account (Email / Password)"""
    sheets = [
        {
            "tab_name": "Normal Flow",
            "tc_id": "TC-AUTH-01_NF",
            "tc_desc": "Register Account – Normal Flow (Successful Registration)",
            "prerequisites": [
                "The email 'newguest@example.com' is NOT registered.",
                "The registration page is publicly accessible.",
            ],
            "test_data": [
                "Full Name: 'John Doe'",
                "Email: 'newguest@example.com'",
                "Password: 'Pass@1234'",
                "Confirm Password: 'Pass@1234'",
            ],
            "test_scenario": "Verify that a new visitor can successfully create a HomeLodge account with valid details and is redirected to the sign-in page.",
            "steps": [
                ("Open the registration page.", "The registration form is displayed with fields: Full Name, Email, Password, Confirm Password."),
                ("Enter Full Name: 'John Doe', Email: 'newguest@example.com'.", "Fields accept the input."),
                ("Enter Password: 'Pass@1234' and Confirm Password: 'Pass@1234'.", "Password strength indicator shows all rules met."),
                ("Click the 'Register' button.", "Form is submitted."),
                ("Observe the system response.", "A success message is shown. User is redirected to the sign-in page."),
                ("Check the database.", "A new account record exists with email 'newguest@example.com', role 'Guest', status 'active'."),
            ],
            "flow_color": NORMAL_FILL,
        },
        {
            "tab_name": "Alternate Flow",
            "tc_id": "TC-AUTH-01_AF",
            "tc_desc": "Register Account – Alternate Flows (Validation Failures)",
            "prerequisites": [
                "The email 'existing@example.com' is already registered (for A1).",
                "Registration page is accessible.",
            ],
            "test_data": [
                "A1 Email: 'existing@example.com'",
                "A2 Password: 'weak' (fails strength)",
                "A3 Confirm Password: 'Different1!'",
                "A4 Email: 'not-an-email'",
            ],
            "test_scenario": "Verify the system handles all alternate validation scenarios for registration correctly.",
            "steps": [
                ("(A1) Enter Email: 'existing@example.com' with valid other fields. Click Register.", "Email field is highlighted. Error: 'Email already in use.' Suggestions to Sign In or use Forgot Password are shown. No account is created."),
                ("(A2) Enter Password: 'weak' (too short, no uppercase/symbol). Click Register.", "Password strength rules that are not met are shown. Submission is blocked."),
                ("(A3) Enter Password: 'Pass@1234' and Confirm Password: 'Different1!'. Click Register.", "Confirm password field is highlighted. Error: 'Passwords do not match.' Submission is blocked."),
                ("(A4) Enter Email: 'not-an-email'. Click Register.", "Email field is highlighted. Error indicating invalid email format is shown. Submission is blocked."),
            ],
            "flow_color": ALT_FILL,
        },
        {
            "tab_name": "Exception Flow",
            "tc_id": "TC-AUTH-01_EF",
            "tc_desc": "Register Account – Exception Flow (System Error)",
            "prerequisites": [
                "All form fields are filled with valid data.",
                "A system/database error is simulated during account creation.",
            ],
            "test_data": [
                "Full Name: 'Jane Doe'",
                "Email: 'jane@example.com'",
                "Password: 'Pass@1234'",
                "Confirm Password: 'Pass@1234'",
            ],
            "test_scenario": "Verify that a system error during account creation displays an error message and does NOT create a partial account.",
            "steps": [
                ("Fill in all valid registration details.", "Fields accept input."),
                ("Click 'Register' while a database/server error is simulated.", "Form is submitted."),
                ("Observe system response.", "An error message is shown: 'Registration failed. Please try again.' No new account record exists in the database."),
            ],
            "flow_color": EXCEP_FILL,
        },
    ]
    create_workbook("TC-AUTH-01_RegisterAccount.xlsx", sheets)


def gen_auth02():
    """UC-AUTH-02: Login / Login via Google SSO"""
    sheets = [
        {
            "tab_name": "Normal Flow – Email",
            "tc_id": "TC-AUTH-02_NF_Email",
            "tc_desc": "Login – Normal Flow (Email & Password, No Forced Change)",
            "prerequisites": [
                "User 'user@example.com' / 'Pass@1234' exists and is active.",
                "No forced password change flag is set on the account.",
            ],
            "test_data": [
                "Email: 'user@example.com'",
                "Password: 'Pass@1234'",
            ],
            "test_scenario": "Verify that a registered user can sign in with valid email and password and is redirected to the dashboard.",
            "steps": [
                ("Navigate to the sign-in page.", "Sign-in form is displayed with Email and Password fields."),
                ("Enter Email: 'user@example.com' and Password: 'Pass@1234'.", "Fields accept input."),
                ("Click 'Sign In'.", "Credentials are submitted."),
                ("Observe the system response.", "User is signed in. Sign-in timestamp is recorded. User is redirected to the dashboard."),
            ],
            "flow_color": NORMAL_FILL,
        },
        {
            "tab_name": "Normal Flow – Google",
            "tc_id": "TC-AUTH-02_NF_Google",
            "tc_desc": "Login – Normal Flow (Google SSO, First-Time & Existing Account)",
            "prerequisites": [
                "Valid Google account is available.",
                "Scenario A: Google email 'google@example.com' already has a HomeLodge account.",
                "Scenario B: Google email 'newgoogle@example.com' has no HomeLodge account.",
            ],
            "test_data": [
                "Google Account A: 'google@example.com' (existing HomeLodge account)",
                "Google Account B: 'newgoogle@example.com' (no HomeLodge account)",
            ],
            "test_scenario": "Verify Google SSO sign-in for both existing and new accounts.",
            "steps": [
                ("(Scenario A) Click 'Continue with Google'. Authenticate with Google account A.", "Google returns name and email. System finds existing account and links Google identity if not already linked. User is redirected to dashboard."),
                ("(Scenario B) Click 'Continue with Google'. Authenticate with Google account B.", "Google returns name and email. No existing account found. System creates a new 'Guest' account. User is redirected to dashboard."),
                ("(Scenario B) Verify new account in database.", "New account exists with email 'newgoogle@example.com', role 'Guest', status 'active'."),
            ],
            "flow_color": NORMAL_FILL,
        },
        {
            "tab_name": "Normal Flow – Force Chg",
            "tc_id": "TC-AUTH-02_NF_Force",
            "tc_desc": "Login – Normal Flow (Forced Password Change on First Login)",
            "prerequisites": [
                "User 'forcechange@example.com' has a forced password change flag set.",
                "User knows the temporary password 'Temp@1234'.",
            ],
            "test_data": [
                "Email: 'forcechange@example.com'",
                "Temporary Password: 'Temp@1234'",
            ],
            "test_scenario": "Verify that a user with a forced password change flag is redirected to the change password page after login and cannot access other pages until done.",
            "steps": [
                ("Navigate to sign-in page. Enter email and temporary password. Click 'Sign In'.", "Credentials are verified. Forced change flag is detected."),
                ("Observe redirect.", "User is redirected to the 'Change Your Password' page. All other pages are blocked."),
                ("Attempt to navigate to the dashboard directly (e.g., via URL).", "System redirects back to the Change Password page. Dashboard is inaccessible."),
            ],
            "flow_color": NORMAL_FILL,
        },
        {
            "tab_name": "Alternate Flow",
            "tc_id": "TC-AUTH-02_AF",
            "tc_desc": "Login – Alternate Flows (Invalid Credentials, Deactivated, Locked, Google Cancelled)",
            "prerequisites": [
                "Account 'deactivated@example.com' is deactivated.",
                "Account 'locked@example.com' is locked (too many failed attempts).",
                "Account 'valid@example.com' exists with password 'Pass@1234'.",
            ],
            "test_data": [
                "A1 Email/Pass: 'valid@example.com' / 'WrongPass!'",
                "A2 Email: 'deactivated@example.com'",
                "A3 Email: 'locked@example.com'",
                "A4: Cancel Google OAuth consent screen",
            ],
            "test_scenario": "Verify system responses for all invalid or blocked login attempts.",
            "steps": [
                ("(A1) Enter 'valid@example.com' and wrong password 'WrongPass!'. Click Sign In.", "Generic error shown: 'Invalid email or password.' No details on which field is wrong. No session created."),
                ("(A2) Enter 'deactivated@example.com' with any password. Click Sign In.", "Message shown: 'Your account has been deactivated. Please contact support.' No session created."),
                ("(A3) Enter 'locked@example.com' with any password. Click Sign In.", "Lockout message displayed with estimated unlock time. Option to reset password to unlock immediately is shown."),
                ("(A4) Click 'Continue with Google', then cancel/deny on the Google consent screen.", "System returns to sign-in page with message that Google sign-in was not completed. No session created."),
            ],
            "flow_color": ALT_FILL,
        },
        {
            "tab_name": "Exception Flow",
            "tc_id": "TC-AUTH-02_EF",
            "tc_desc": "Login – Exception Flows (System Unavailable, Google Service Down)",
            "prerequisites": [
                "E1: System/database is temporarily unavailable.",
                "E2: Google OAuth service is unreachable.",
            ],
            "test_data": [
                "E1 Email: 'user@example.com' / 'Pass@1234'",
                "E2: Google Sign-In button clicked",
            ],
            "test_scenario": "Verify graceful error handling when the system or Google service is unavailable.",
            "steps": [
                ("(E1) Attempt email/password login while system is unavailable.", "Error message is shown. No session is created."),
                ("(E2) Click 'Continue with Google' when Google OAuth service is unreachable.", "Error message is shown. User is returned to the sign-in page. No session created."),
            ],
            "flow_color": EXCEP_FILL,
        },
    ]
    create_workbook("TC-AUTH-02_Login.xlsx", sheets)


def gen_auth03():
    """UC-AUTH-03: Logout"""
    sheets = [
        {
            "tab_name": "Normal Flow",
            "tc_id": "TC-AUTH-03_NF",
            "tc_desc": "Logout – Normal Flow (Successful Sign-Out)",
            "prerequisites": [
                "User is currently signed in.",
            ],
            "test_data": [
                "Logged-in user session active.",
            ],
            "test_scenario": "Verify that clicking Logout ends the session and redirects to the sign-in page, and that protected pages are inaccessible without re-authentication.",
            "steps": [
                ("While signed in, click the 'Logout' button in the navigation menu.", "Logout request is sent."),
                ("Observe the system response.", "Session is ended. 'Remember Me' tokens are cleared. User is redirected to the sign-in page."),
                ("Press the browser back button.", "Browser shows the sign-in page, not the previously viewed protected content."),
                ("Attempt to navigate directly to a protected URL (e.g., /dashboard).", "System redirects to sign-in page. Protected content is not accessible."),
            ],
            "flow_color": NORMAL_FILL,
        },
        {
            "tab_name": "Exception Flow",
            "tc_id": "TC-AUTH-03_EF",
            "tc_desc": "Logout – Exception Flow (Session Already Expired)",
            "prerequisites": [
                "User's session has already expired before clicking Logout.",
            ],
            "test_data": [
                "Expired session token.",
            ],
            "test_scenario": "Verify that clicking Logout when the session is already expired still redirects to the sign-in page without error.",
            "steps": [
                ("With an expired session, click 'Logout'.", "Logout request is sent."),
                ("Observe system response.", "User is redirected to the sign-in page. No error message is displayed."),
            ],
            "flow_color": EXCEP_FILL,
        },
    ]
    create_workbook("TC-AUTH-03_Logout.xlsx", sheets)


def gen_auth04():
    """UC-AUTH-04: Forgot Password (Reset via Email)"""
    sheets = [
        {
            "tab_name": "Normal Flow",
            "tc_id": "TC-AUTH-04_NF",
            "tc_desc": "Forgot Password – Normal Flow (Successful Reset)",
            "prerequisites": [
                "Account 'resetme@example.com' is registered.",
                "Email service is configured and working.",
            ],
            "test_data": [
                "Email: 'resetme@example.com'",
                "New Password: 'NewPass@123'",
                "Confirm New Password: 'NewPass@123'",
            ],
            "test_scenario": "Verify a registered user can reset their password via email link and the reset link is invalidated after use.",
            "steps": [
                ("On the sign-in page, click 'Forgot Password'.", "Forgot Password page is displayed."),
                ("Enter email 'resetme@example.com'. Click Submit.", "Generic message displayed: 'If an account exists with this email, a reset link has been sent.'"),
                ("Open the reset link received via email.", "Password reset form is displayed."),
                ("Enter New Password 'NewPass@123' and Confirm Password 'NewPass@123'. Click Save.", "New password is saved. Reset link is marked as used."),
                ("Observe redirect.", "User is redirected to sign-in page with a success message."),
                ("Attempt to use the same reset link again.", "Error: 'This reset link is invalid or has expired.' No change is made."),
                ("Sign in with new password 'NewPass@123'.", "User successfully signs in."),
            ],
            "flow_color": NORMAL_FILL,
        },
        {
            "tab_name": "Alternate Flow",
            "tc_id": "TC-AUTH-04_AF",
            "tc_desc": "Forgot Password – Alternate Flows (Email Not Found, Expired Link, Weak Password)",
            "prerequisites": [
                "A1: Email 'notfound@example.com' is NOT registered.",
                "A2: A reset link for 'resetme@example.com' has already expired (>60 min) or been used.",
                "A3: Email 'resetme@example.com' is registered.",
            ],
            "test_data": [
                "A1 Email: 'notfound@example.com'",
                "A2: Expired/used reset link URL",
                "A3 New Password: 'weak' (fails strength rules)",
            ],
            "test_scenario": "Verify system responses for alternate scenarios in the forgot password flow.",
            "steps": [
                ("(A1) Enter 'notfound@example.com'. Click Submit.", "Same generic message displayed: 'If an account exists with this email, a reset link has been sent.' No link is sent. Privacy is protected."),
                ("(A2) Open an expired or already-used reset link.", "Error displayed: 'This reset link is invalid or has expired.' User is prompted to request a new one."),
                ("(A3) Open a valid reset link. Enter New Password: 'weak'. Click Save.", "Password strength rules not met are shown. Submission is blocked. Flow returns to the reset form."),
            ],
            "flow_color": ALT_FILL,
        },
        {
            "tab_name": "Exception Flow",
            "tc_id": "TC-AUTH-04_EF",
            "tc_desc": "Forgot Password – Exception Flow (Email Delivery Failure)",
            "prerequisites": [
                "Account 'resetme@example.com' is registered.",
                "Email service is configured to fail/unreachable during test.",
            ],
            "test_data": [
                "Email: 'resetme@example.com'",
            ],
            "test_scenario": "Verify that an email delivery failure does not expose sensitive error details and the user can retry.",
            "steps": [
                ("Enter 'resetme@example.com'. Click Submit while email service is down.", "Request is submitted."),
                ("Observe system response.", "Generic message is displayed regardless. No internal error is exposed. A reset link record is created in the system. User can try again."),
            ],
            "flow_color": EXCEP_FILL,
        },
    ]
    create_workbook("TC-AUTH-04_ForgotPassword.xlsx", sheets)


def gen_auth05():
    """UC-AUTH-05: View / Update Profile"""
    sheets = [
        {
            "tab_name": "Normal Flow",
            "tc_id": "TC-AUTH-05_NF",
            "tc_desc": "View / Update Profile – Normal Flow (Successful Profile Update & Password Change)",
            "prerequisites": [
                "User is signed in.",
                "User profile exists with name, email, phone number, and profile photo.",
            ],
            "test_data": [
                "New Phone: '+60123456789'",
                "New Photo: valid JPEG < 2MB",
                "Current Password: 'Pass@1234'",
                "New Password: 'NewPass@5678'",
            ],
            "test_scenario": "Verify a signed-in user can view and update profile information and change their password.",
            "steps": [
                ("Navigate to the Profile page.", "Current profile data is displayed: name, email, phone, profile photo."),
                ("Update phone number to '+60123456789'. Upload a new valid JPEG photo. Click 'Save'.", "Validation passes. Updated profile is saved. Confirmation: 'Profile updated successfully.'"),
                ("Verify updated information is reflected system-wide.", "Phone and photo are updated in the UI."),
                ("Click 'Change Password'. Enter Current Password 'Pass@1234', New Password 'NewPass@5678', Confirm 'NewPass@5678'.", "Passwords are submitted for validation."),
                ("Observe system response.", "New password is saved. Confirmation: 'Password changed successfully.'"),
                ("Sign out and sign back in with 'NewPass@5678'.", "Login succeeds with new password."),
            ],
            "flow_color": NORMAL_FILL,
        },
        {
            "tab_name": "Alternate Flow",
            "tc_id": "TC-AUTH-05_AF",
            "tc_desc": "View / Update Profile – Alternate Flows (Invalid Phone, Invalid Photo, Password Errors)",
            "prerequisites": [
                "User is signed in.",
            ],
            "test_data": [
                "A1 Phone: 'abc12345' (invalid)",
                "A2 Photo: a .exe file (invalid type)",
                "A3 New Password: 'weak'",
                "A4 Current Password: 'WrongCurrent!'",
                "A5 New Password: 'Pass@1234', Confirm: 'Different@1'",
            ],
            "test_scenario": "Verify validation error handling for profile update and password change.",
            "steps": [
                ("(A1) Enter phone 'abc12345'. Click Save.", "Phone field is highlighted. Error: invalid phone number format."),
                ("(A2) Upload a .exe file as profile photo. Click Save.", "Error: 'Invalid file type. Please upload a valid image.'"),
                ("(A3) In Change Password, enter new password 'weak'. Submit.", "Failing password strength rules are displayed. Submission blocked."),
                ("(A4) In Change Password, enter wrong current password 'WrongCurrent!'. Submit.", "Error: 'Current password does not match.'"),
                ("(A5) Enter new password 'Pass@1234' and confirm 'Different@1'. Submit.", "Error: 'Passwords do not match.'"),
            ],
            "flow_color": ALT_FILL,
        },
        {
            "tab_name": "Exception Flow",
            "tc_id": "TC-AUTH-05_EF",
            "tc_desc": "View / Update Profile – Exception Flows (Upload Failure, System Error)",
            "prerequisites": [
                "User is signed in.",
                "A photo upload or system error is simulated.",
            ],
            "test_data": [
                "E1: Photo upload simulated to fail",
                "E2: System error simulated on Save",
            ],
            "test_scenario": "Verify that upload or system errors roll back changes and notify the user.",
            "steps": [
                ("(E1) Upload a valid photo. Simulate upload failure. Click Save.", "Photo upload fails. System rolls back the change. Error message notifies the user. Existing photo remains unchanged."),
                ("(E2) Make profile changes. Simulate a system error on Save.", "System rolls back the change. Error notification shown. No partial data is saved."),
            ],
            "flow_color": EXCEP_FILL,
        },
    ]
    create_workbook("TC-AUTH-05_ViewUpdateProfile.xlsx", sheets)


def gen_auth06():
    """UC-AUTH-06: Force Change Password"""
    sheets = [
        {
            "tab_name": "Normal Flow",
            "tc_id": "TC-AUTH-06_NF",
            "tc_desc": "Force Change Password – Normal Flow (Successful New Password Set)",
            "prerequisites": [
                "Admin has reset user 'user@example.com' password to temporary 'Temp@1234'.",
                "Forced password change flag is set on the account.",
            ],
            "test_data": [
                "Temp Password: 'Temp@1234'",
                "New Password: 'MyNewPass@99'",
                "Confirm Password: 'MyNewPass@99'",
            ],
            "test_scenario": "Verify that a user with a forced-change flag must change password before accessing the system, and the flag is cleared after.",
            "steps": [
                ("Sign in with 'user@example.com' / 'Temp@1234'.", "Login credentials verified. Forced change flag detected."),
                ("Observe redirect.", "User is redirected to 'Change Your Password' page. All other pages are blocked."),
                ("Enter new password 'MyNewPass@99' and confirm. Click 'Save'.", "New password is submitted."),
                ("Observe system response.", "New password is saved. Forced-change flag is cleared. User is redirected to dashboard."),
                ("Verify forced-change flag is cleared.", "User can now access all pages normally. Signing in again with 'MyNewPass@99' succeeds."),
            ],
            "flow_color": NORMAL_FILL,
        },
        {
            "tab_name": "Alternate Flow",
            "tc_id": "TC-AUTH-06_AF",
            "tc_desc": "Force Change Password – Alternate Flows (Same Temp Password, Weak Password)",
            "prerequisites": [
                "User is on the forced Change Password page.",
                "Temporary password is 'Temp@1234'.",
            ],
            "test_data": [
                "A1 New Password: 'Temp@1234' (same as temp)",
                "A2 New Password: 'weak'",
            ],
            "test_scenario": "Verify that using the same temporary password or a weak password is rejected.",
            "steps": [
                ("(A1) Enter 'Temp@1234' as new password. Click Save.", "Error: 'Please choose a different password.' Same temporary password is not accepted."),
                ("(A2) Enter 'weak' as new password. Click Save.", "Password strength rules not met are highlighted. Submission blocked."),
            ],
            "flow_color": ALT_FILL,
        },
        {
            "tab_name": "Exception Flow",
            "tc_id": "TC-AUTH-06_EF",
            "tc_desc": "Force Change Password – Exception Flow (Navigation Bypass Attempt)",
            "prerequisites": [
                "User has forced-change flag active and is on the Change Password page.",
            ],
            "test_data": [
                "Attempt to navigate to: /dashboard (direct URL)",
            ],
            "test_scenario": "Verify that a user with an active forced-change flag cannot bypass the change password page by navigating directly.",
            "steps": [
                ("While on Change Password page, attempt to navigate to /dashboard via URL bar.", "Navigation is attempted."),
                ("Observe system response.", "System intercepts the navigation and redirects user back to the Change Password page. Dashboard is not accessible."),
            ],
            "flow_color": EXCEP_FILL,
        },
    ]
    create_workbook("TC-AUTH-06_ForceChangePassword.xlsx", sheets)


# ══════════════════════════════════════════════════════════════
# MODULE 2 – HOMESTAY MANAGEMENT
# ══════════════════════════════════════════════════════════════

def gen_hs01():
    sheets = [
        {
            "tab_name": "Normal Flow",
            "tc_id": "TC-HS-01_NF",
            "tc_desc": "Browse Homestay Units – Normal Flow (Units Available)",
            "prerequisites": ["At least two active homestay units exist in the system.", "Guest may or may not be signed in."],
            "test_data": ["Guest user (signed out)", "At least 2 active units in DB"],
            "test_scenario": "Verify the guest can view all active homestay unit cards on the listing page and navigate to a unit's detail page.",
            "steps": [
                ("Open the homestay listing page.", "All active units are displayed as cards showing: name, main photo, price per night, location, average rating."),
                ("Scroll through the listing.", "All active units are visible. Inactive units are NOT shown."),
                ("Click on any unit card.", "User is navigated to that unit's detail page (UC-HS-02)."),
            ],
            "flow_color": NORMAL_FILL,
        },
        {
            "tab_name": "Alternate Flow",
            "tc_id": "TC-HS-01_AF",
            "tc_desc": "Browse Homestay Units – Alternate Flow (No Active Units)",
            "prerequisites": ["No active homestay units exist in the system."],
            "test_data": ["DB: all units are inactive or no units exist"],
            "test_scenario": "Verify the system shows an informational message when no active units are available.",
            "steps": [
                ("Open the homestay listing page.", "Page loads."),
                ("Observe the displayed content.", "Informational message shown: 'No homestay units are currently available.' No cards are displayed."),
            ],
            "flow_color": ALT_FILL,
        },
    ]
    create_workbook("TC-HS-01_BrowseHomestayUnits.xlsx", sheets)


def gen_hs02():
    sheets = [
        {
            "tab_name": "Normal Flow",
            "tc_id": "TC-HS-02_NF",
            "tc_desc": "View Unit Details & Availability – Normal Flow",
            "prerequisites": ["An active unit 'Cozy Cottage' exists with photos, description, pricing, and booking records."],
            "test_data": ["Unit: 'Cozy Cottage'", "Has available, booked, and blocked dates"],
            "test_scenario": "Verify the unit detail page shows complete information including availability calendar with colour-coded dates.",
            "steps": [
                ("Click on 'Cozy Cottage' unit card from the listing page.", "Unit detail page is loaded."),
                ("Review the displayed information.", "Displayed: unit name, photo gallery, full description, location, base price/night, deposit, check-in/out times, house rules, average rating, guest reviews."),
                ("Observe the availability calendar.", "Calendar shows colour-coded dates: Available (bookable), Booked (reserved), Temporarily Held, Blocked (admin-set)."),
                ("Click 'Book Now' button.", "User is navigated to the booking form (UC-BK-01)."),
            ],
            "flow_color": NORMAL_FILL,
        },
    ]
    create_workbook("TC-HS-02_ViewUnitDetails.xlsx", sheets)


def gen_hs03():
    sheets = [
        {
            "tab_name": "Normal Flow",
            "tc_id": "TC-HS-03_NF",
            "tc_desc": "Create Homestay Unit – Normal Flow (Successful Creation)",
            "prerequisites": ["Admin is signed in with permission to create homestay units.", "Default house rules (No Pets, No Durians, No Smoking) are configured."],
            "test_data": ["Name: 'Sunset Villa'", "Description: 'A beautiful villa...'", "Location: 'Langkawi'", "Price/Night: RM 250", "Deposit: RM 100", "Check-in: 2:00 PM", "Check-out: 12:00 PM", "Photo: valid JPEG"],
            "test_scenario": "Verify an admin can create a new homestay unit with all required details, and default house rules are automatically applied.",
            "steps": [
                ("Navigate to Homestay Management → Create New Unit.", "Unit creation form is displayed."),
                ("Fill in all required fields: name, description, location, price, deposit, check-in/out times.", "All fields accept input."),
                ("Upload at least one valid JPEG photo.", "Photo is accepted."),
                ("Optionally set a custom extension payment window.", "Field accepts the value."),
                ("Click 'Save'.", "Form is submitted."),
                ("Observe system response.", "Success message shown. Unit is created with 'active' status. Unit appears on guest listing page."),
                ("Verify default house rules.", "Unit's house rules list includes: No Pets, No Durians, No Smoking (all active defaults)."),
                ("Verify audit log.", "An audit log entry is recorded for the unit creation."),
            ],
            "flow_color": NORMAL_FILL,
        },
        {
            "tab_name": "Alternate Flow",
            "tc_id": "TC-HS-03_AF",
            "tc_desc": "Create Homestay Unit – Alternate Flows (Missing Fields, Invalid Photo)",
            "prerequisites": ["Admin is signed in with create permission."],
            "test_data": ["A1: Leave 'Name' field empty", "A2: Upload a .exe file as photo"],
            "test_scenario": "Verify validation errors are shown for missing required fields or invalid photo types.",
            "steps": [
                ("(A1) Fill all fields except Name. Click Save.", "Missing Name field is highlighted. Error: 'This field is required.' No unit is created."),
                ("(A2) Upload a .exe file as photo. Click Save.", "File is rejected. Error: 'Invalid file format. Please upload a valid image.' No unit is created."),
            ],
            "flow_color": ALT_FILL,
        },
        {
            "tab_name": "Exception Flow",
            "tc_id": "TC-HS-03_EF",
            "tc_desc": "Create Homestay Unit – Exception Flow (Photo Upload Failure)",
            "prerequisites": ["Admin is signed in. All fields are filled with valid data. Photo upload is simulated to fail."],
            "test_data": ["All valid unit fields", "Photo upload: simulated failure"],
            "test_scenario": "Verify that a photo upload failure does not prevent unit creation; unit is saved with a warning about the failed upload.",
            "steps": [
                ("Fill all required fields and attempt to upload a photo. Simulate upload failure. Click Save.", "Form submitted."),
                ("Observe system response.", "Unit record is saved successfully. Warning shown: 'Unit created, but photo upload failed. You can upload photos by editing the unit.' Unit is visible on the listing page (without photo)."),
            ],
            "flow_color": EXCEP_FILL,
        },
    ]
    create_workbook("TC-HS-03_CreateHomestayUnit.xlsx", sheets)


def gen_hs04():
    sheets = [
        {
            "tab_name": "Normal Flow",
            "tc_id": "TC-HS-04_NF",
            "tc_desc": "Edit Homestay Unit – Normal Flow (Successful Edit)",
            "prerequisites": ["Unit 'Sunset Villa' exists. Admin is signed in with edit permission."],
            "test_data": ["Updated Name: 'Sunset Villa Premium'", "Updated Price: RM 300/night", "New photo: valid JPEG"],
            "test_scenario": "Verify an admin can edit an existing unit's details and changes are reflected immediately.",
            "steps": [
                ("Navigate to Homestay Management → select 'Sunset Villa' → click 'Edit'.", "Pre-filled edit form is displayed with current values."),
                ("Change Name to 'Sunset Villa Premium' and Price to RM 300. Upload a new photo.", "Fields accept changes."),
                ("Click 'Save'.", "Changes are submitted."),
                ("Observe system response.", "Success message shown. Updated values are live on the guest listing page immediately."),
                ("Verify audit log.", "Audit log entry recorded for the edit."),
            ],
            "flow_color": NORMAL_FILL,
        },
        {
            "tab_name": "Alternate Flow",
            "tc_id": "TC-HS-04_AF",
            "tc_desc": "Edit Homestay Unit – Alternate Flow (Invalid Input)",
            "prerequisites": ["Unit exists. Admin is signed in."],
            "test_data": ["Price: '-50' (negative, invalid)"],
            "test_scenario": "Verify invalid input is rejected and previously entered values are preserved.",
            "steps": [
                ("Open unit edit form. Enter price '-50'. Click Save.", "Form submitted."),
                ("Observe system response.", "Problem fields are highlighted. Error is shown. Previously entered values are preserved (not cleared). Record is NOT saved."),
            ],
            "flow_color": ALT_FILL,
        },
    ]
    create_workbook("TC-HS-04_EditHomestayUnit.xlsx", sheets)


def gen_hs05():
    sheets = [
        {
            "tab_name": "Normal Flow",
            "tc_id": "TC-HS-05_NF",
            "tc_desc": "Deactivate / Delete Unit – Normal Flow (No Conflicts)",
            "prerequisites": ["Unit 'Sunset Villa' exists with no upcoming confirmed bookings.", "Admin is signed in."],
            "test_data": ["Unit: 'Sunset Villa'"],
            "test_scenario": "Verify admin can deactivate a unit with no future bookings, and it is removed from the guest listing.",
            "steps": [
                ("Navigate to Homestay Management → open 'Sunset Villa' → click 'Deactivate'.", "System checks for upcoming confirmed bookings."),
                ("System shows confirmation dialog.", "Confirmation: 'Are you sure you want to deactivate this unit?'"),
                ("Click 'Confirm'.", "Deactivation is submitted."),
                ("Observe system response.", "Success message shown. Unit status is 'inactive'. Unit no longer appears on guest listing page. Unit data is retained."),
                ("Verify audit log.", "Audit log entry recorded."),
            ],
            "flow_color": NORMAL_FILL,
        },
        {
            "tab_name": "Alternate Flow",
            "tc_id": "TC-HS-05_AF",
            "tc_desc": "Deactivate / Delete Unit – Alternate Flows (Has Bookings, Admin Cancels)",
            "prerequisites": ["Unit 'Cottage A' has at least one upcoming confirmed booking."],
            "test_data": ["A1: Unit with future confirmed booking", "A2: Admin clicks Cancel on confirmation dialog"],
            "test_scenario": "Verify deactivation is blocked when future confirmed bookings exist, and confirm that cancelling the dialog makes no changes.",
            "steps": [
                ("(A1) Attempt to deactivate 'Cottage A' which has an upcoming confirmed booking.", "Action is blocked. Warning shown listing the conflicting bookings. Admin must cancel or reassign those bookings first."),
                ("(A2) On a unit with no conflicts, click 'Deactivate'. In the confirmation dialog, click 'Cancel'.", "Dialog is dismissed. No changes are made. Unit remains active."),
            ],
            "flow_color": ALT_FILL,
        },
        {
            "tab_name": "Exception Flow",
            "tc_id": "TC-HS-05_EF",
            "tc_desc": "Deactivate / Delete Unit – Exception Flow (Cancel Confirmation Prompt)",
            "prerequisites": ["Unit exists. No upcoming bookings."],
            "test_data": ["Admin clicks Cancel on confirmation dialog"],
            "test_scenario": "Verify no changes are made when admin cancels the confirmation dialog.",
            "steps": [
                ("Click 'Deactivate'. When confirmation dialog appears, click 'Cancel'.", "Dialog closes. No changes are made. Unit remains active."),
            ],
            "flow_color": EXCEP_FILL,
        },
    ]
    create_workbook("TC-HS-05_DeactivateDeleteUnit.xlsx", sheets)


def gen_hs06():
    sheets = [
        {
            "tab_name": "Normal Flow",
            "tc_id": "TC-HS-06_NF",
            "tc_desc": "View All Units List – Normal Flow (Admin Overview)",
            "prerequisites": ["Admin is signed in with homestay management access.", "Multiple units (active and inactive) exist."],
            "test_data": ["Filter: Status = 'Active'"],
            "test_scenario": "Verify admin can view all units including inactive ones, and apply status filter.",
            "steps": [
                ("Navigate to Homestay Management section.", "System retrieves all unit records (active and inactive)."),
                ("Observe displayed list.", "List/table shows each unit's: name, status (active/inactive), number of upcoming confirmed bookings, base price, action buttons (Edit, Deactivate, Manage Policies)."),
                ("Apply status filter 'Active'.", "List updates to show only active units."),
                ("Clear the filter.", "Full list (all units) is restored."),
            ],
            "flow_color": NORMAL_FILL,
        },
    ]
    create_workbook("TC-HS-06_ViewAllUnitsList.xlsx", sheets)


# ══════════════════════════════════════════════════════════════
# MODULE 3 – BOOKING
# ══════════════════════════════════════════════════════════════

def gen_bk01():
    sheets = [
        {
            "tab_name": "Normal Flow",
            "tc_id": "TC-BK-01_NF",
            "tc_desc": "View Availability & Select Dates – Normal Flow (Valid Selection)",
            "prerequisites": ["Guest is on the unit detail/booking page.", "Dates 2026-07-10 to 2026-07-13 are available."],
            "test_data": ["Check-in: 2026-07-10", "Check-out: 2026-07-13"],
            "test_scenario": "Verify guest can view colour-coded availability calendar and select available dates, with booking summary updated.",
            "steps": [
                ("Open the unit detail/booking form.", "Colour-coded availability calendar is displayed (Available, Booked, Temporarily Held, Blocked)."),
                ("Click check-in date: 2026-07-10.", "Check-in date is highlighted."),
                ("Click check-out date: 2026-07-13.", "System performs real-time availability check."),
                ("Observe result.", "All dates confirmed available. Booking summary updates: 3 nights, estimated cost calculated. Guest can proceed to submit booking."),
            ],
            "flow_color": NORMAL_FILL,
        },
        {
            "tab_name": "Alternate Flow",
            "tc_id": "TC-BK-01_AF",
            "tc_desc": "View Availability & Select Dates – Alternate Flows (Unavailable, Invalid Range, Min Stay)",
            "prerequisites": ["Date 2026-07-11 is already booked.", "Unit has a minimum 2-night stay requirement."],
            "test_data": ["A1: Check-in 2026-07-10, Check-out 2026-07-13 (2026-07-11 is booked)", "A2: Check-in 2026-07-13, Check-out 2026-07-10 (invalid)", "A3: Check-in 2026-07-10, Check-out 2026-07-10 (0 nights, < minimum)"],
            "test_scenario": "Verify conflict detection, invalid date range, and minimum stay validation.",
            "steps": [
                ("(A1) Select check-in 2026-07-10, check-out 2026-07-13 (2026-07-11 is booked).", "Conflict highlighted. Message: 'Selected dates are not available.' Guest must re-select."),
                ("(A2) Select check-out 2026-07-10 BEFORE check-in 2026-07-13.", "Validation message shown: 'Check-out must be after check-in.' Guest must re-select."),
                ("(A3) Select same-day check-in and check-out (0 nights).", "Minimum stay requirement message displayed. Booking summary not updated."),
            ],
            "flow_color": ALT_FILL,
        },
    ]
    create_workbook("TC-BK-01_ViewAvailabilitySelectDates.xlsx", sheets)


def gen_bk02():
    sheets = [
        {
            "tab_name": "Normal Flow",
            "tc_id": "TC-BK-02_NF",
            "tc_desc": "Submit Booking – Normal Flow (Successful Booking Creation)",
            "prerequisites": ["Guest is signed in.", "Dates 2026-07-10 to 2026-07-13 are confirmed available.", "Guest is on booking summary page."],
            "test_data": ["Unit: 'Cozy Cottage'", "Check-in: 2026-07-10", "Check-out: 2026-07-13", "Total: RM 850"],
            "test_scenario": "Verify a guest can confirm a booking, record is created in 'awaiting payment' status, bill is auto-generated, and notifications are sent.",
            "steps": [
                ("Review booking summary (unit, dates, total cost). Click 'Confirm Booking'.", "System double-checks date availability."),
                ("Observe system response.", "Booking record created with status 'awaiting payment' and 1-day payment deadline. Bill generated with unique bill number. Dates are temporarily reserved."),
                ("Verify notifications.", "In-app and email notification sent to guest with bill and payment deadline."),
                ("Observe redirect.", "Guest is directed to the payment page."),
            ],
            "flow_color": NORMAL_FILL,
        },
        {
            "tab_name": "Alternate Flow",
            "tc_id": "TC-BK-02_AF",
            "tc_desc": "Submit Booking – Alternate Flow (Race Condition – Dates Became Unavailable)",
            "prerequisites": ["Another guest booked the same dates between the guest's selection and submission."],
            "test_data": ["Selected dates: 2026-07-10 to 2026-07-13 (now booked by another guest)"],
            "test_scenario": "Verify the system detects a race condition when dates become unavailable between selection and confirmation.",
            "steps": [
                ("Click 'Confirm Booking' while selected dates were just taken by another user.", "System re-checks availability."),
                ("Observe system response.", "Conflict message shown: 'These dates are no longer available. Please select new dates.' No booking record is created."),
            ],
            "flow_color": ALT_FILL,
        },
        {
            "tab_name": "Exception Flow",
            "tc_id": "TC-BK-02_EF",
            "tc_desc": "Submit Booking – Exception Flow (System Error During Creation)",
            "prerequisites": ["Guest is signed in. Valid dates selected. System error simulated."],
            "test_data": ["Simulated system/DB error during booking creation"],
            "test_scenario": "Verify no partial booking record is created on system error.",
            "steps": [
                ("Click 'Confirm Booking' while a system error is simulated.", "Submission attempted."),
                ("Observe system response.", "Error message shown. No booking record is created. Guest is prompted to try again."),
            ],
            "flow_color": EXCEP_FILL,
        },
    ]
    create_workbook("TC-BK-02_SubmitBooking.xlsx", sheets)


def gen_bk03():
    sheets = [
        {
            "tab_name": "Normal Flow – Guest",
            "tc_id": "TC-BK-03_NF_Guest",
            "tc_desc": "View Bookings – Normal Flow (Guest View)",
            "prerequisites": ["Guest is signed in and has at least one current and one past booking."],
            "test_data": ["Guest with 2+ bookings (one 'confirmed', one 'completed')"],
            "test_scenario": "Verify a guest can view their own bookings separated into Current and History tabs.",
            "steps": [
                ("Navigate to 'My Bookings'.", "System retrieves all bookings for the guest."),
                ("Observe displayed tabs.", "Bookings shown in 'Current' (active/upcoming) and 'History' (completed/cancelled) tabs."),
                ("Check each booking card.", "Displays: unit name, check-in/out dates, status, total cost."),
                ("Click on a booking.", "Navigated to booking detail page (UC-BK-04)."),
                ("On a completed booking, observe links.", "'View Receipt' and 'Leave a Review' (if not yet submitted) links are shown."),
            ],
            "flow_color": NORMAL_FILL,
        },
        {
            "tab_name": "Normal Flow – Admin",
            "tc_id": "TC-BK-03_NF_Admin",
            "tc_desc": "View Bookings – Normal Flow (Admin View – All Bookings)",
            "prerequisites": ["Admin is signed in with booking management access.", "Multiple bookings exist across different units."],
            "test_data": ["Filter by status: 'confirmed'", "Filter by date range: July 2026"],
            "test_scenario": "Verify admin can view all system-wide bookings with filters and calendar view.",
            "steps": [
                ("Navigate to Booking Management.", "System retrieves all bookings across all units."),
                ("Observe the displayed list.", "Bookings displayed with filters available: status, date range, unit, booking reference."),
                ("Apply filter: status 'confirmed'.", "List updates to show only confirmed bookings."),
                ("Open the booking calendar view.", "All-unit calendar is displayed showing all reservations."),
                ("Click on a booking entry.", "Navigated to booking detail or action options."),
            ],
            "flow_color": NORMAL_FILL,
        },
    ]
    create_workbook("TC-BK-03_ViewBookings.xlsx", sheets)


def gen_bk04():
    sheets = [
        {
            "tab_name": "Normal Flow",
            "tc_id": "TC-BK-04_NF",
            "tc_desc": "View Booking Details – Normal Flow",
            "prerequisites": ["User is signed in.", "Booking BK-2026-001 exists with confirmed status and associated QR code."],
            "test_data": ["Booking reference: BK-2026-001"],
            "test_scenario": "Verify all booking details are displayed including QR code for confirmed bookings, and bill/receipt download is available.",
            "steps": [
                ("Click on booking BK-2026-001 from the bookings list.", "System retrieves full booking record."),
                ("Observe displayed information.", "Shown: unit name and photo, check-in/out date and time, total amount, payment status, booking status, cancellation policy with estimated refund, QR code (since booking is confirmed), extension history (if any)."),
                ("Click 'Download Bill'.", "Bill document is downloaded."),
            ],
            "flow_color": NORMAL_FILL,
        },
    ]
    create_workbook("TC-BK-04_ViewBookingDetails.xlsx", sheets)


def gen_bk05():
    sheets = [
        {
            "tab_name": "Normal Flow",
            "tc_id": "TC-BK-05_NF",
            "tc_desc": "Cancel Booking – Normal Flow (Confirmed Booking with Refund)",
            "prerequisites": ["Booking BK-2026-001 is in 'confirmed' status. Check-in is 20 days away (>14 days → 100% refund tier)."],
            "test_data": ["Booking: BK-2026-001", "Refund tier: >14 days = 100%"],
            "test_scenario": "Verify guest can cancel a confirmed booking, refund is calculated and processed, dates are released, and notifications are sent.",
            "steps": [
                ("Open booking BK-2026-001. Click 'Cancel Booking'.", "System calculates refund based on cancellation policy tiers."),
                ("Observe confirmation dialog.", "Dialog shown: 'You will receive a refund of [100% amount]. This cannot be undone.'"),
                ("Click 'Confirm Cancellation'.", "Cancellation is submitted."),
                ("Observe system response.", "Booking status set to 'cancelled'. Reserved dates released. Refund is initiated through payment service. In-app and email notification sent to guest and admin. Audit log entry recorded."),
            ],
            "flow_color": NORMAL_FILL,
        },
        {
            "tab_name": "Alternate Flow",
            "tc_id": "TC-BK-05_AF",
            "tc_desc": "Cancel Booking – Alternate Flows (User Cancels Dialog, Awaiting Payment)",
            "prerequisites": ["Booking BK-2026-002 is 'awaiting payment' (no payment made). Booking BK-2026-003 confirmed."],
            "test_data": ["A1: User clicks Cancel in confirmation dialog", "A2: Booking in 'awaiting payment' status"],
            "test_scenario": "Verify cancelling the confirmation dialog makes no changes, and 'awaiting payment' bookings have no refund.",
            "steps": [
                ("(A1) Click 'Cancel Booking', then in dialog click 'Cancel'.", "Dialog dismissed. No changes made. Booking remains in original status."),
                ("(A2) Cancel a booking in 'awaiting payment' status.", "No refund is processed (no payment was made). Booking is simply cancelled. Dates are released. Notification sent."),
            ],
            "flow_color": ALT_FILL,
        },
        {
            "tab_name": "Exception Flow",
            "tc_id": "TC-BK-05_EF",
            "tc_desc": "Cancel Booking – Exception Flow (Refund Processing Failure)",
            "prerequisites": ["Payment gateway refund is simulated to fail."],
            "test_data": ["Simulated refund gateway failure"],
            "test_scenario": "Verify that a refund failure still completes the cancellation and flags it for manual admin processing.",
            "steps": [
                ("Confirm cancellation while payment gateway refund fails.", "Cancellation proceeds."),
                ("Observe system response.", "Booking is cancelled. Refund flagged for manual processing by admin. Admin is notified. Guest receives cancellation notification (without confirming refund amount yet)."),
            ],
            "flow_color": EXCEP_FILL,
        },
    ]
    create_workbook("TC-BK-05_CancelBooking.xlsx", sheets)


def gen_bk06():
    sheets = [
        {
            "tab_name": "Normal Flow – Create",
            "tc_id": "TC-BK-06_NF_Create",
            "tc_desc": "Manage Booking (Admin) – Normal Flow: Create Booking on Behalf of Guest",
            "prerequisites": ["Admin signed in with booking management permission.", "Guest account 'guest@example.com' exists.", "Dates 2026-08-01 to 2026-08-05 are available."],
            "test_data": ["Guest: 'guest@example.com'", "Unit: 'Cozy Cottage'", "Check-in: 2026-08-01", "Check-out: 2026-08-05"],
            "test_scenario": "Verify admin can create a booking on behalf of a guest.",
            "steps": [
                ("Navigate to Bookings → Create Booking.", "Create booking form displayed."),
                ("Select guest 'guest@example.com', unit 'Cozy Cottage', dates 2026-08-01 to 2026-08-05.", "Real-time availability check performed."),
                ("Observe booking summary.", "Total cost is displayed."),
                ("Click Confirm.", "Booking created with status 'awaiting payment'. Bill generated. Guest notified to pay."),
            ],
            "flow_color": NORMAL_FILL,
        },
        {
            "tab_name": "Normal Flow – Edit",
            "tc_id": "TC-BK-06_NF_Edit",
            "tc_desc": "Manage Booking (Admin) – Normal Flow: Edit Booking",
            "prerequisites": ["Booking BK-2026-005 exists. Admin signed in."],
            "test_data": ["New Check-out: 2026-08-07"],
            "test_scenario": "Verify admin can edit booking details including dates.",
            "steps": [
                ("Open booking BK-2026-005. Click 'Edit'.", "Edit form displayed with current booking details."),
                ("Change check-out date to 2026-08-07. Click Submit.", "System checks availability for new dates."),
                ("Observe system response.", "Updated record saved. Audit log entry recorded. Guest notified of the change."),
            ],
            "flow_color": NORMAL_FILL,
        },
        {
            "tab_name": "Normal Flow – Delete & Block",
            "tc_id": "TC-BK-06_NF_Delete",
            "tc_desc": "Manage Booking (Admin) – Normal Flow: Delete Booking & Block Dates",
            "prerequisites": ["Booking BK-2026-006 exists. Admin signed in."],
            "test_data": ["Delete: BK-2026-006", "Block: Unit 'Cozy Cottage', 2026-09-10 to 2026-09-15"],
            "test_scenario": "Verify admin can permanently delete a booking and block specific dates.",
            "steps": [
                ("Select BK-2026-006. Click 'Delete'.", "Warning: 'This is permanent and cannot be undone.' Confirmation required."),
                ("Confirm deletion.", "Booking deleted. Dates released. Guest notified. Audit log recorded."),
                ("Select unit 'Cozy Cottage'. Set date range 2026-09-10–2026-09-15. Click 'Block Dates'. Add optional internal note.", "Blocked dates stored. They immediately appear as 'unavailable/blocked' on guest-facing calendar. Internal note stored (not shown to guests)."),
            ],
            "flow_color": NORMAL_FILL,
        },
        {
            "tab_name": "Alternate Flow",
            "tc_id": "TC-BK-06_AF",
            "tc_desc": "Manage Booking (Admin) – Alternate Flows (Date Conflict, Cancel Delete, Block Conflict)",
            "prerequisites": ["Dates 2026-08-06 are booked when admin tries to edit/create to that date."],
            "test_data": ["A1: Conflicting new dates during edit", "A2: Admin clicks Cancel on delete confirmation", "A3: Blocked dates overlap with confirmed booking"],
            "test_scenario": "Verify conflict and cancel handling for admin booking management.",
            "steps": [
                ("(A1) Edit booking to dates that are already taken.", "System displays conflict. Admin prompted to select different dates."),
                ("(A2) Click Delete on a booking, then Cancel on the warning dialog.", "Dialog dismissed. No changes made. Booking record intact."),
                ("(A3) Attempt to block dates that have a confirmed booking.", "System warns admin of the conflict. Admin must cancel existing bookings first before blocking."),
            ],
            "flow_color": ALT_FILL,
        },
    ]
    create_workbook("TC-BK-06_ManageBookingAdmin.xlsx", sheets)


def gen_bk07():
    sheets = [
        {
            "tab_name": "Normal Flow",
            "tc_id": "TC-BK-07_NF",
            "tc_desc": "Auto-Cancel Expired Booking – Normal Flow (Scheduled Job)",
            "prerequisites": ["Booking BK-2026-010 is in 'awaiting payment' status. Payment deadline has passed."],
            "test_data": ["Booking BK-2026-010: awaiting payment, deadline passed > 1 hour ago"],
            "test_scenario": "Verify the scheduled job auto-cancels overdue 'awaiting payment' bookings, releases dates, and notifies guests.",
            "steps": [
                ("Trigger or wait for the scheduled job to run (~hourly).", "Scheduled job executes."),
                ("Verify system finds overdue bookings.", "System queries for bookings with status 'awaiting payment' and payment deadline < now."),
                ("Observe result for BK-2026-010.", "Booking status set to 'cancelled'. Reserved dates released."),
                ("Verify notification.", "In-app and email cancellation notification sent to guest."),
                ("Verify audit log.", "Audit log entry recorded: 'System auto-cancelled booking BK-2026-010'."),
            ],
            "flow_color": NORMAL_FILL,
        },
    ]
    create_workbook("TC-BK-07_AutoCancelExpiredBooking.xlsx", sheets)


# ══════════════════════════════════════════════════════════════
# MODULE 4 – PAYMENT
# ══════════════════════════════════════════════════════════════

def gen_pay01():
    sheets = [
        {
            "tab_name": "Normal Flow",
            "tc_id": "TC-PAY-01_NF",
            "tc_desc": "Make Payment – Normal Flow (Successful Payment)",
            "prerequisites": ["Booking BK-2026-001 is 'awaiting payment'. Bill generated. Payment deadline has not passed."],
            "test_data": ["Booking: BK-2026-001", "Payment method: test credit card 4242 4242 4242 4242"],
            "test_scenario": "Verify a guest can pay a bill, booking becomes 'confirmed', QR code is generated, receipt created, and notifications sent.",
            "steps": [
                ("Click 'Pay Now' from booking detail or bill notification.", "System creates payment request with gateway. Guest redirected to secure payment page."),
                ("Complete payment with test card on the payment gateway page.", "Payment submitted to gateway."),
                ("Gateway sends payment confirmation to HomeLodge.", "System receives confirmation."),
                ("System verifies confirmation is genuine.", "Verification passes."),
                ("Observe system actions.", "Payment recorded as successful. Booking status updated to 'confirmed'. QR access code generated (valid check-in to check-out). Payment receipt generated."),
                ("Verify notifications.", "Guest receives in-app + email booking confirmation with receipt and QR code. Admin notified of new confirmed booking."),
            ],
            "flow_color": NORMAL_FILL,
        },
        {
            "tab_name": "Alternate Flow",
            "tc_id": "TC-PAY-01_AF",
            "tc_desc": "Make Payment – Alternate Flows (Payment Declined, Guest Abandons)",
            "prerequisites": ["Booking BK-2026-001 is 'awaiting payment'."],
            "test_data": ["A1: Declined card 4000 0000 0000 0002", "A2: Guest closes payment page without completing"],
            "test_scenario": "Verify system handles payment failure and abandoned payment gracefully.",
            "steps": [
                ("(A1) Enter a declined card on the payment page.", "Payment gateway rejects the card. Failure notification sent to HomeLodge. Payment recorded as 'failed'. Guest returned to HomeLodge with error message and 'Try Again' option. Booking remains 'awaiting payment'."),
                ("(A2) Guest is redirected to the payment page but closes the browser/tab.", "No confirmation received. Booking remains in 'awaiting payment' status until the deadline. No payment record for this attempt."),
            ],
            "flow_color": ALT_FILL,
        },
        {
            "tab_name": "Exception Flow",
            "tc_id": "TC-PAY-01_EF",
            "tc_desc": "Make Payment – Exception Flows (Duplicate Confirmation, Unverifiable Confirmation)",
            "prerequisites": ["A valid payment confirmation has already been processed for BK-2026-001."],
            "test_data": ["E1: Same payment confirmation sent twice", "E2: Forged/invalid confirmation payload"],
            "test_scenario": "Verify system handles duplicate and unverifiable payment confirmations securely.",
            "steps": [
                ("(E1) Simulate duplicate payment confirmation webhook.", "System detects duplicate. Second confirmation is ignored without re-processing. Booking remains 'confirmed' (not double-confirmed)."),
                ("(E2) Simulate a forged/unverifiable confirmation webhook.", "System rejects the confirmation. Security alert is recorded in the system. Booking status is NOT changed. No QR code or receipt is generated."),
            ],
            "flow_color": EXCEP_FILL,
        },
    ]
    create_workbook("TC-PAY-01_MakePayment.xlsx", sheets)


def gen_pay02():
    sheets = [
        {
            "tab_name": "Normal Flow – Guest",
            "tc_id": "TC-PAY-02_NF_Guest",
            "tc_desc": "View Payment & Billing Records – Normal Flow (Guest)",
            "prerequisites": ["Guest is signed in and has payment and bill records."],
            "test_data": ["Guest with at least 1 paid booking"],
            "test_scenario": "Verify guest can view their payment and billing records and download documents.",
            "steps": [
                ("Navigate to My Bookings → booking detail or Payment History.", "System retrieves all payment records for the guest."),
                ("Observe displayed records.", "Each payment shown: payment number, booking reference, date, amount, status."),
                ("Click on a bill to view.", "Itemised bill displayed: nightly rate × nights, deposit, total, payment deadline."),
                ("Click 'Download Bill'.", "Bill PDF is downloaded."),
                ("For a completed payment, click 'View Receipt'.", "Receipt displayed. 'Download Receipt' available."),
            ],
            "flow_color": NORMAL_FILL,
        },
        {
            "tab_name": "Normal Flow – Admin",
            "tc_id": "TC-PAY-02_NF_Admin",
            "tc_desc": "View Payment & Billing Records – Normal Flow (Admin)",
            "prerequisites": ["Admin is signed in with payment management access."],
            "test_data": ["Filter: status = 'successful'", "Filter: date range = July 2026"],
            "test_scenario": "Verify admin can view all billing and payment records system-wide with filters.",
            "steps": [
                ("Navigate to Payment Management.", "System retrieves all billing and payment records."),
                ("Observe billing list.", "Each entry: reference number, guest name, unit, amount, status, date."),
                ("Apply filter: date range July 2026, status 'successful'.", "List updates to show matching records only."),
                ("Click an entry.", "Detail view displayed. Admin can take action."),
            ],
            "flow_color": NORMAL_FILL,
        },
    ]
    create_workbook("TC-PAY-02_ViewPaymentBillingRecords.xlsx", sheets)


def gen_pay03():
    sheets = [
        {
            "tab_name": "Normal Flow",
            "tc_id": "TC-PAY-03_NF",
            "tc_desc": "Regenerate Bill / Receipt – Normal Flow (Admin)",
            "prerequisites": ["Booking BK-2026-001 has corresponding billing and payment records.", "Admin is signed in with appropriate permission."],
            "test_data": ["Booking: BK-2026-001"],
            "test_scenario": "Verify admin can regenerate a bill or receipt, which is available for download and optionally resent to the guest.",
            "steps": [
                ("Navigate to billing or payment detail page for BK-2026-001.", "Detail page loaded."),
                ("Click 'Regenerate Bill'.", "System fetches latest booking and payment data."),
                ("Observe system response.", "Fresh bill document generated. Document is available for download. Option to resend to guest via email."),
                ("Click 'Regenerate Receipt'.", "Fresh receipt document generated. Available for download."),
            ],
            "flow_color": NORMAL_FILL,
        },
    ]
    create_workbook("TC-PAY-03_RegenerateBillReceipt.xlsx", sheets)


# ══════════════════════════════════════════════════════════════
# MODULE 5 – NOTIFICATION
# ══════════════════════════════════════════════════════════════

def gen_notif01():
    sheets = [
        {
            "tab_name": "Normal Flow – In-App",
            "tc_id": "TC-NOTIF-01_NF_InApp",
            "tc_desc": "Receive System Notification – Normal Flow (In-App Notification)",
            "prerequisites": ["User is signed in and online.", "A booking confirmation event occurs for the user."],
            "test_data": ["Triggering event: booking confirmed for BK-2026-001"],
            "test_scenario": "Verify the in-app notification bell updates in real time and the notification navigates to the relevant page.",
            "steps": [
                ("Trigger a booking confirmation (payment made for BK-2026-001).", "System creates a notification record for the guest."),
                ("Observe bell icon (user is online).", "Bell icon badge updates in real time without page refresh."),
                ("Click the bell icon.", "Notifications panel opens showing list of notifications (newest first, read and unread)."),
                ("Click the booking confirmation notification.", "Notification marked as read. User navigated to booking detail page for BK-2026-001."),
            ],
            "flow_color": NORMAL_FILL,
        },
        {
            "tab_name": "Normal Flow – Email",
            "tc_id": "TC-NOTIF-01_NF_Email",
            "tc_desc": "Receive System Notification – Normal Flow (Email Notification)",
            "prerequisites": ["Email notifications are globally enabled (UC-SET-01).", "User has a valid registered email address."],
            "test_data": ["Triggering event: cancellation processed"],
            "test_scenario": "Verify email notification is sent when a significant event occurs and email is globally enabled.",
            "steps": [
                ("Trigger a booking cancellation event.", "System checks if email notifications are globally enabled."),
                ("Email is enabled.", "System sends email to user's registered address."),
                ("Open email inbox.", "Email received with cancellation details."),
            ],
            "flow_color": NORMAL_FILL,
        },
        {
            "tab_name": "Normal Flow – Reminders",
            "tc_id": "TC-NOTIF-01_NF_Reminders",
            "tc_desc": "Receive System Notification – Normal Flow (Automated Reminders)",
            "prerequisites": ["Booking BK-2026-020 is 'awaiting payment' with deadline tomorrow.", "Booking BK-2026-021 check-in is tomorrow."],
            "test_data": ["BK-2026-020: payment deadline tomorrow", "BK-2026-021: check-in tomorrow"],
            "test_scenario": "Verify scheduled daily reminders are sent for approaching payment deadlines and check-in/out dates.",
            "steps": [
                ("Daily scheduled job runs.", "System finds 'awaiting payment' bookings with deadline approaching (tomorrow)."),
                ("Observe payment reminder.", "Payment reminder sent to guest for BK-2026-020: booking reference, amount due, deadline, direct payment link."),
                ("System finds confirmed bookings with check-in tomorrow.", "Check-in reminder sent to guest and admin for BK-2026-021."),
            ],
            "flow_color": NORMAL_FILL,
        },
        {
            "tab_name": "Alternate Flow",
            "tc_id": "TC-NOTIF-01_AF",
            "tc_desc": "Receive System Notification – Alternate Flow (Email Globally Disabled)",
            "prerequisites": ["Email notifications are globally disabled in system settings."],
            "test_data": ["Email toggle: disabled"],
            "test_scenario": "Verify that in-app notifications are still delivered when email is globally disabled.",
            "steps": [
                ("Trigger a significant event (booking confirmed).", "System checks email notification toggle."),
                ("Email is disabled.", "System skips email sending. No email is sent."),
                ("Observe in-app notification.", "In-app notification is still created and delivered normally."),
            ],
            "flow_color": ALT_FILL,
        },
        {
            "tab_name": "Exception Flow",
            "tc_id": "TC-NOTIF-01_EF",
            "tc_desc": "Receive System Notification – Exception Flow (Email Server Unreachable)",
            "prerequisites": ["Email notifications are enabled. Mail server is simulated to be unreachable."],
            "test_data": ["Mail server: simulated unreachable"],
            "test_scenario": "Verify that email server failure is logged, retried, and does not affect in-app notifications.",
            "steps": [
                ("Trigger a significant event while mail server is unreachable.", "System attempts to send email."),
                ("Email send fails.", "System retries. After max retries, failure is recorded. In-app notification is unaffected and delivered normally."),
            ],
            "flow_color": EXCEP_FILL,
        },
    ]
    create_workbook("TC-NOTIF-01_ReceiveSystemNotification.xlsx", sheets)


def gen_notif02():
    sheets = [
        {
            "tab_name": "Normal Flow",
            "tc_id": "TC-NOTIF-02_NF",
            "tc_desc": "Google Calendar Integration – Normal Flow (Booking Added to Calendar)",
            "prerequisites": ["User has connected their Google account to HomeLodge.", "Booking BK-2026-001 is confirmed."],
            "test_data": ["Confirmed booking: BK-2026-001", "Check-in: 2026-07-10 2PM", "Check-out: 2026-07-13 12PM"],
            "test_scenario": "Verify a confirmed booking is automatically added to the user's connected Google Calendar.",
            "steps": [
                ("Payment confirmed for BK-2026-001.", "System detects booking confirmed event."),
                ("System retrieves user's stored Google account connection.", "Connection found."),
                ("System creates Google Calendar event.", "Event created: title = 'Unit Name Stay', start = 2026-07-10 2PM, end = 2026-07-13 12PM, description = booking reference + unit address."),
                ("Verify event in Google Calendar.", "Event appears in user's Google Calendar."),
            ],
            "flow_color": NORMAL_FILL,
        },
        {
            "tab_name": "Alternate Flow",
            "tc_id": "TC-NOTIF-02_AF",
            "tc_desc": "Google Calendar Integration – Alternate Flow (Google Account Not Connected)",
            "prerequisites": ["User has NOT connected Google account to HomeLodge."],
            "test_data": ["User: no Google connection stored"],
            "test_scenario": "Verify the system silently skips calendar integration when Google account is not connected.",
            "steps": [
                ("Booking confirmed for a user without Google connection.", "System checks for stored Google account connection."),
                ("No connection found.", "System skips calendar step. No error is shown to the user."),
            ],
            "flow_color": ALT_FILL,
        },
        {
            "tab_name": "Exception Flow",
            "tc_id": "TC-NOTIF-02_EF",
            "tc_desc": "Google Calendar Integration – Exception Flow (Calendar Service Error)",
            "prerequisites": ["User has connected Google account. Google Calendar API returns an error or connection is expired."],
            "test_data": ["Simulated Google Calendar API error (e.g., token expired)"],
            "test_scenario": "Verify that Google Calendar API failures are logged without affecting the booking confirmation flow.",
            "steps": [
                ("Booking confirmed. System attempts to create calendar event. Google Calendar API returns error.", "API call fails."),
                ("System records the failure.", "Failure logged in DB. User may be prompted to reconnect Google account. Booking confirmation is NOT affected."),
            ],
            "flow_color": EXCEP_FILL,
        },
    ]
    create_workbook("TC-NOTIF-02_GoogleCalendarIntegration.xlsx", sheets)


# ══════════════════════════════════════════════════════════════
# MODULE 6 – CHAT
# ══════════════════════════════════════════════════════════════

def gen_chat01():
    sheets = [
        {
            "tab_name": "Normal Flow",
            "tc_id": "TC-CHAT-01_NF",
            "tc_desc": "Send / Receive Messages – Normal Flow (Real-Time Delivery)",
            "prerequisites": ["Both Guest and Admin accounts exist and are signed in.", "Guest is on Chat page."],
            "test_data": ["Message: 'Hello, I have a question about my booking.'"],
            "test_scenario": "Verify a guest can send a message that is instantly delivered to the admin and stored in the system.",
            "steps": [
                ("Guest opens the Chat page.", "Chat interface displayed."),
                ("Guest types 'Hello, I have a question about my booking.' in the input field.", "Text appears in the input field."),
                ("Guest clicks 'Send' or presses Enter.", "Message submitted."),
                ("Message is saved: sender = guest, recipient = admin, content, timestamp.", "Saved to DB."),
                ("Admin is currently online.", "Message delivered instantly to admin's chat window. Unread count badge updated on admin's chat icon."),
                ("Message appears in guest's chat window.", "Message shown as 'sent'."),
            ],
            "flow_color": NORMAL_FILL,
        },
        {
            "tab_name": "Alternate Flow",
            "tc_id": "TC-CHAT-01_AF",
            "tc_desc": "Send / Receive Messages – Alternate Flows (Empty Message, Offline Recipient)",
            "prerequisites": ["Guest is signed in. Admin is offline."],
            "test_data": ["A1: Empty message (no text)", "A2: Admin is offline when message sent"],
            "test_scenario": "Verify empty messages are blocked and offline recipient messages are stored for later delivery.",
            "steps": [
                ("(A1) Leave message input blank. Click 'Send'.", "Send button is disabled when input is blank. Empty message cannot be sent."),
                ("(A2) Send a valid message while admin is offline.", "Message is saved to the system. When admin next signs in, the message is visible in their chat."),
            ],
            "flow_color": ALT_FILL,
        },
        {
            "tab_name": "Exception Flow",
            "tc_id": "TC-CHAT-01_EF",
            "tc_desc": "Send / Receive Messages – Exception Flow (Real-Time Connection Lost)",
            "prerequisites": ["Guest is on Chat page. Network connection is interrupted."],
            "test_data": ["Simulated network disconnection"],
            "test_scenario": "Verify the system shows a reconnecting indicator and still preserves the message on reconnection.",
            "steps": [
                ("Simulate a network disconnection while on Chat page.", "'Reconnecting...' indicator is shown."),
                ("Reconnection restored.", "Message is still saved in the system. Message becomes visible on page reload."),
            ],
            "flow_color": EXCEP_FILL,
        },
    ]
    create_workbook("TC-CHAT-01_SendReceiveMessages.xlsx", sheets)


def gen_chat02():
    sheets = [
        {
            "tab_name": "Normal Flow",
            "tc_id": "TC-CHAT-02_NF",
            "tc_desc": "View Chat History – Normal Flow",
            "prerequisites": ["Conversation with at least 5 messages exists between guest and admin.", "Some messages are unread."],
            "test_data": ["User opens Chat page with existing conversation"],
            "test_scenario": "Verify all messages are displayed in chronological order, unread messages marked as read, and view auto-scrolls to the latest.",
            "steps": [
                ("Open the Chat page.", "System retrieves all messages in the conversation (oldest to newest)."),
                ("Observe message display.", "Sent messages on the right. Received messages on the left. Each with sender name and timestamp."),
                ("Observe unread status.", "All unread messages are marked as read upon view."),
                ("Observe scroll position.", "Chat auto-scrolls to the most recent message."),
            ],
            "flow_color": NORMAL_FILL,
        },
    ]
    create_workbook("TC-CHAT-02_ViewChatHistory.xlsx", sheets)


# ══════════════════════════════════════════════════════════════
# MODULE 7 – USER & ACCESS MANAGEMENT
# ══════════════════════════════════════════════════════════════

def gen_usr01():
    sheets = [
        {
            "tab_name": "Normal Flow",
            "tc_id": "TC-USR-01_NF",
            "tc_desc": "Create User Account – Normal Flow (Successful Creation)",
            "prerequisites": ["Admin is signed in with user management permission.", "Email 'newstaff@example.com' is not registered."],
            "test_data": ["Name: 'Staff Member One'", "Email: 'newstaff@example.com'", "Role: 'Admin'"],
            "test_scenario": "Verify admin can create a new user account with a temporary password, forced-change flag set, and notification sent to the new user.",
            "steps": [
                ("Navigate to User Management → Create User.", "Create user form displayed."),
                ("Enter name 'Staff Member One', email 'newstaff@example.com', role 'Admin'.", "Fields accept input."),
                ("Click 'Create'.", "System generates a temporary password."),
                ("Observe system response.", "User record created with forced password change flag set. Email sent to 'newstaff@example.com' with temporary password and sign-in link. Audit log entry recorded. Admin sees success message."),
            ],
            "flow_color": NORMAL_FILL,
        },
        {
            "tab_name": "Alternate Flow",
            "tc_id": "TC-USR-01_AF",
            "tc_desc": "Create User Account – Alternate Flow (Email Already Registered)",
            "prerequisites": ["Email 'existing@example.com' is already registered."],
            "test_data": ["Email: 'existing@example.com'"],
            "test_scenario": "Verify that user creation is rejected if the email is already in use.",
            "steps": [
                ("Enter email 'existing@example.com' and fill other fields. Click Create.", "Validation runs."),
                ("Observe error.", "Validation error shown: 'Email already registered.' Admin can edit the existing account instead. No new account created."),
            ],
            "flow_color": ALT_FILL,
        },
    ]
    create_workbook("TC-USR-01_CreateUserAccount.xlsx", sheets)


def gen_usr02():
    sheets = [
        {
            "tab_name": "Normal Flow – Edit",
            "tc_id": "TC-USR-02_NF_Edit",
            "tc_desc": "Edit / Activate / Deactivate User – Normal Flow: Edit User",
            "prerequisites": ["User 'staff@example.com' exists. Admin is signed in."],
            "test_data": ["Updated Name: 'Staff Member Updated'", "Updated Role: 'Guest'"],
            "test_scenario": "Verify admin can edit user details and the changes are saved.",
            "steps": [
                ("Navigate to User Management → select 'staff@example.com' → click 'Edit'.", "Pre-filled edit form displayed."),
                ("Update Name to 'Staff Member Updated' and Role to 'Guest'. Click Save.", "Validation runs (email uniqueness)."),
                ("Observe system response.", "Updated record saved. If role changed, new permissions take effect immediately. Audit log entry recorded. Success message shown."),
            ],
            "flow_color": NORMAL_FILL,
        },
        {
            "tab_name": "Normal Flow – Activate/Deactivate",
            "tc_id": "TC-USR-02_NF_ActivateDeactivate",
            "tc_desc": "Edit / Activate / Deactivate User – Normal Flow: Activate / Deactivate",
            "prerequisites": ["Active user 'staff@example.com' exists. Admin signed in."],
            "test_data": ["Target user: 'staff@example.com'"],
            "test_scenario": "Verify admin can deactivate a user (ending active sessions) and reactivate them.",
            "steps": [
                ("Select 'staff@example.com'. Click 'Deactivate'.", "Confirmation dialog shown."),
                ("Confirm deactivation.", "Account disabled. All active sessions for that user are immediately ended. Audit log recorded. Success message."),
                ("Attempt to sign in as 'staff@example.com'.", "Login rejected: 'Your account has been deactivated.'"),
                ("In admin panel, select 'staff@example.com'. Click 'Activate'. Confirm.", "Account restored to active. Audit log recorded. Success message."),
                ("Attempt to sign in as 'staff@example.com' again.", "Login succeeds."),
            ],
            "flow_color": NORMAL_FILL,
        },
    ]
    create_workbook("TC-USR-02_EditActivateDeactivateUser.xlsx", sheets)


def gen_usr03():
    sheets = [
        {
            "tab_name": "Normal Flow",
            "tc_id": "TC-USR-03_NF",
            "tc_desc": "Reset User Password – Normal Flow (Send Reset Link & Set to Default)",
            "prerequisites": ["User 'staff@example.com' exists. Admin signed in with permission.", "Account is optionally locked."],
            "test_data": ["Option A: Send reset link", "Option B: Set to system default temp password"],
            "test_scenario": "Verify admin can reset a user's password via both methods, forced-change flag is set, and locked account is unlocked.",
            "steps": [
                ("Navigate to User Management → select 'staff@example.com' → click 'Reset Password'.", "Reset method selection displayed."),
                ("(Option A) Choose 'Send reset link'. Submit.", "Password reset email sent to 'staff@example.com'. Forced change flag set. If account was locked: unlocked, failed-attempt counter reset. User notified. Audit log recorded."),
                ("(Option B) Choose 'Set to default'. Submit.", "Password immediately set to default temporary password. Forced change flag set. Account unlocked if applicable. User notified. Audit log recorded."),
                ("User signs in with new/temp password.", "Login succeeds. User is immediately prompted to change password (forced change page)."),
            ],
            "flow_color": NORMAL_FILL,
        },
    ]
    create_workbook("TC-USR-03_ResetUserPassword.xlsx", sheets)


def gen_usr04():
    sheets = [
        {
            "tab_name": "Normal Flow – Create & Edit",
            "tc_id": "TC-USR-04_NF_CreateEdit",
            "tc_desc": "Manage Roles – Normal Flow: Create & Edit Role",
            "prerequisites": ["Admin is signed in with role management access."],
            "test_data": ["Role Name: 'Property Manager'", "Updated Name: 'Senior Property Manager'"],
            "test_scenario": "Verify admin can create and edit roles.",
            "steps": [
                ("Navigate to Role Management → Create Role.", "Create role form displayed."),
                ("Enter name 'Property Manager'. Click Create.", "System checks name is unique. Role created. Admin navigated to role detail page to assign permissions."),
                ("Select 'Property Manager' role. Click Edit.", "Edit form shown."),
                ("Update name to 'Senior Property Manager'. Save.", "Name uniqueness checked. Updated role saved."),
            ],
            "flow_color": NORMAL_FILL,
        },
        {
            "tab_name": "Normal Flow – Assign & Delete",
            "tc_id": "TC-USR-04_NF_AssignDelete",
            "tc_desc": "Manage Roles – Normal Flow: Assign Permissions & Delete Role",
            "prerequisites": ["Role 'Temp Role' exists with no assigned users."],
            "test_data": ["Permissions to assign: 'manage_bookings', 'view_reports'"],
            "test_scenario": "Verify admin can assign permissions to a role and delete a role with no assigned users.",
            "steps": [
                ("Navigate to Role Management → select 'Senior Property Manager' → Manage Permissions.", "Checklist of all permissions displayed; currently assigned are ticked."),
                ("Tick 'manage_bookings' and 'view_reports'. Click Save.", "Role permissions updated. Takes effect immediately for all users with this role. Audit log recorded."),
                ("Select 'Temp Role'. Click Delete.", "System checks: no users have this role."),
                ("Confirm deletion.", "Role deleted. Audit log recorded."),
            ],
            "flow_color": NORMAL_FILL,
        },
        {
            "tab_name": "Alternate Flow",
            "tc_id": "TC-USR-04_AF",
            "tc_desc": "Manage Roles – Alternate Flow (Delete Role Assigned to Users)",
            "prerequisites": ["Role 'Guest' is currently assigned to multiple users."],
            "test_data": ["Role: 'Guest' (assigned to 10 users)"],
            "test_scenario": "Verify deletion is blocked when a role is assigned to one or more users.",
            "steps": [
                ("Select 'Guest' role. Click Delete.", "System checks: 10 users have this role."),
                ("Observe error.", "Action blocked. Message: 'This role is assigned to 10 user(s). Reassign users before deleting.'"),
            ],
            "flow_color": ALT_FILL,
        },
    ]
    create_workbook("TC-USR-04_ManageRoles.xlsx", sheets)


def gen_usr05():
    sheets = [
        {
            "tab_name": "Normal Flow",
            "tc_id": "TC-USR-05_NF",
            "tc_desc": "Manage Permissions – Normal Flow: Create, Edit, Delete",
            "prerequisites": ["Admin is signed in with permission management access.", "Permission 'temp_permission' exists and is NOT attached to any role."],
            "test_data": ["Create: name 'can_export_reports'", "Edit: to 'can_export_all_reports'", "Delete: 'temp_permission'"],
            "test_scenario": "Verify admin can create, edit, and delete permissions.",
            "steps": [
                ("Navigate to Permission Management → Create Permission. Enter name 'can_export_reports'. Save.", "Name uniqueness checked. Permission created. Success shown."),
                ("Select 'can_export_reports'. Click Edit. Update to 'can_export_all_reports'. Save.", "Update saved."),
                ("Select 'temp_permission'. Click Delete.", "System checks: not attached to any roles."),
                ("Confirm deletion.", "Permission deleted. Audit log recorded."),
            ],
            "flow_color": NORMAL_FILL,
        },
        {
            "tab_name": "Alternate Flow",
            "tc_id": "TC-USR-05_AF",
            "tc_desc": "Manage Permissions – Alternate Flow (Delete Permission Attached to Roles)",
            "prerequisites": ["Permission 'manage_bookings' is attached to 2 roles."],
            "test_data": ["Permission: 'manage_bookings' (used by 2 roles)"],
            "test_scenario": "Verify deletion is blocked when a permission is still attached to roles.",
            "steps": [
                ("Select 'manage_bookings'. Click Delete.", "System checks: attached to 2 roles."),
                ("Observe response.", "Action blocked. Message: 'This permission is attached to 2 role(s). Remove from roles before deleting.'"),
            ],
            "flow_color": ALT_FILL,
        },
    ]
    create_workbook("TC-USR-05_ManagePermissions.xlsx", sheets)


# ══════════════════════════════════════════════════════════════
# MODULE 8 – SYSTEM SETTINGS
# ══════════════════════════════════════════════════════════════

def gen_set01():
    sheets = [
        {
            "tab_name": "Normal Flow",
            "tc_id": "TC-SET-01_NF",
            "tc_desc": "Configure System Settings – Normal Flow (Save Valid Settings)",
            "prerequisites": ["Admin is signed in with system settings access."],
            "test_data": ["Security: Max failed attempts = 5, Lockout = 30 min", "Cancellation: >14d = 100%, 7-14d = 50%, <7d = 0%", "Extension: RM 30/hr, RM 150/night", "Email toggle: enabled"],
            "test_scenario": "Verify admin can update all system settings sections and saved values take effect immediately.",
            "steps": [
                ("Navigate to System Settings.", "Settings page displayed with multiple sections."),
                ("Update Security settings: max attempts = 5, lockout = 30 min, session timeout = 120 min.", "Fields accept values."),
                ("Update Cancellation Policy: >14d=100%, 7-14d=50%, <7d=0%.", "Tier values entered."),
                ("Update Extension Charges: RM 30/hr, RM 150/night. Extension payment window: 60 min.", "Fields accept values."),
                ("Toggle Email Notifications to 'Enabled'.", "Toggle set."),
                ("Click Save.", "System validates settings."),
                ("Observe system response.", "Settings saved. Confirmation message shown: 'Settings saved. All future operations use new values.'"),
            ],
            "flow_color": NORMAL_FILL,
        },
        {
            "tab_name": "Alternate Flow",
            "tc_id": "TC-SET-01_AF",
            "tc_desc": "Configure System Settings – Alternate Flow (Email Connection Test Fails, Validation Error)",
            "prerequisites": ["Admin is on System Settings page. Mail server is simulated to be unreachable."],
            "test_data": ["A1: Invalid mail server address", "A2: Refund tier: 80% for >14 days and 90% for 7-14 days (overlapping)"],
            "test_scenario": "Verify email connection test failure and validation errors are reported.",
            "steps": [
                ("(A1) Enter an invalid mail server address. Click 'Test Connection'.", "Connection test fails. Error message shown with specific connection error. Admin can correct and retry."),
                ("(A2) Enter overlapping refund tiers (e.g., both tiers = 80%). Click Save.", "Validation detects overlapping/invalid tiers. Error shown. Settings NOT saved."),
            ],
            "flow_color": ALT_FILL,
        },
    ]
    create_workbook("TC-SET-01_ConfigureSystemSettings.xlsx", sheets)


def gen_set02():
    sheets = [
        {
            "tab_name": "Normal Flow",
            "tc_id": "TC-SET-02_NF",
            "tc_desc": "Manage Default Homestay Policies – Normal Flow (Add, Edit, Remove)",
            "prerequisites": ["Admin is signed in with system settings access.", "Default policies: No Pets, No Durians, No Smoking."],
            "test_data": ["Add: 'No Parties'", "Edit: 'No Smoking' → 'No Smoking or Vaping'", "Remove: 'No Durians'"],
            "test_scenario": "Verify admin can add, edit, and remove default house rules, and changes apply only to new units.",
            "steps": [
                ("Navigate to System Settings → Default Policies.", "Current list displayed: No Pets, No Durians, No Smoking."),
                ("Click 'Add'. Enter 'No Parties'. Save.", "New rule added to default list."),
                ("Select 'No Smoking'. Click 'Edit'. Update to 'No Smoking or Vaping'. Save.", "Rule text updated."),
                ("Select 'No Durians'. Click 'Delete'. Confirm.", "Confirmation prompt shown. After confirm, rule removed from list."),
                ("Create a new unit after these changes.", "New unit inherits: No Pets, No Smoking or Vaping, No Parties. No Durians is not applied."),
                ("Verify existing units.", "Existing units' house rules are NOT changed by this action."),
            ],
            "flow_color": NORMAL_FILL,
        },
    ]
    create_workbook("TC-SET-02_ManageDefaultHomestayPolicies.xlsx", sheets)


# ══════════════════════════════════════════════════════════════
# MODULE 9 – AUDIT LOGS
# ══════════════════════════════════════════════════════════════

def gen_audit01():
    sheets = [
        {
            "tab_name": "Normal Flow",
            "tc_id": "TC-AUDIT-01_NF",
            "tc_desc": "View / Filter Audit Trail – Normal Flow",
            "prerequisites": ["Admin is signed in with audit log access.", "Multiple log entries exist."],
            "test_data": ["Filter: date range = 2026-07-01 to 2026-07-15", "Filter: event type = 'booking_created'"],
            "test_scenario": "Verify admin can view the full audit trail and apply filters to find specific events.",
            "steps": [
                ("Navigate to Audit Logs section.", "System retrieves all log entries in reverse chronological order (newest first)."),
                ("Observe displayed entries.", "Each entry shows: timestamp, actor (user name or 'System'), event type, what was affected."),
                ("Page through entries.", "Pagination works. Next/previous pages load correctly."),
                ("Apply date range filter: 2026-07-01 to 2026-07-15.", "List filters to entries within the date range only."),
                ("Apply event type filter: 'booking_created'.", "List shows only booking creation events."),
                ("Click 'Clear Filters'.", "Full unfiltered audit log is displayed again."),
                ("Attempt to edit or delete a log entry.", "No edit/delete option available. Log entries are read-only."),
            ],
            "flow_color": NORMAL_FILL,
        },
    ]
    create_workbook("TC-AUDIT-01_ViewFilterAuditTrail.xlsx", sheets)


def gen_audit02():
    sheets = [
        {
            "tab_name": "Normal Flow – User Action",
            "tc_id": "TC-AUDIT-02_NF_User",
            "tc_desc": "Automatic Event Logging – Normal Flow: User / Admin Action",
            "prerequisites": ["User 'guest@example.com' is signed in.", "A booking creation is about to be performed."],
            "test_data": ["Action: submit booking for BK-2026-030"],
            "test_scenario": "Verify a log entry is automatically created when a user performs a significant action.",
            "steps": [
                ("Guest submits a booking (BK-2026-030).", "Booking created."),
                ("Admin navigates to Audit Logs.", "Log entries listed."),
                ("Locate entry for BK-2026-030.", "Entry exists: timestamp = now, actor = 'guest@example.com', event type = 'booking_created', affected record = BK-2026-030."),
                ("Verify entry is permanent.", "No edit or delete option. Entry cannot be modified."),
            ],
            "flow_color": NORMAL_FILL,
        },
        {
            "tab_name": "Normal Flow – System Action",
            "tc_id": "TC-AUDIT-02_NF_System",
            "tc_desc": "Automatic Event Logging – Normal Flow: System / Automated Action",
            "prerequisites": ["Auto-cancel job runs. Booking BK-2026-031 was auto-cancelled."],
            "test_data": ["Auto-cancelled booking: BK-2026-031"],
            "test_scenario": "Verify that automated system actions are logged with 'System' as the actor.",
            "steps": [
                ("Scheduled auto-cancel job runs and cancels BK-2026-031.", "System cancels booking."),
                ("Admin navigates to Audit Logs.", "Log entries listed."),
                ("Locate entry for BK-2026-031 auto-cancel.", "Entry exists: timestamp = job run time, actor = 'System', event type = 'booking_auto_cancelled', affected = BK-2026-031."),
                ("Verify immutability.", "Entry cannot be edited or deleted."),
            ],
            "flow_color": NORMAL_FILL,
        },
    ]
    create_workbook("TC-AUDIT-02_AutomaticEventLogging.xlsx", sheets)


# ══════════════════════════════════════════════════════════════
# MODULE 10 – QR CODE & ACCESS
# ══════════════════════════════════════════════════════════════

def gen_qr01():
    sheets = [
        {
            "tab_name": "Normal Flow",
            "tc_id": "TC-QR-01_NF",
            "tc_desc": "Receive & Use QR Code – Normal Flow (QR Generated & Access Granted)",
            "prerequisites": ["Booking BK-2026-001 confirmed (payment received). Check-in: 2026-07-10 2PM, Check-out: 2026-07-13 12PM."],
            "test_data": ["Booking: BK-2026-001", "Current time: within valid window (e.g., 2026-07-10 3PM)"],
            "test_scenario": "Verify QR code is generated on payment, delivered to guest, and grants access when presented within the valid window.",
            "steps": [
                ("Payment confirmed for BK-2026-001.", "System generates unique QR code: status = active, valid from 2026-07-10 2PM to 2026-07-13 12PM."),
                ("Verify notifications.", "In-app notification sent with QR code. Email sent with QR code and booking details."),
                ("Guest views QR code on booking detail page.", "QR code is visible."),
                ("At 2026-07-10 3PM, guest presents QR code to smart lock scanner.", "Smart lock queries system for validation."),
                ("System validates: status = active, current time within valid window.", "Validation passes."),
                ("Observe result.", "Access granted. Door opens."),
            ],
            "flow_color": NORMAL_FILL,
        },
        {
            "tab_name": "Alternate Flow",
            "tc_id": "TC-QR-01_AF",
            "tc_desc": "Receive & Use QR Code – Alternate Flows (Expired, Inactive/Revoked, No Device)",
            "prerequisites": ["QR code for BK-2026-001 is expired (past check-out). QR code for BK-2026-002 is inactive/revoked."],
            "test_data": ["A1: QR code past check-out time", "A2: QR code status = inactive", "A3: Guest device unavailable (dead battery)"],
            "test_scenario": "Verify access is denied for expired, inactive, or revoked QR codes.",
            "steps": [
                ("(A1) Present expired QR code (past 2026-07-13 12PM) to scanner.", "Smart lock queries system. System detects code is expired. Response: 'QR Code Expired.' Access denied."),
                ("(A2) Present revoked/inactive QR code to scanner.", "System detects code is inactive. Access denied."),
                ("(A3) Guest's device is unavailable (no battery/no signal).", "Guest cannot present QR code. Guest must contact administrator for manual assistance."),
            ],
            "flow_color": ALT_FILL,
        },
    ]
    create_workbook("TC-QR-01_ReceiveUseQRCode.xlsx", sheets)


def gen_qr02():
    sheets = [
        {
            "tab_name": "Normal Flow – QR Expiry",
            "tc_id": "TC-QR-02_NF_Expiry",
            "tc_desc": "Manage Housekeeping Cycle – Normal Flow: Automatic QR Expiry",
            "prerequisites": ["Active QR code for BK-2026-001. Check-out was at 12PM (now past)."],
            "test_data": ["Booking BK-2026-001 check-out: 2026-07-13 12PM (now past)"],
            "test_scenario": "Verify the scheduled job expires QR codes past their valid-until time and marks bookings as completed.",
            "steps": [
                ("Scheduled job runs at/after 2026-07-13 12PM.", "System finds active QR codes with valid-until time passed."),
                ("For BK-2026-001 QR code:", "QR code status set to 'expired'. Booking BK-2026-001 status updated to 'completed'."),
            ],
            "flow_color": NORMAL_FILL,
        },
        {
            "tab_name": "Normal Flow – Housekeeping",
            "tc_id": "TC-QR-02_NF_Housekeeping",
            "tc_desc": "Manage Housekeeping Cycle – Normal Flow: Housekeeping QR & Complete",
            "prerequisites": ["Booking BK-2026-001 is 'completed'. Next booking BK-2026-002 confirmed on same unit."],
            "test_data": ["Housekeeping validity: 4 hours"],
            "test_scenario": "Verify admin can generate a housekeeping QR, mark housekeeping complete, and next guest automatically receives their QR code.",
            "steps": [
                ("Admin clicks 'Generate Housekeeping QR' for the unit. Set validity: 4 hours.", "System generates temporary QR code type 'housekeeping' with 4-hour validity."),
                ("QR code displayed to admin.", "Admin can share QR code with cleaning staff."),
                ("Admin clicks 'Mark Housekeeping Complete'.", "Housekeeping QR code set to 'expired'. System checks for next confirmed booking."),
                ("Next booking BK-2026-002 found.", "New QR code generated for next guest. In-app and email notification with QR code sent to next guest."),
            ],
            "flow_color": NORMAL_FILL,
        },
        {
            "tab_name": "Alternate Flow",
            "tc_id": "TC-QR-02_AF",
            "tc_desc": "Manage Housekeeping Cycle – Alternate Flow (No Next Booking)",
            "prerequisites": ["After marking housekeeping complete, no next confirmed booking exists on the unit."],
            "test_data": ["No upcoming confirmed booking for unit after current checkout"],
            "test_scenario": "Verify no QR code is generated when there is no next confirmed booking after housekeeping.",
            "steps": [
                ("Admin marks housekeeping complete.", "Housekeeping QR code expired. System checks for next confirmed booking."),
                ("No next booking found.", "System skips QR code generation. No action needed from admin. No notification sent."),
            ],
            "flow_color": ALT_FILL,
        },
    ]
    create_workbook("TC-QR-02_ManageHousekeepingCycle.xlsx", sheets)


def gen_qr03():
    sheets = [
        {
            "tab_name": "Normal Flow",
            "tc_id": "TC-QR-03_NF",
            "tc_desc": "Initiate Booking Extension – Normal Flow",
            "prerequisites": ["Booking BK-2026-001 is 'confirmed'. Check-out currently 2026-07-13 12PM. Dates 2026-07-13 12PM to 2026-07-14 12PM are free."],
            "test_data": ["Extension type: Date extension (add 1 night)", "New check-out: 2026-07-14 12PM", "Rate: RM 150/night"],
            "test_scenario": "Verify admin can initiate a booking extension, extension record is created, bill generated, and guest notified.",
            "steps": [
                ("Admin opens BK-2026-001. Click 'Extend Stay'.", "Extension form displayed."),
                ("Select 'Date extension'. Enter new check-out: 2026-07-14 12PM.", "Availability checked for 2026-07-13 to 2026-07-14."),
                ("Dates are available.", "Additional charge calculated: 1 night × RM 150 = RM 150."),
                ("Observe system actions.", "Extension record created (status: awaiting payment). Extension bill generated. Payment deadline set (unit-specific window or system default). Guest notified with charge amount, deadline, and direct payment link. Audit log recorded."),
                ("Verify QR code.", "QR code validity is NOT extended at this point. Still valid until 2026-07-13 12PM."),
            ],
            "flow_color": NORMAL_FILL,
        },
        {
            "tab_name": "Alternate Flow",
            "tc_id": "TC-QR-03_AF",
            "tc_desc": "Initiate Booking Extension – Alternate Flow (Date Conflict)",
            "prerequisites": ["Another confirmed booking occupies 2026-07-13 12PM to 2026-07-14 12PM on the same unit."],
            "test_data": ["Conflicting booking: BK-2026-005 on same unit"],
            "test_scenario": "Verify extension is blocked when the requested extended period conflicts with another booking.",
            "steps": [
                ("Admin attempts to extend BK-2026-001 to 2026-07-14 12PM.", "System checks availability for extended period."),
                ("Conflict detected with BK-2026-005.", "System shows conflicting booking details. Message: 'Select a different extension period.' No extension record created."),
            ],
            "flow_color": ALT_FILL,
        },
    ]
    create_workbook("TC-QR-03_InitiateBookingExtension.xlsx", sheets)


def gen_qr04():
    sheets = [
        {
            "tab_name": "Normal Flow",
            "tc_id": "TC-QR-04_NF",
            "tc_desc": "Pay Extension Charge – Normal Flow",
            "prerequisites": ["Extension EXT-001 for BK-2026-001 is 'awaiting payment'. Payment deadline has not passed."],
            "test_data": ["Extension charge: RM 150", "Payment method: test card 4242 4242 4242 4242"],
            "test_scenario": "Verify guest can pay the extension charge, booking is updated to new check-out time, and QR code validity is extended.",
            "steps": [
                ("Guest opens booking detail. Finds pending extension charge EXT-001. Click 'Pay Extension Charge'.", "Redirected to secure payment page with RM 150."),
                ("Complete payment with test card.", "Payment submitted."),
                ("Gateway sends payment confirmation.", "System verifies and records payment as successful."),
                ("Observe system updates.", "Extension EXT-001 status = 'confirmed'. Booking check-out updated to 2026-07-14 12PM. QR code validity extended to 2026-07-14 12PM."),
                ("Verify notification.", "Guest notified: 'Stay extended. QR code valid until 2026-07-14 12PM.' Audit log recorded."),
            ],
            "flow_color": NORMAL_FILL,
        },
    ]
    create_workbook("TC-QR-04_PayExtensionCharge.xlsx", sheets)


def gen_qr05():
    sheets = [
        {
            "tab_name": "Normal Flow",
            "tc_id": "TC-QR-05_NF",
            "tc_desc": "Auto-Cancel Extension – Normal Flow (Extension Deadline Passed)",
            "prerequisites": ["Extension EXT-002 for BK-2026-002 is 'awaiting payment'. Payment deadline has passed."],
            "test_data": ["EXT-002: awaiting payment, deadline passed"],
            "test_scenario": "Verify the scheduler auto-cancels overdue extensions, booking reverts to original check-out, and guest is notified.",
            "steps": [
                ("Scheduled job runs (every few minutes).", "System finds extension records: status = 'awaiting payment' AND deadline has passed."),
                ("EXT-002 found.", "Extension status set to 'cancelled'. Booking BK-2026-002 check-out reverted to original value."),
                ("Verify QR code.", "QR code is NOT modified. Still reflects the original check-out time."),
                ("Verify notification.", "Guest notified: 'Extension cancelled – payment not received. Original check-out: [original date/time] remains.' Audit log recorded."),
            ],
            "flow_color": NORMAL_FILL,
        },
    ]
    create_workbook("TC-QR-05_AutoCancelExtension.xlsx", sheets)


# ══════════════════════════════════════════════════════════════
# MODULE 11 – REPORTING & ANALYTICS
# ══════════════════════════════════════════════════════════════

def gen_rpt01():
    sheets = [
        {
            "tab_name": "Normal Flow",
            "tc_id": "TC-RPT-01_NF",
            "tc_desc": "View Analytics Dashboard – Normal Flow",
            "prerequisites": ["Admin is signed in with reporting access.", "Multiple bookings, payments, and reviews exist."],
            "test_data": ["Date filter: July 2026", "Unit filter: 'Cozy Cottage'"],
            "test_scenario": "Verify admin can view all KPI cards and interactive charts on the analytics dashboard.",
            "steps": [
                ("Navigate to Reporting & Analytics section.", "Dashboard loads."),
                ("Observe KPI cards.", "Displayed: Total bookings (month + YTD), Total revenue (month + YTD), Occupancy rate, Cancellation rate, Average guest rating."),
                ("Observe charts.", "Booking trends line chart (daily/weekly/monthly toggle), Revenue summary, Per-unit bar chart, Feedback & rating summary per unit."),
                ("Click on a chart to interact.", "More detail or drill-down is available."),
                ("Apply date filter: July 2026.", "Dashboard updates to reflect filtered date range."),
                ("Apply unit filter: 'Cozy Cottage'.", "Data filtered to show only 'Cozy Cottage' metrics."),
            ],
            "flow_color": NORMAL_FILL,
        },
    ]
    create_workbook("TC-RPT-01_ViewAnalyticsDashboard.xlsx", sheets)


def gen_rpt02():
    sheets = [
        {
            "tab_name": "Normal Flow",
            "tc_id": "TC-RPT-02_NF",
            "tc_desc": "View Revenue Report – Normal Flow",
            "prerequisites": ["Admin is signed in with reporting access. Payment records exist."],
            "test_data": ["Filter: date range = 2026-07-01 to 2026-07-31", "Filter: unit = all", "Filter: status = successful"],
            "test_scenario": "Verify admin can view a filtered revenue report with totals, itemised breakdown, and unit summary.",
            "steps": [
                ("Navigate to Reporting → Revenue Report.", "Revenue report page displayed."),
                ("Apply filters: date range July 2026, all units, status successful.", "Filters applied."),
                ("Observe report content.", "Displayed: total revenue for period, itemised breakdown per booking (booking ref, guest, unit, amount), summary per unit."),
                ("Click 'Export CSV'.", "Export triggered (see TC-RPT-03)."),
            ],
            "flow_color": NORMAL_FILL,
        },
    ]
    create_workbook("TC-RPT-02_ViewRevenueReport.xlsx", sheets)


def gen_rpt03():
    sheets = [
        {
            "tab_name": "Normal Flow",
            "tc_id": "TC-RPT-03_NF",
            "tc_desc": "Export Report – Normal Flow (PDF & CSV Download)",
            "prerequisites": ["Admin is viewing a report page with data loaded."],
            "test_data": ["Export type: PDF", "Export type: CSV"],
            "test_scenario": "Verify admin can export report data as both PDF and CSV files.",
            "steps": [
                ("On a report page with data loaded, click 'Export PDF'.", "System collects filtered report data."),
                ("Observe result.", "PDF document generated and automatically downloaded to admin's device."),
                ("On the same (or another) report page, click 'Export CSV'.", "System collects filtered report data."),
                ("Observe result.", "CSV file generated and automatically downloaded to admin's device."),
            ],
            "flow_color": NORMAL_FILL,
        },
    ]
    create_workbook("TC-RPT-03_ExportReport.xlsx", sheets)


# ══════════════════════════════════════════════════════════════
# MODULE 12 – GUEST FEEDBACK
# ══════════════════════════════════════════════════════════════

def gen_fb01():
    sheets = [
        {
            "tab_name": "Normal Flow",
            "tc_id": "TC-FB-01_NF",
            "tc_desc": "Submit Rating & Feedback – Normal Flow",
            "prerequisites": ["Booking BK-2026-001 has status 'completed'. No review submitted yet. Guest is signed in."],
            "test_data": ["Star rating: 4", "Comment: 'Great stay, very clean and comfortable!'"],
            "test_scenario": "Verify guest can submit a review for a completed booking, it is visible, and average rating is recalculated.",
            "steps": [
                ("Navigate to My Bookings → History. Find BK-2026-001. Click 'Leave a Review'.", "System verifies: booking is completed AND no review submitted."),
                ("Feedback form displayed.", "Star rating (1–5) and optional comment text area shown."),
                ("Select 4 stars. Enter comment 'Great stay, very clean and comfortable!' Click Submit.", "Review submitted."),
                ("Observe system response.", "Review saved (linked to booking, unit, guest; visibility = visible). Unit's average rating recalculated. Thank-you message shown."),
                ("Observe 'Leave a Review' button for BK-2026-001.", "Button is hidden/replaced with 'View Your Review'."),
                ("Verify review is visible on unit listing/detail page.", "Review appears with 4 stars and comment."),
            ],
            "flow_color": NORMAL_FILL,
        },
        {
            "tab_name": "Alternate Flow",
            "tc_id": "TC-FB-01_AF",
            "tc_desc": "Submit Rating & Feedback – Alternate Flows (No Star Rating, Already Reviewed)",
            "prerequisites": ["Booking BK-2026-002 is 'completed'. BK-2026-003 already has a submitted review."],
            "test_data": ["A1: Submit without selecting star rating", "A2: Attempt to review already-reviewed booking"],
            "test_scenario": "Verify that missing star rating blocks submission, and already-reviewed bookings show 'View Your Review'.",
            "steps": [
                ("(A1) Open review form for BK-2026-002. Write a comment. Do NOT select a star rating. Click Submit.", "Star rating field highlighted as required. Submission blocked."),
                ("(A2) Navigate to BK-2026-003 (already reviewed).", "'Leave a Review' button is hidden. 'View Your Review' is shown instead. Submission form is not accessible."),
            ],
            "flow_color": ALT_FILL,
        },
        {
            "tab_name": "Exception Flow",
            "tc_id": "TC-FB-01_EF",
            "tc_desc": "Submit Rating & Feedback – Exception Flow (Non-Completed Booking Direct URL Access)",
            "prerequisites": ["Booking BK-2026-004 is 'confirmed' (not yet completed)."],
            "test_data": ["Direct URL access to review form for BK-2026-004"],
            "test_scenario": "Verify that review submission is blocked for non-completed bookings even via direct URL.",
            "steps": [
                ("Attempt to access the review form for BK-2026-004 via direct URL.", "System verifies booking status."),
                ("Observe system response.", "Access blocked. Validation error shown. No review can be submitted."),
            ],
            "flow_color": EXCEP_FILL,
        },
    ]
    create_workbook("TC-FB-01_SubmitRatingFeedback.xlsx", sheets)


def gen_fb02():
    sheets = [
        {
            "tab_name": "Normal Flow",
            "tc_id": "TC-FB-02_NF",
            "tc_desc": "View Submitted Feedback – Normal Flow (Guest Views Own Reviews)",
            "prerequisites": ["Guest has submitted at least 2 reviews. One has an admin reply."],
            "test_data": ["Guest with 2 submitted reviews"],
            "test_scenario": "Verify guest can view all their submitted reviews in read-only format, including admin replies.",
            "steps": [
                ("Navigate to 'My Reviews' or view a completed booking in history.", "System retrieves all review records for the signed-in guest."),
                ("Observe displayed entries.", "Each entry (read-only): unit name, check-in/out dates, star rating, comment, submission date, admin reply (if any)."),
                ("Attempt to edit a submitted review.", "No edit option available. Reviews are read-only after submission."),
            ],
            "flow_color": NORMAL_FILL,
        },
    ]
    create_workbook("TC-FB-02_ViewSubmittedFeedback.xlsx", sheets)


def gen_fb03():
    sheets = [
        {
            "tab_name": "Normal Flow – View",
            "tc_id": "TC-FB-03_NF_View",
            "tc_desc": "View / Manage All Feedback (Admin) – Normal Flow: View & Filter",
            "prerequisites": ["Admin is signed in with feedback management access. Multiple reviews exist (some hidden)."],
            "test_data": ["Filter: unit = 'Cozy Cottage'", "Filter: visibility = 'hidden'"],
            "test_scenario": "Verify admin can view all reviews including hidden ones and apply filters.",
            "steps": [
                ("Navigate to Guest Feedback management.", "System retrieves all review records including hidden ones."),
                ("Observe displayed list.", "Each entry: guest name, unit, booking ref, rating, comment, date, visibility status, admin reply status."),
                ("Apply filter: unit = 'Cozy Cottage'.", "List updates to show only 'Cozy Cottage' reviews."),
                ("Apply filter: visibility = 'hidden'.", "Only hidden reviews displayed."),
            ],
            "flow_color": NORMAL_FILL,
        },
        {
            "tab_name": "Normal Flow – Reply",
            "tc_id": "TC-FB-03_NF_Reply",
            "tc_desc": "View / Manage All Feedback (Admin) – Normal Flow: Reply to Review",
            "prerequisites": ["Review REV-001 exists with no admin reply."],
            "test_data": ["Reply text: 'Thank you for your feedback! We hope to see you again.'"],
            "test_scenario": "Verify admin can publish a reply to a review that is immediately visible on the unit page.",
            "steps": [
                ("Select REV-001. Click 'Reply'.", "Reply text area displayed."),
                ("Type 'Thank you for your feedback! We hope to see you again.' Click 'Publish Reply'.", "Reply submitted."),
                ("Observe system response.", "Reply saved with reply date. Reply immediately visible alongside REV-001 on the unit detail page."),
                ("Verify guest sees reply.", "When guest views their submitted feedback (UC-FB-02), admin reply is visible."),
            ],
            "flow_color": NORMAL_FILL,
        },
        {
            "tab_name": "Normal Flow – Hide & Restore",
            "tc_id": "TC-FB-03_NF_HideRestore",
            "tc_desc": "View / Manage All Feedback (Admin) – Normal Flow: Hide & Restore Review",
            "prerequisites": ["Review REV-002 with inappropriate content is visible."],
            "test_data": ["Review REV-002 with offensive content"],
            "test_scenario": "Verify admin can hide a review (removing from public view, recalculating rating) and restore it.",
            "steps": [
                ("Select REV-002. Click 'Hide'.", "Confirmation: 'Hide this review from public view?'"),
                ("Confirm.", "REV-002 visibility set to 'hidden'. Unit's average rating recalculated (excluding hidden review). Review removed from guest-facing unit detail page."),
                ("Select hidden REV-002. Click 'Restore'.", "System sets visibility back to 'visible'. Average rating recalculated (including REV-002 again)."),
            ],
            "flow_color": NORMAL_FILL,
        },
        {
            "tab_name": "Alternate Flow",
            "tc_id": "TC-FB-03_AF",
            "tc_desc": "View / Manage All Feedback (Admin) – Alternate Flows (Cancel Hide, Edit Reply)",
            "prerequisites": ["Review REV-003 is visible. Review REV-001 has an existing admin reply."],
            "test_data": ["A1: Admin clicks Cancel on hide dialog", "A2: Updated reply: 'Thank you for staying!'"],
            "test_scenario": "Verify cancelling the hide dialog makes no changes, and admin can edit an existing reply.",
            "steps": [
                ("(A1) Click 'Hide' on REV-003. When dialog appears, click 'Cancel'.", "Dialog dismissed. No changes made. REV-003 remains visible."),
                ("(A2) Select REV-001 (has existing reply). Click 'Edit Reply'. Update text to 'Thank you for staying!' Click Save.", "Reply updated. Reply date updated to edit time. Updated reply visible on unit page."),
            ],
            "flow_color": ALT_FILL,
        },
    ]
    create_workbook("TC-FB-03_ViewManageAllFeedbackAdmin.xlsx", sheets)


def gen_fb04():
    sheets = [
        {
            "tab_name": "Normal Flow",
            "tc_id": "TC-FB-04_NF",
            "tc_desc": "Display Average Rating – Normal Flow",
            "prerequisites": ["Unit 'Cozy Cottage' has 3 visible reviews: 4★, 5★, 3★. 1 hidden review: 2★."],
            "test_data": ["Visible reviews: [4, 5, 3] → avg = 4.0", "Hidden reviews: [2] → excluded"],
            "test_scenario": "Verify the average rating is calculated from visible reviews only and displayed correctly on listing and detail pages.",
            "steps": [
                ("Navigate to homestay listing page.", "System retrieves visible reviews for each unit."),
                ("Observe 'Cozy Cottage' card.", "Rating displayed: '4.0 ★ (3 reviews)'. Hidden review (2★) is NOT included."),
                ("Navigate to 'Cozy Cottage' detail page.", "Same rating '4.0 ★ (3 reviews)' displayed."),
                ("Admin submits a new visible review (5★) via UC-FB-01.", "Average recalculated: (4+5+3+5)/4 = 4.25 ★ (4 reviews). Updated on both listing and detail pages."),
            ],
            "flow_color": NORMAL_FILL,
        },
        {
            "tab_name": "Alternate Flow",
            "tc_id": "TC-FB-04_AF",
            "tc_desc": "Display Average Rating – Alternate Flow (No Visible Reviews)",
            "prerequisites": ["Unit 'New Unit' has no submitted reviews or all reviews are hidden."],
            "test_data": ["Unit with 0 visible reviews"],
            "test_scenario": "Verify 'No reviews yet' is displayed when no visible reviews exist.",
            "steps": [
                ("Navigate to listing page.", "System retrieves visible reviews for 'New Unit'."),
                ("Observe 'New Unit' card.", "'No reviews yet' is displayed instead of a rating."),
                ("Navigate to 'New Unit' detail page.", "'No reviews yet' displayed. No star rating or count shown."),
            ],
            "flow_color": ALT_FILL,
        },
    ]
    create_workbook("TC-FB-04_DisplayAverageRating.xlsx", sheets)


# ══════════════════════════════════════════════════════════════
# RUN ALL GENERATORS
# ══════════════════════════════════════════════════════════════

if __name__ == "__main__":
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    print("Generating HomeLodge Test Case Excel files...\n")

    # Module 1 – Authentication (6 UCs)
    print("Module 1 – Authentication")
    gen_auth01()
    gen_auth02()
    gen_auth03()
    gen_auth04()
    gen_auth05()
    gen_auth06()

    # Module 2 – Homestay Management (6 UCs)
    print("\nModule 2 – Homestay Management")
    gen_hs01()
    gen_hs02()
    gen_hs03()
    gen_hs04()
    gen_hs05()
    gen_hs06()

    # Module 3 – Booking (7 UCs)
    print("\nModule 3 – Booking")
    gen_bk01()
    gen_bk02()
    gen_bk03()
    gen_bk04()
    gen_bk05()
    gen_bk06()
    gen_bk07()

    # Module 4 – Payment (3 UCs)
    print("\nModule 4 – Payment")
    gen_pay01()
    gen_pay02()
    gen_pay03()

    # Module 5 – Notification (2 UCs)
    print("\nModule 5 – Notification")
    gen_notif01()
    gen_notif02()

    # Module 6 – Chat (2 UCs)
    print("\nModule 6 – Chat")
    gen_chat01()
    gen_chat02()

    # Module 7 – User & Access Management (5 UCs)
    print("\nModule 7 – User & Access Management")
    gen_usr01()
    gen_usr02()
    gen_usr03()
    gen_usr04()
    gen_usr05()

    # Module 8 – System Settings (2 UCs)
    print("\nModule 8 – System Settings")
    gen_set01()
    gen_set02()

    # Module 9 – Audit Logs (2 UCs)
    print("\nModule 9 – Audit Logs")
    gen_audit01()
    gen_audit02()

    # Module 10 – QR Code & Access (5 UCs)
    print("\nModule 10 – QR Code & Access")
    gen_qr01()
    gen_qr02()
    gen_qr03()
    gen_qr04()
    gen_qr05()

    # Module 11 – Reporting & Analytics (3 UCs)
    print("\nModule 11 – Reporting & Analytics")
    gen_rpt01()
    gen_rpt02()
    gen_rpt03()

    # Module 12 – Guest Feedback (4 UCs)
    print("\nModule 12 – Guest Feedback")
    gen_fb01()
    gen_fb02()
    gen_fb03()
    gen_fb04()

    print("\n✅ All 47 test case Excel files generated successfully!")
    print(f"   Output directory: {OUTPUT_DIR}")
