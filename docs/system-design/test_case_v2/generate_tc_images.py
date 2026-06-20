"""
Generate images from Excel test case tabs for TC-AUTH-01_RegisterAccount.xlsx
Creates one image per sheet tab, saved in 'test case img' folder.
"""

import os
import openpyxl
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont
import textwrap

# ── Configuration ──────────────────────────────────────────────────────────────
EXCEL_FILE = r"c:\Users\Admin\Documents\HomeLodge-Doc\docs\system-design\test_case_v2\TC-AUTH-01_RegisterAccount.xlsx"
OUTPUT_DIR = r"c:\Users\Admin\Documents\HomeLodge-Doc\docs\system-design\test_case_v2\test case img"
SCALE = 2          # Retina-like scaling for crisp output
PADDING = 32       # Outer padding in pixels (pre-scale)

# Fonts – fall back to a basic truetype if Calibri not found
FONT_PATHS = [
    r"C:\Windows\Fonts\calibri.ttf",
    r"C:\Windows\Fonts\arial.ttf",
    r"C:\Windows\Fonts\segoeui.ttf",
]
FONT_BOLD_PATHS = [
    r"C:\Windows\Fonts\calibrib.ttf",
    r"C:\Windows\Fonts\arialbd.ttf",
    r"C:\Windows\Fonts\segoeuib.ttf",
]

def load_font(paths, size):
    for p in paths:
        if os.path.exists(p):
            return ImageFont.truetype(p, size)
    return ImageFont.load_default()

# ── Colour helpers ──────────────────────────────────────────────────────────────
def hex_to_rgb(hex_str: str):
    """Convert AARRGGBB or RRGGBB to (R,G,B)."""
    h = hex_str.lstrip('#')
    if len(h) == 8:
        h = h[2:]          # strip alpha channel
    if len(h) == 6:
        return tuple(int(h[i:i+2], 16) for i in (0, 2, 4))
    return (0, 0, 0)

DARK_BLUE  = hex_to_rgb("1F4E79")   # header dark blue
MID_BLUE   = hex_to_rgb("2E75B6")   # sub-header mid blue
LIGHT_BLUE = hex_to_rgb("D6E4F0")   # value cells light blue
WHITE      = (255, 255, 255)
BLACK      = (0, 0, 0)
LIGHT_GRAY = (242, 242, 242)
BORDER     = (180, 180, 180)

# ── Column width mapping ───────────────────────────────────────────────────────
# Excel col widths (in "character units") → pixels
COL_WIDTHS_EXCEL = {1: 5, 2: 28, 3: 18, 4: 13, 5: 32, 6: 13, 7: 22, 8: 13}
UNIT = 7.5   # pixels per Excel character unit

COL_WIDTHS_PX = {col: max(int(w * UNIT), 1) for col, w in COL_WIDTHS_EXCEL.items()}

# Row height mapping
ROW_HEIGHTS_EXCEL = {1: 25, 4: 32, 10: 35, 11: 28}   # special rows
DEFAULT_ROW_H = 15

def row_height_px(row_idx, ws):
    """Return pixel height for a row."""
    rd = ws.row_dimensions.get(row_idx)
    pt = rd.height if (rd and rd.height) else ROW_HEIGHTS_EXCEL.get(row_idx, DEFAULT_ROW_H)
    if pt is None:
        pt = DEFAULT_ROW_H
    return max(int(pt * 1.25), 1)   # approx pt→px

# ── Merged cell helper ────────────────────────────────────────────────────────
def build_merge_map(ws):
    """
    Returns dict: (row, col) → (row_span, col_span)
    and a set of 'slave' cells that should be skipped.
    """
    master = {}
    slaves = set()
    for mr in ws.merged_cells.ranges:
        min_r, min_c = mr.min_row, mr.min_col
        max_r, max_c = mr.max_row, mr.max_col
        master[(min_r, min_c)] = (max_r - min_r + 1, max_c - min_c + 1)
        for r in range(min_r, max_r + 1):
            for c in range(min_c, max_c + 1):
                if (r, c) != (min_r, min_c):
                    slaves.add((r, c))
    return master, slaves

# ── Cell styling from openpyxl ────────────────────────────────────────────────
def get_cell_style(cell):
    """Return (bg_rgb, fg_rgb, is_bold) for a cell."""
    # Background
    fill = cell.fill
    if fill and fill.patternType == 'solid':
        fgc = fill.fgColor
        if fgc and fgc.type == 'rgb' and fgc.rgb not in ('00000000', '000000'):
            bg = hex_to_rgb(fgc.rgb)
        else:
            bg = WHITE
    else:
        bg = WHITE

    # Font colour
    font = cell.font
    if font and font.color:
        fc = font.color
        if fc.type == 'rgb' and fc.rgb not in ('00000000', '000000'):
            fg = hex_to_rgb(fc.rgb)
        else:
            fg = BLACK
    else:
        fg = BLACK

    bold = bool(font and font.bold)
    return bg, fg, bold

# ── Draw a single sheet ────────────────────────────────────────────────────────
def draw_sheet(ws, sheet_name: str, output_path: str):
    num_cols = ws.max_column
    num_rows = ws.max_row

    # Pre-compute column x positions
    col_x = {}
    x = PADDING
    for c in range(1, num_cols + 1):
        col_x[c] = x
        x += COL_WIDTHS_PX.get(c, int(DEFAULT_ROW_H * UNIT))
    total_width = x + PADDING

    # Pre-compute row y positions
    row_y = {}
    y = PADDING
    for r in range(1, num_rows + 1):
        row_y[r] = y
        y += row_height_px(r, ws)
    total_height = y + PADDING

    # Build merge map
    merge_map, slaves = build_merge_map(ws)

    # Create canvas
    img = Image.new("RGB", (total_width * SCALE, total_height * SCALE), WHITE)
    draw = ImageDraw.Draw(img)

    FONT_SIZE = int(9.5 * SCALE)
    font_reg  = load_font(FONT_PATHS, FONT_SIZE)
    font_bold = load_font(FONT_BOLD_PATHS, FONT_SIZE)

    # Draw cells
    for r in range(1, num_rows + 1):
        for c in range(1, num_cols + 1):
            if (r, c) in slaves:
                continue

            cell = ws.cell(row=r, column=c)
            bg, fg, bold = get_cell_style(cell)
            font = font_bold if bold else font_reg

            # Compute bounding rect (handle merged cells)
            row_span, col_span = merge_map.get((r, c), (1, 1))

            x0 = col_x[c] * SCALE
            y0 = row_y[r] * SCALE
            x1 = (col_x[c + col_span - 1] + COL_WIDTHS_PX.get(c + col_span - 1, 40)) * SCALE - 1
            y1 = (row_y[r + row_span - 1] + row_height_px(r + row_span - 1, ws)) * SCALE - 1

            cell_w = x1 - x0
            cell_h = y1 - y0

            # Fill background
            draw.rectangle([x0, y0, x1, y1], fill=bg)

            # Draw border
            draw.rectangle([x0, y0, x1, y1], outline=BORDER)

            # Draw text
            value = cell.value
            if value is not None:
                text = str(value)

                # Estimate chars per line based on cell width
                avg_char_w = FONT_SIZE * 0.55
                chars_per_line = max(int(cell_w / avg_char_w), 5)

                # Wrap text
                lines = []
                for raw_line in text.splitlines():
                    wrapped = textwrap.wrap(raw_line, width=chars_per_line) or ['']
                    lines.extend(wrapped)

                # Vertical: distribute lines, centered in cell
                line_h = FONT_SIZE + int(FONT_SIZE * 0.35)
                total_text_h = len(lines) * line_h
                text_y = y0 + (cell_h - total_text_h) // 2
                if text_y < y0:
                    text_y = y0 + 2 * SCALE

                TEXT_PAD = 3 * SCALE
                for line in lines:
                    draw.text((x0 + TEXT_PAD, text_y), line, fill=fg, font=font)
                    text_y += line_h

    # Save
    img.save(output_path, dpi=(150 * SCALE, 150 * SCALE))
    print(f"  Saved: {output_path}")


# ── Folder of Excel files → images ────────────────────────────────────────────
EXCEL_DIR = os.path.dirname(EXCEL_FILE)   # same folder as the original file


# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    xlsx_files = sorted([
        f for f in os.listdir(EXCEL_DIR)
        if f.lower().endswith(".xlsx")
    ])

    print(f"Found {len(xlsx_files)} Excel file(s) to process.\n")

    for xlsx_name in xlsx_files:
        xlsx_path = os.path.join(EXCEL_DIR, xlsx_name)
        file_base = os.path.splitext(xlsx_name)[0]

        try:
            wb = openpyxl.load_workbook(xlsx_path)
        except Exception as exc:
            print(f"  [SKIP] Could not open {xlsx_name}: {exc}")
            continue

        print(f"Processing: {xlsx_name}  ({len(wb.sheetnames)} sheet(s))")
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            safe_name = sheet_name.replace(" ", "_").replace("/", "-")
            out_path = os.path.join(OUTPUT_DIR, f"{file_base}_{safe_name}.png")

            # Skip if already generated
            if os.path.exists(out_path):
                print(f"  [SKIP] Already exists: {os.path.basename(out_path)}")
                continue

            print(f"  -> {os.path.basename(out_path)}")
            try:
                draw_sheet(ws, sheet_name, out_path)
            except Exception as exc:
                print(f"  [ERROR] {sheet_name}: {exc}")

        wb.close()

    print("\nAll done! Images saved to:", OUTPUT_DIR)


if __name__ == "__main__":
    main()
