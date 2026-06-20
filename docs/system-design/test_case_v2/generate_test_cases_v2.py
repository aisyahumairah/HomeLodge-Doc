"""
HomeLodge Test Case Generator v2
Generates one Excel file per use case with tabs for each flow type.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

OUTPUT_DIR = r"c:\Users\Admin\Documents\HomeLodge-Doc\docs\system-design\test_case_v2"
CREATED_BY = "Aisyah"
VERSION = "1.0"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
VALUE_FILL  = PatternFill("solid", fgColor="D6E4F0")
TABLE_HEAD  = PatternFill("solid", fgColor="2E75B6")
ALT_ROW     = PatternFill("solid", fgColor="EBF3FB")
WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")
NORMAL_FILL = PatternFill("solid", fgColor="E2EFDA")
ALT_FILL    = PatternFill("solid", fgColor="FFF2CC")
EXCEP_FILL  = PatternFill("solid", fgColor="FCE4D6")

THIN  = Side(style="thin", color="AAAAAA")
INNER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

LABEL_F = Font(name="Calibri", color="FFFFFF", bold=True, size=10)
VALUE_F = Font(name="Calibri", color="1F4E79", size=10)
BODY_F  = Font(name="Calibri", size=10)
HEAD_F  = Font(name="Calibri", color="FFFFFF", bold=True, size=10)
TITLE_F = Font(name="Calibri", bold=True, size=11, color="1F4E79")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def put(ws, row, col, val, font, fill, align=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = font
    c.fill = fill
    c.alignment = align or LEFT
    c.border = INNER
    return c


def build_sheet(ws, tc_id, tc_desc, prereqs, data, scenario, steps, color=None):
    for col, w in zip("ABCDEFGH", [5, 28, 18, 18, 32, 32, 22, 22]):
        ws.column_dimensions[col].width = w

    # Row 1
    put(ws,1,1,"Test Case ID",          LABEL_F, HEADER_FILL, Alignment(horizontal="left", vertical="center"))
    ws.merge_cells("A1:B1")
    put(ws,1,3, tc_id,                  VALUE_F, VALUE_FILL)
    ws.merge_cells("C1:D1")
    put(ws,1,5,"Test Case Description", LABEL_F, HEADER_FILL, Alignment(horizontal="left", vertical="center"))
    ws.merge_cells("E1:F1")
    put(ws,1,7, tc_desc,                VALUE_F, VALUE_FILL)
    ws.merge_cells("G1:H1")
    ws.row_dimensions[1].height = 25

    # Row 2
    put(ws,2,1,"Created By",   LABEL_F, HEADER_FILL, Alignment(horizontal="left", vertical="center"))
    ws.merge_cells("A2:B2")
    put(ws,2,3, CREATED_BY,   VALUE_F, VALUE_FILL, Alignment(horizontal="left", vertical="center"))
    ws.merge_cells("C2:D2")
    put(ws,2,5,"Reviewed By",  LABEL_F, HEADER_FILL, Alignment(horizontal="left", vertical="center"))
    put(ws,2,6,"",             VALUE_F, VALUE_FILL,  Alignment(horizontal="left", vertical="center"))
    put(ws,2,7,"Version",      LABEL_F, HEADER_FILL, Alignment(horizontal="left", vertical="center"))
    put(ws,2,8, VERSION,       VALUE_F, VALUE_FILL,  Alignment(horizontal="left", vertical="center"))

    # Row 3
    put(ws,3,1,"QA Tester's Log", LABEL_F, HEADER_FILL, Alignment(horizontal="left", vertical="center"))
    ws.merge_cells("A3:B3")
    put(ws,3,3,"",                VALUE_F, VALUE_FILL)
    ws.merge_cells("C3:H3")

    # Row 4
    put(ws,4,1,"Tester's Name",                     LABEL_F, HEADER_FILL, Alignment(horizontal="left", vertical="center"))
    ws.merge_cells("A4:B4")
    put(ws,4,3,"",                                   VALUE_F, VALUE_FILL, Alignment(horizontal="left", vertical="center"))
    ws.merge_cells("C4:D4")
    put(ws,4,5,"Date Tested",                        LABEL_F, HEADER_FILL, Alignment(horizontal="left", vertical="center"))
    put(ws,4,6,"",                                   VALUE_F, VALUE_FILL, Alignment(horizontal="left", vertical="center"))
    put(ws,4,7,"Test Case (Pass/Fail/Not Executed)", LABEL_F, HEADER_FILL, Alignment(horizontal="left", vertical="center", wrap_text=True))
    put(ws,4,8,"",                                   VALUE_F, VALUE_FILL, Alignment(horizontal="left", vertical="center"))
    ws.row_dimensions[4].height = 32

    # Row 5 – section headers
    put(ws,5,1,"S #",            HEAD_F, TABLE_HEAD, CENTER)
    put(ws,5,2,"Prerequisites:", HEAD_F, TABLE_HEAD, Alignment(horizontal="left", vertical="center"))
    ws.merge_cells("B5:E5")
    put(ws,5,6,"S #",            HEAD_F, TABLE_HEAD, CENTER)
    put(ws,5,7,"Test Data",      HEAD_F, TABLE_HEAD, Alignment(horizontal="left", vertical="center"))
    ws.merge_cells("G5:H5")

    max_r = max(len(prereqs), len(data), 3)
    for i in range(max_r):
        r   = 6 + i
        f   = ALT_ROW if i % 2 == 0 else WHITE_FILL
        put(ws, r, 1, i+1, BODY_F, f, CENTER)
        put(ws, r, 2, prereqs[i] if i < len(prereqs) else None, BODY_F, f)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        put(ws, r, 6, i+1, BODY_F, f, CENTER)
        put(ws, r, 7, data[i] if i < len(data) else None, BODY_F, f)
        ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=8)

    sr = 6 + max_r
    put(ws, sr, 1, "Test Scenario", LABEL_F, HEADER_FILL, Alignment(horizontal="left", vertical="center"))
    ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=2)
    put(ws, sr, 3, scenario, TITLE_F, VALUE_FILL)
    ws.merge_cells(start_row=sr, start_column=3, end_row=sr, end_column=8)
    ws.row_dimensions[sr].height = 35

    hr = sr + 1
    put(ws, hr, 1, "Step #",                                     HEAD_F, TABLE_HEAD, CENTER)
    put(ws, hr, 2, "Step Details",                               HEAD_F, TABLE_HEAD, Alignment(horizontal="left", vertical="center"))
    ws.merge_cells(start_row=hr, start_column=2, end_row=hr, end_column=4)
    put(ws, hr, 5, "Expected Results",                           HEAD_F, TABLE_HEAD, Alignment(horizontal="left", vertical="center"))
    ws.merge_cells(start_row=hr, start_column=5, end_row=hr, end_column=6)
    put(ws, hr, 7, "Actual Results",                             HEAD_F, TABLE_HEAD, Alignment(horizontal="left", vertical="center"))
    put(ws, hr, 8, "Pass / Fail / Not executed / Suspended",     HEAD_F, TABLE_HEAD, CENTER)
    ws.row_dimensions[hr].height = 28

    for i, (detail, expected) in enumerate(steps):
        r = hr + 1 + i
        f = ALT_ROW if i % 2 == 0 else WHITE_FILL
        ws.row_dimensions[r].height = 50
        put(ws, r, 1, i+1,     BODY_F, f, CENTER)
        put(ws, r, 2, detail,   BODY_F, f)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        put(ws, r, 5, expected, BODY_F, f)
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
        put(ws, r, 7, "",      BODY_F, WHITE_FILL)
        put(ws, r, 8, "",      BODY_F, WHITE_FILL, CENTER)

    if color:
        ws.sheet_properties.tabColor = color.fgColor


def make_wb(filename, sheets_data):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sd in sheets_data:
        ws = wb.create_sheet(title=sd["tab"])
        build_sheet(ws, sd["id"], sd["desc"],
                    sd["prereqs"], sd["data"],
                    sd["scenario"], sd["steps"],
                    sd.get("color"))
    out = os.path.join(OUTPUT_DIR, filename)
    wb.save(out)
    print(f"  Saved: {filename}")


# ══════════════════════════════════════════════════
#  MODULE 1 – AUTHENTICATION
# ══════════════════════════════════════════════════
def gen_auth01():
    make_wb("TC-AUTH-01_RegisterAccount.xlsx", [
        {"tab": "Normal Flow", "id": "TC-AUTH-01_NF",
         "desc": "Register Account – Normal Flow (Successful Registration)",
         "prereqs": ["Email 'newguest@example.com' is NOT registered.",
                     "Registration page is publicly accessible."],
         "data": ["Name: 'John Doe'", "Email: 'newguest@example.com'",
                  "Password: 'Pass@1234'", "Confirm Password: 'Pass@1234'"],
         "scenario": "Verify a new visitor can successfully register a HomeLodge account with valid details and is redirected to the sign-in page.",
         "steps": [
             ("Open the registration page.", "Registration form displayed with Name, Email, Password, Confirm Password fields."),
             ("Enter all fields with the test data.", "Fields accept input. Password strength indicator shows all rules met."),
             ("Click 'Register'.", "Success message shown. User redirected to sign-in page."),
             ("Check database.", "Account exists with email 'newguest@example.com', role 'Guest', status 'active'."),
         ], "color": NORMAL_FILL},
        {"tab": "Alternate Flow", "id": "TC-AUTH-01_AF",
         "desc": "Register Account – Alternate Flows (Validation Failures)",
         "prereqs": ["Email 'existing@example.com' is already registered (for A1).",
                     "Registration page is accessible."],
         "data": ["A1 Email: 'existing@example.com'", "A2 Password: 'weak'",
                  "A3 Confirm: 'Different1!'", "A4 Email: 'not-an-email'"],
         "scenario": "Verify the system handles all alternate validation scenarios for registration correctly.",
         "steps": [
             ("(A1) Enter 'existing@example.com' with valid other fields. Click Register.", "Email field highlighted. Error: 'Email already in use.' Suggestion to Sign In or use Forgot Password shown. No account created."),
             ("(A2) Enter Password: 'weak'. Click Register.", "Password strength rules not met are shown. Submission blocked."),
             ("(A3) Enter Password 'Pass@1234' and Confirm 'Different1!'. Click Register.", "Confirm password field highlighted. Error: 'Passwords do not match.' Submission blocked."),
             ("(A4) Enter Email: 'not-an-email'. Click Register.", "Email field highlighted. Invalid email format error shown. Submission blocked."),
         ], "color": ALT_FILL},
        {"tab": "Exception Flow", "id": "TC-AUTH-01_EF",
         "desc": "Register Account – Exception Flow (System Error)",
         "prereqs": ["All form fields valid.", "System/DB error simulated during account creation."],
         "data": ["Name: 'Jane Doe'", "Email: 'jane@example.com'",
                  "Password: 'Pass@1234'", "Confirm Password: 'Pass@1234'"],
         "scenario": "Verify that a system error during registration shows an error message and does NOT create a partial account.",
         "steps": [
             ("Fill in all valid registration details.", "Fields accept input."),
             ("Click 'Register' while DB error is simulated.", "Form submitted."),
             ("Observe system response.", "Error shown: 'Registration failed. Please try again.' No account record created in DB."),
         ], "color": EXCEP_FILL},
    ])


def gen_auth02():
    make_wb("TC-AUTH-02_Login.xlsx", [
        {"tab": "Normal Flow – Email", "id": "TC-AUTH-02_NF_Email",
         "desc": "Login – Normal Flow (Email & Password)",
         "prereqs": ["User 'user@example.com' / 'Pass@1234' exists and is active.",
                     "No forced password change flag set."],
         "data": ["Email: 'user@example.com'", "Password: 'Pass@1234'"],
         "scenario": "Verify a registered user can sign in with valid email and password and is redirected to dashboard.",
         "steps": [
             ("Navigate to sign-in page.", "Sign-in form displayed."),
             ("Enter email and password. Click Sign In.", "Credentials verified. Sign-in timestamp recorded."),
             ("Observe redirect.", "User taken to their dashboard."),
         ], "color": NORMAL_FILL},
        {"tab": "Normal Flow – Google", "id": "TC-AUTH-02_NF_Google",
         "desc": "Login – Normal Flow (Google SSO)",
         "prereqs": ["Google account 'google@example.com' has a HomeLodge account.",
                     "Google account 'newgoogle@example.com' has no HomeLodge account."],
         "data": ["Google Account A: 'google@example.com' (existing)", "Google Account B: 'newgoogle@example.com' (new)"],
         "scenario": "Verify Google SSO for both existing and new accounts.",
         "steps": [
             ("(Scenario A) Click 'Continue with Google'. Authenticate with Account A.", "Google returns name/email. Existing account found and Google identity linked. User redirected to dashboard."),
             ("(Scenario B) Click 'Continue with Google'. Authenticate with Account B.", "Google returns name/email. No account found. New 'Guest' account created. User redirected to dashboard."),
             ("(Scenario B) Verify new account.", "Account exists with email 'newgoogle@example.com', role 'Guest', status 'active'."),
         ], "color": NORMAL_FILL},
        {"tab": "Normal Flow – Force Chg", "id": "TC-AUTH-02_NF_Force",
         "desc": "Login – Normal Flow (Forced Password Change)",
         "prereqs": ["User 'forcechange@example.com' has forced change flag active.",
                     "Temporary password is 'Temp@1234'."],
         "data": ["Email: 'forcechange@example.com'", "Temp Password: 'Temp@1234'"],
         "scenario": "Verify a user with forced change flag is redirected to change password page and cannot access other pages.",
         "steps": [
             ("Enter email and temp password. Click Sign In.", "Credentials verified. Forced change flag detected."),
             ("Observe redirect.", "User redirected to 'Change Your Password' page. All other pages blocked."),
             ("Attempt to navigate to /dashboard via URL.", "System redirects back to Change Password page."),
         ], "color": NORMAL_FILL},
        {"tab": "Alternate Flow", "id": "TC-AUTH-02_AF",
         "desc": "Login – Alternate Flows (Wrong Creds, Deactivated, Locked, Google Cancelled)",
         "prereqs": ["'deactivated@example.com' account is deactivated.",
                     "'locked@example.com' account is locked.",
                     "'valid@example.com' exists with password 'Pass@1234'."],
         "data": ["A1: 'valid@example.com' / 'WrongPass!'", "A2: 'deactivated@example.com'",
                  "A3: 'locked@example.com'", "A4: Cancel Google consent screen"],
         "scenario": "Verify error handling for invalid credentials, deactivated, locked accounts, and cancelled Google sign-in.",
         "steps": [
             ("(A1) Enter valid email with wrong password. Click Sign In.", "Generic error: 'Invalid email or password.' No session created."),
             ("(A2) Enter 'deactivated@example.com'. Click Sign In.", "Message: 'Your account has been deactivated. Please contact support.' No session."),
             ("(A3) Enter 'locked@example.com'. Click Sign In.", "Lockout message shown with estimated unlock time. Option to reset password to unlock."),
             ("(A4) Click 'Continue with Google'. Cancel on Google consent screen.", "Returned to sign-in page with message that Google sign-in was not completed."),
         ], "color": ALT_FILL},
        {"tab": "Exception Flow", "id": "TC-AUTH-02_EF",
         "desc": "Login – Exception Flows (System Down, Google Service Down)",
         "prereqs": ["E1: System/DB temporarily unavailable.", "E2: Google OAuth service unreachable."],
         "data": ["E1: Valid credentials", "E2: Google Sign-In button click"],
         "scenario": "Verify graceful error handling when system or Google service is unavailable.",
         "steps": [
             ("(E1) Attempt email/password login while system is down.", "Error message shown. No session created."),
             ("(E2) Click 'Continue with Google' while Google service is unreachable.", "Error shown. User returned to sign-in page. No session."),
         ], "color": EXCEP_FILL},
    ])


def gen_auth03():
    make_wb("TC-AUTH-03_Logout.xlsx", [
        {"tab": "Normal Flow", "id": "TC-AUTH-03_NF",
         "desc": "Logout – Normal Flow (Successful Sign-Out)",
         "prereqs": ["User is currently signed in."],
         "data": ["Logged-in user session active."],
         "scenario": "Verify clicking Logout ends the session, redirects to sign-in, and protected pages are inaccessible.",
         "steps": [
             ("While signed in, click 'Logout' in navigation menu.", "Logout request sent."),
             ("Observe response.", "Session ended. Remember Me tokens cleared. Redirected to sign-in page."),
             ("Press browser back button.", "Sign-in page shown, not protected content."),
             ("Navigate directly to /dashboard.", "Redirected to sign-in. Protected content not accessible."),
         ], "color": NORMAL_FILL},
        {"tab": "Exception Flow", "id": "TC-AUTH-03_EF",
         "desc": "Logout – Exception Flow (Session Already Expired)",
         "prereqs": ["User's session has already expired."],
         "data": ["Expired session token."],
         "scenario": "Verify clicking Logout with an expired session still redirects to sign-in without error.",
         "steps": [
             ("With expired session, click 'Logout'.", "Logout request sent."),
             ("Observe response.", "Redirected to sign-in page. No error message shown."),
         ], "color": EXCEP_FILL},
    ])


def gen_auth04():
    make_wb("TC-AUTH-04_ForgotPassword.xlsx", [
        {"tab": "Normal Flow", "id": "TC-AUTH-04_NF",
         "desc": "Forgot Password – Normal Flow (Successful Reset)",
         "prereqs": ["Account 'resetme@example.com' is registered.", "Email service configured and working."],
         "data": ["Email: 'resetme@example.com'", "New Password: 'NewPass@123'", "Confirm: 'NewPass@123'"],
         "scenario": "Verify a registered user can reset password via email link and the link is invalidated after use.",
         "steps": [
             ("Click 'Forgot Password' on sign-in page.", "Forgot Password page displayed."),
             ("Enter 'resetme@example.com'. Click Submit.", "Generic message: 'If an account exists, a reset link has been sent.'"),
             ("Open reset link from email.", "Password reset form displayed."),
             ("Enter new password 'NewPass@123' and confirm. Click Save.", "New password saved. Link marked as used."),
             ("Observe redirect.", "Redirected to sign-in page with success message."),
             ("Attempt to use same reset link again.", "Error: 'This reset link is invalid or has expired.'"),
             ("Sign in with 'NewPass@123'.", "Login succeeds."),
         ], "color": NORMAL_FILL},
        {"tab": "Alternate Flow", "id": "TC-AUTH-04_AF",
         "desc": "Forgot Password – Alternate Flows",
         "prereqs": ["A1: 'notfound@example.com' is NOT registered.", "A2: Reset link already expired/used."],
         "data": ["A1 Email: 'notfound@example.com'", "A2: Expired reset link URL", "A3 New Password: 'weak'"],
         "scenario": "Verify privacy protection for unfound email, expired link error, and weak password rejection.",
         "steps": [
             ("(A1) Enter 'notfound@example.com'. Click Submit.", "Same generic message shown. Privacy protected. No link sent."),
             ("(A2) Open expired or already-used reset link.", "Error: 'This reset link is invalid or has expired.' User prompted to request new one."),
             ("(A3) On valid reset form, enter 'weak' as new password. Click Save.", "Strength rules shown. Submission blocked."),
         ], "color": ALT_FILL},
        {"tab": "Exception Flow", "id": "TC-AUTH-04_EF",
         "desc": "Forgot Password – Exception Flow (Email Delivery Failure)",
         "prereqs": ["Account 'resetme@example.com' registered.", "Email service simulated to fail."],
         "data": ["Email: 'resetme@example.com'"],
         "scenario": "Verify email delivery failure does not expose errors and user can retry.",
         "steps": [
             ("Enter 'resetme@example.com'. Click Submit while email service is down.", "Request submitted."),
             ("Observe response.", "Generic message displayed. No internal error exposed. Reset link record created. User can try again."),
         ], "color": EXCEP_FILL},
    ])


def gen_auth05():
    make_wb("TC-AUTH-05_ViewUpdateProfile.xlsx", [
        {"tab": "Normal Flow", "id": "TC-AUTH-05_NF",
         "desc": "View / Update Profile – Normal Flow",
         "prereqs": ["User is signed in.", "Profile exists with name, email, phone, photo."],
         "data": ["New Phone: '+60123456789'", "New Photo: valid JPEG <2MB",
                  "Current Password: 'Pass@1234'", "New Password: 'NewPass@5678'"],
         "scenario": "Verify a signed-in user can update profile info and change their password successfully.",
         "steps": [
             ("Navigate to Profile page.", "Current profile data displayed: name, email, phone, photo."),
             ("Update phone to '+60123456789'. Upload new JPEG photo. Click Save.", "Validation passes. Profile saved. Confirmation: 'Profile updated successfully.'"),
             ("Verify updated info.", "Phone and photo updated in UI."),
             ("Click 'Change Password'. Enter current 'Pass@1234', new 'NewPass@5678', confirm 'NewPass@5678'.", "Submitted for validation."),
             ("Observe response.", "New password saved. Confirmation: 'Password changed successfully.'"),
             ("Sign out and sign back in with 'NewPass@5678'.", "Login succeeds."),
         ], "color": NORMAL_FILL},
        {"tab": "Alternate Flow", "id": "TC-AUTH-05_AF",
         "desc": "View / Update Profile – Alternate Flows (Validation Errors)",
         "prereqs": ["User is signed in."],
         "data": ["A1 Phone: 'abc12345'", "A2 Photo: .exe file",
                  "A3 New Password: 'weak'", "A4 Current Password: 'WrongPass!'", "A5 Confirm: 'Different@1'"],
         "scenario": "Verify validation errors for invalid phone, invalid photo, and password rule violations.",
         "steps": [
             ("(A1) Enter phone 'abc12345'. Click Save.", "Phone field highlighted. Error: invalid phone number format."),
             ("(A2) Upload .exe file as photo. Click Save.", "Error: 'Invalid file type. Please upload a valid image.'"),
             ("(A3) In Change Password, enter new password 'weak'. Submit.", "Failing strength rules displayed. Blocked."),
             ("(A4) Enter wrong current password 'WrongPass!'. Submit.", "Error: 'Current password does not match.'"),
             ("(A5) Enter new 'Pass@1234', confirm 'Different@1'. Submit.", "Error: 'Passwords do not match.'"),
         ], "color": ALT_FILL},
        {"tab": "Exception Flow", "id": "TC-AUTH-05_EF",
         "desc": "View / Update Profile – Exception Flows",
         "prereqs": ["User signed in.", "Photo upload or save error simulated."],
         "data": ["E1: Photo upload failure simulated", "E2: System error on Save"],
         "scenario": "Verify upload or system errors roll back changes and notify the user.",
         "steps": [
             ("(E1) Upload valid photo. Simulate upload failure. Click Save.", "Photo upload fails. Changes rolled back. Error notifies user. Existing photo unchanged."),
             ("(E2) Make profile changes. Simulate system error on Save.", "Changes rolled back. Error notification shown. No partial data saved."),
         ], "color": EXCEP_FILL},
    ])


def gen_auth06():
    make_wb("TC-AUTH-06_ForceChangePassword.xlsx", [
        {"tab": "Normal Flow", "id": "TC-AUTH-06_NF",
         "desc": "Force Change Password – Normal Flow",
         "prereqs": ["Admin reset user password to 'Temp@1234'.", "Forced change flag is set."],
         "data": ["Temp Password: 'Temp@1234'", "New Password: 'MyNewPass@99'", "Confirm: 'MyNewPass@99'"],
         "scenario": "Verify user with forced-change flag must change password before accessing system, and flag clears after.",
         "steps": [
             ("Sign in with temp password 'Temp@1234'.", "Credentials verified. Forced change flag detected."),
             ("Observe redirect.", "User taken to 'Change Your Password' page. All other pages blocked."),
             ("Enter new password 'MyNewPass@99' and confirm. Click Save.", "New password submitted."),
             ("Observe response.", "New password saved. Flag cleared. User redirected to dashboard."),
             ("Verify access restored.", "All pages accessible. Sign-in with 'MyNewPass@99' succeeds."),
         ], "color": NORMAL_FILL},
        {"tab": "Alternate Flow", "id": "TC-AUTH-06_AF",
         "desc": "Force Change Password – Alternate Flows",
         "prereqs": ["User is on forced Change Password page.", "Temp password is 'Temp@1234'."],
         "data": ["A1 New Password: 'Temp@1234' (same as temp)", "A2 New Password: 'weak'"],
         "scenario": "Verify same temp password and weak passwords are rejected.",
         "steps": [
             ("(A1) Enter 'Temp@1234' as new password. Click Save.", "Error: 'Please choose a different password.' Same temp not accepted."),
             ("(A2) Enter 'weak' as new password. Click Save.", "Strength rules highlighted. Submission blocked."),
         ], "color": ALT_FILL},
        {"tab": "Exception Flow", "id": "TC-AUTH-06_EF",
         "desc": "Force Change Password – Exception Flow (Navigation Bypass)",
         "prereqs": ["Forced-change flag is active."],
         "data": ["Direct URL: /dashboard"],
         "scenario": "Verify user cannot bypass Change Password page via direct URL navigation.",
         "steps": [
             ("While on Change Password page, navigate to /dashboard via URL bar.", "Navigation attempted."),
             ("Observe response.", "System redirects back to Change Password page. Dashboard inaccessible."),
         ], "color": EXCEP_FILL},
    ])


# ══════════════════════════════════════════════════
#  MODULE 2 – HOMESTAY MANAGEMENT
# ══════════════════════════════════════════════════
def gen_hs01():
    make_wb("TC-HS-01_BrowseHomestayUnits.xlsx", [
        {"tab": "Normal Flow", "id": "TC-HS-01_NF",
         "desc": "Browse Homestay Units – Normal Flow",
         "prereqs": ["At least 2 active units exist.", "Guest may or may not be signed in."],
         "data": ["Guest (signed out)", "2+ active units in DB"],
         "scenario": "Verify guest can view all active unit cards and navigate to a unit detail page.",
         "steps": [
             ("Open homestay listing page.", "All active units displayed as cards: name, photo, price/night, location, avg rating."),
             ("Scroll through listing.", "All active units visible. Inactive units NOT shown."),
             ("Click on any unit card.", "Navigated to that unit's detail page (UC-HS-02)."),
         ], "color": NORMAL_FILL},
        {"tab": "Alternate Flow", "id": "TC-HS-01_AF",
         "desc": "Browse Homestay Units – Alternate Flow (No Active Units)",
         "prereqs": ["No active units exist."],
         "data": ["All units: inactive or none"],
         "scenario": "Verify informational message shown when no active units available.",
         "steps": [
             ("Open listing page.", "Page loads."),
             ("Observe displayed content.", "Message: 'No homestay units are currently available.' No cards shown."),
         ], "color": ALT_FILL},
    ])


def gen_hs02():
    make_wb("TC-HS-02_ViewUnitDetails.xlsx", [
        {"tab": "Normal Flow", "id": "TC-HS-02_NF",
         "desc": "View Unit Details & Availability – Normal Flow",
         "prereqs": ["Active unit 'Cozy Cottage' with photos, description, pricing, bookings."],
         "data": ["Unit: 'Cozy Cottage'", "Has available, booked, and blocked dates"],
         "scenario": "Verify the unit detail page shows complete info including colour-coded availability calendar.",
         "steps": [
             ("Click 'Cozy Cottage' from listing page.", "Unit detail page loaded."),
             ("Review displayed information.", "Shown: name, photo gallery, description, location, base price/night, deposit, check-in/out times, house rules, avg rating, reviews."),
             ("Observe availability calendar.", "Colour-coded dates: Available, Booked, Temporarily Held, Blocked."),
             ("Click 'Book Now'.", "Navigated to booking form (UC-BK-01)."),
         ], "color": NORMAL_FILL},
    ])


def gen_hs03():
    make_wb("TC-HS-03_CreateHomestayUnit.xlsx", [
        {"tab": "Normal Flow", "id": "TC-HS-03_NF",
         "desc": "Create Homestay Unit – Normal Flow",
         "prereqs": ["Admin signed in with create permission.", "Default policies: No Pets, No Durians, No Smoking."],
         "data": ["Name: 'Sunset Villa'", "Location: 'Langkawi'", "Price: RM 250/night",
                  "Deposit: RM 100", "Check-in: 2:00 PM", "Check-out: 12:00 PM", "Photo: valid JPEG"],
         "scenario": "Verify admin can create a new unit and default house rules are auto-applied.",
         "steps": [
             ("Navigate to Homestay Management > Create New Unit.", "Unit creation form displayed."),
             ("Fill all required fields: name, description, location, price, deposit, check-in/out times.", "All fields accept input."),
             ("Upload valid JPEG photo.", "Photo accepted."),
             ("Optionally set custom extension window. Click Save.", "Form submitted."),
             ("Observe response.", "Success message. Unit created with 'active' status. Visible on guest listing."),
             ("Verify default house rules.", "Unit has: No Pets, No Durians, No Smoking rules attached."),
             ("Verify audit log.", "Audit log entry recorded for the unit creation."),
         ], "color": NORMAL_FILL},
        {"tab": "Alternate Flow", "id": "TC-HS-03_AF",
         "desc": "Create Homestay Unit – Alternate Flows",
         "prereqs": ["Admin signed in with create permission."],
         "data": ["A1: Leave 'Name' field empty", "A2: Upload .exe file"],
         "scenario": "Verify validation for missing required fields or invalid photo type.",
         "steps": [
             ("(A1) Fill all fields except Name. Click Save.", "Name field highlighted. Error: required. No unit created."),
             ("(A2) Upload .exe file as photo. Click Save.", "File rejected. Error: invalid format. No unit created."),
         ], "color": ALT_FILL},
        {"tab": "Exception Flow", "id": "TC-HS-03_EF",
         "desc": "Create Homestay Unit – Exception Flow (Photo Upload Failure)",
         "prereqs": ["All fields valid. Photo upload simulated to fail."],
         "data": ["All valid unit fields", "Photo upload: simulated failure"],
         "scenario": "Verify photo upload failure does not prevent unit creation; unit saved with warning.",
         "steps": [
             ("Fill all fields, upload photo (simulated to fail). Click Save.", "Submitted."),
             ("Observe response.", "Unit record saved successfully. Warning: 'Unit created, but photo upload failed. Upload photos by editing the unit.' Unit visible on listing (without photo)."),
         ], "color": EXCEP_FILL},
    ])


def gen_hs04():
    make_wb("TC-HS-04_EditHomestayUnit.xlsx", [
        {"tab": "Normal Flow", "id": "TC-HS-04_NF",
         "desc": "Edit Homestay Unit – Normal Flow",
         "prereqs": ["Unit 'Sunset Villa' exists. Admin signed in with edit permission."],
         "data": ["Updated Name: 'Sunset Villa Premium'", "Updated Price: RM 300/night", "New photo: valid JPEG"],
         "scenario": "Verify admin can edit unit details and changes reflect immediately.",
         "steps": [
             ("Navigate to Homestay Management > select 'Sunset Villa' > click 'Edit'.", "Pre-filled edit form displayed."),
             ("Change Name to 'Sunset Villa Premium', Price to RM 300. Upload new photo.", "Fields accept changes."),
             ("Click Save.", "Changes submitted."),
             ("Observe response.", "Success message. Updated values live on guest listing immediately."),
             ("Verify audit log.", "Audit log entry recorded."),
         ], "color": NORMAL_FILL},
        {"tab": "Alternate Flow", "id": "TC-HS-04_AF",
         "desc": "Edit Homestay Unit – Alternate Flow (Invalid Input)",
         "prereqs": ["Unit exists. Admin signed in."],
         "data": ["Price: '-50' (negative, invalid)"],
         "scenario": "Verify invalid input is rejected and entered values are preserved.",
         "steps": [
             ("Open edit form. Enter price '-50'. Click Save.", "Submitted."),
             ("Observe response.", "Problem field highlighted. Error shown. Values preserved. Record NOT saved."),
         ], "color": ALT_FILL},
    ])


def gen_hs05():
    make_wb("TC-HS-05_DeactivateDeleteUnit.xlsx", [
        {"tab": "Normal Flow", "id": "TC-HS-05_NF",
         "desc": "Deactivate / Delete Unit – Normal Flow (No Conflicts)",
         "prereqs": ["Unit 'Sunset Villa' with no upcoming confirmed bookings.", "Admin signed in."],
         "data": ["Unit: 'Sunset Villa'"],
         "scenario": "Verify admin can deactivate a unit with no future bookings.",
         "steps": [
             ("Navigate to Homestay Management > open 'Sunset Villa' > click 'Deactivate'.", "System checks for upcoming confirmed bookings (none found)."),
             ("Confirmation dialog shown.", "Confirmation: 'Are you sure you want to deactivate this unit?'"),
             ("Click Confirm.", "Deactivation submitted."),
             ("Observe response.", "Success. Unit status = 'inactive'. Removed from guest listing. Data retained."),
             ("Verify audit log.", "Audit log entry recorded."),
         ], "color": NORMAL_FILL},
        {"tab": "Alternate Flow", "id": "TC-HS-05_AF",
         "desc": "Deactivate / Delete Unit – Alternate Flows",
         "prereqs": ["'Cottage A' has an upcoming confirmed booking."],
         "data": ["A1: Unit with future confirmed booking", "A2: Admin clicks Cancel on dialog"],
         "scenario": "Verify deactivation blocked when future bookings exist; cancelling dialog makes no changes.",
         "steps": [
             ("(A1) Attempt to deactivate 'Cottage A' (has future booking).", "Action blocked. Warning listing conflicting bookings. Must cancel/reassign bookings first."),
             ("(A2) On a conflict-free unit, click Deactivate, then Cancel on dialog.", "Dialog dismissed. No changes. Unit remains active."),
         ], "color": ALT_FILL},
        {"tab": "Exception Flow", "id": "TC-HS-05_EF",
         "desc": "Deactivate / Delete Unit – Exception Flow (Cancel Confirmation)",
         "prereqs": ["Unit exists. No upcoming bookings."],
         "data": ["Admin clicks Cancel on confirmation dialog"],
         "scenario": "Verify no changes on cancel.",
         "steps": [
             ("Click 'Deactivate'. When dialog appears, click 'Cancel'.", "Dialog closes. No changes. Unit remains active."),
         ], "color": EXCEP_FILL},
    ])


def gen_hs06():
    make_wb("TC-HS-06_ViewAllUnitsList.xlsx", [
        {"tab": "Normal Flow", "id": "TC-HS-06_NF",
         "desc": "View All Units List – Normal Flow",
         "prereqs": ["Admin signed in with homestay management access.", "Multiple units (active and inactive) exist."],
         "data": ["Filter: Status = 'Active'"],
         "scenario": "Verify admin can view all units including inactive, and apply status filter.",
         "steps": [
             ("Navigate to Homestay Management section.", "System retrieves all unit records."),
             ("Observe displayed list.", "Shows: name, status (active/inactive), upcoming confirmed bookings count, base price, action buttons."),
             ("Apply status filter 'Active'.", "List shows only active units."),
             ("Clear filter.", "Full list restored."),
         ], "color": NORMAL_FILL},
    ])


# ══════════════════════════════════════════════════
#  MODULE 3 – BOOKING
# ══════════════════════════════════════════════════
def gen_bk01():
    make_wb("TC-BK-01_ViewAvailabilitySelectDates.xlsx", [
        {"tab": "Normal Flow", "id": "TC-BK-01_NF",
         "desc": "View Availability & Select Dates – Normal Flow",
         "prereqs": ["Guest on unit detail/booking page.", "Dates 2026-07-10 to 2026-07-13 are all available."],
         "data": ["Check-in: 2026-07-10", "Check-out: 2026-07-13"],
         "scenario": "Verify guest can view colour-coded calendar, select available dates, and see booking summary.",
         "steps": [
             ("Open unit detail/booking form.", "Colour-coded availability calendar displayed."),
             ("Click check-in date 2026-07-10.", "Check-in date highlighted."),
             ("Click check-out date 2026-07-13.", "Real-time availability check performed."),
             ("Observe result.", "All dates available. Booking summary: 3 nights, estimated cost displayed. Guest can proceed."),
         ], "color": NORMAL_FILL},
        {"tab": "Alternate Flow", "id": "TC-BK-01_AF",
         "desc": "View Availability & Select Dates – Alternate Flows",
         "prereqs": ["Date 2026-07-11 is booked. Unit has minimum 2-night stay."],
         "data": ["A1: 2026-07-10 to 2026-07-13 (2026-07-11 booked)", "A2: Check-out before check-in", "A3: Same-day (0 nights)"],
         "scenario": "Verify conflict, invalid range, and minimum stay validation.",
         "steps": [
             ("(A1) Select 2026-07-10 to 2026-07-13 (2026-07-11 is booked).", "Conflict highlighted. 'Selected dates not available.' Must re-select."),
             ("(A2) Select check-out 2026-07-10 before check-in 2026-07-13.", "'Check-out must be after check-in.' Must re-select."),
             ("(A3) Select same day for check-in and check-out.", "Minimum stay message displayed. Summary not updated."),
         ], "color": ALT_FILL},
    ])


def gen_bk02():
    make_wb("TC-BK-02_SubmitBooking.xlsx", [
        {"tab": "Normal Flow", "id": "TC-BK-02_NF",
         "desc": "Submit Booking – Normal Flow",
         "prereqs": ["Guest signed in.", "Dates 2026-07-10 to 2026-07-13 confirmed available.", "On booking summary page."],
         "data": ["Unit: 'Cozy Cottage'", "Check-in: 2026-07-10", "Check-out: 2026-07-13", "Total: RM 850"],
         "scenario": "Verify guest can confirm booking, record created 'awaiting payment', bill generated, and notifications sent.",
         "steps": [
             ("Review booking summary. Click 'Confirm Booking'.", "System double-checks date availability."),
             ("Dates confirmed.", "Booking record: status 'awaiting payment', 1-day deadline. Bill with unique number generated. Dates temporarily reserved."),
             ("Verify notifications.", "In-app + email sent to guest with bill and payment deadline."),
             ("Observe redirect.", "Guest directed to payment page."),
         ], "color": NORMAL_FILL},
        {"tab": "Alternate Flow", "id": "TC-BK-02_AF",
         "desc": "Submit Booking – Alternate Flow (Race Condition)",
         "prereqs": ["Another guest booked same dates between selection and submission."],
         "data": ["Selected dates now taken by another user"],
         "scenario": "Verify system detects race condition when dates become unavailable on confirmation.",
         "steps": [
             ("Click 'Confirm Booking' while selected dates were just taken.", "System re-checks availability."),
             ("Observe response.", "Conflict message. 'These dates are no longer available. Please select new dates.' No booking created."),
         ], "color": ALT_FILL},
        {"tab": "Exception Flow", "id": "TC-BK-02_EF",
         "desc": "Submit Booking – Exception Flow (System Error)",
         "prereqs": ["Guest signed in. Valid dates selected. System error simulated."],
         "data": ["Simulated system/DB error during booking creation"],
         "scenario": "Verify no partial booking record created on system error.",
         "steps": [
             ("Click 'Confirm Booking' while system error simulated.", "Submitted."),
             ("Observe response.", "Error message. No booking record created. Guest prompted to try again."),
         ], "color": EXCEP_FILL},
    ])


def gen_bk03():
    make_wb("TC-BK-03_ViewBookings.xlsx", [
        {"tab": "Normal Flow – Guest", "id": "TC-BK-03_NF_Guest",
         "desc": "View Bookings – Normal Flow (Guest)",
         "prereqs": ["Guest signed in with at least 1 current and 1 past booking."],
         "data": ["Guest with confirmed and completed bookings"],
         "scenario": "Verify guest can view bookings separated into Current and History tabs.",
         "steps": [
             ("Navigate to 'My Bookings'.", "System retrieves all bookings for the guest."),
             ("Observe tabs.", "Bookings in 'Current' (active/upcoming) and 'History' (completed/cancelled) tabs."),
             ("Check booking card.", "Shows: unit name, check-in/out dates, status, total cost."),
             ("Click a booking.", "Navigated to booking detail (UC-BK-04)."),
             ("On completed booking, observe links.", "'View Receipt' and 'Leave a Review' (if not yet done) shown."),
         ], "color": NORMAL_FILL},
        {"tab": "Normal Flow – Admin", "id": "TC-BK-03_NF_Admin",
         "desc": "View Bookings – Normal Flow (Admin)",
         "prereqs": ["Admin signed in with booking management access.", "Multiple bookings exist across units."],
         "data": ["Filter by status: 'confirmed'", "Date range: July 2026"],
         "scenario": "Verify admin can view all system-wide bookings with filters and calendar view.",
         "steps": [
             ("Navigate to Booking Management.", "System retrieves all bookings across all units."),
             ("Observe list.", "Bookings with filter options: status, date range, unit, booking reference."),
             ("Apply filter: status 'confirmed'.", "List updates to confirmed bookings only."),
             ("Open booking calendar view.", "All-unit calendar with all reservations."),
             ("Click a booking entry.", "Navigated to detail or action options."),
         ], "color": NORMAL_FILL},
    ])


def gen_bk04():
    make_wb("TC-BK-04_ViewBookingDetails.xlsx", [
        {"tab": "Normal Flow", "id": "TC-BK-04_NF",
         "desc": "View Booking Details – Normal Flow",
         "prereqs": ["Booking BK-2026-001 confirmed. Associated QR code exists."],
         "data": ["Booking reference: BK-2026-001"],
         "scenario": "Verify all booking details displayed including QR code for confirmed bookings.",
         "steps": [
             ("Click BK-2026-001 from bookings list.", "Full booking record retrieved."),
             ("Observe displayed info.", "Shown: unit name+photo, check-in/out date+time, total amount, payment status, booking status, cancellation policy+estimated refund, QR code (confirmed), extension history."),
             ("Click 'Download Bill'.", "Bill document downloaded."),
         ], "color": NORMAL_FILL},
    ])


def gen_bk05():
    make_wb("TC-BK-05_CancelBooking.xlsx", [
        {"tab": "Normal Flow", "id": "TC-BK-05_NF",
         "desc": "Cancel Booking – Normal Flow (With Refund)",
         "prereqs": ["BK-2026-001 is 'confirmed'. Check-in 20 days away (>14d = 100% refund)."],
         "data": ["Booking: BK-2026-001", "Refund tier: >14 days = 100%"],
         "scenario": "Verify cancellation, refund calculation, date release, and notifications.",
         "steps": [
             ("Open BK-2026-001. Click 'Cancel Booking'.", "System calculates refund per cancellation policy."),
             ("Observe confirmation dialog.", "'You will receive a refund of [100% amount]. This cannot be undone.'"),
             ("Click 'Confirm Cancellation'.", "Cancellation submitted."),
             ("Observe response.", "Status = 'cancelled'. Dates released. Refund processed. In-app + email to guest and admin. Audit log recorded."),
         ], "color": NORMAL_FILL},
        {"tab": "Alternate Flow", "id": "TC-BK-05_AF",
         "desc": "Cancel Booking – Alternate Flows",
         "prereqs": ["BK-2026-002 is 'awaiting payment' (no payment)."],
         "data": ["A1: User clicks Cancel in dialog", "A2: Booking 'awaiting payment'"],
         "scenario": "Verify cancelling dialog makes no changes; awaiting payment has no refund.",
         "steps": [
             ("(A1) Click 'Cancel Booking', then Cancel in dialog.", "Dialog dismissed. No changes. Booking unchanged."),
             ("(A2) Cancel booking in 'awaiting payment' status.", "No refund (no payment made). Simply cancelled. Dates released. Notification sent."),
         ], "color": ALT_FILL},
        {"tab": "Exception Flow", "id": "TC-BK-05_EF",
         "desc": "Cancel Booking – Exception Flow (Refund Failure)",
         "prereqs": ["Payment gateway refund simulated to fail."],
         "data": ["Simulated gateway refund failure"],
         "scenario": "Verify refund failure still completes cancellation and flags for manual admin processing.",
         "steps": [
             ("Confirm cancellation while gateway refund fails.", "Cancellation proceeds."),
             ("Observe response.", "Booking cancelled. Refund flagged for manual admin processing. Admin notified. Guest receives cancellation notification."),
         ], "color": EXCEP_FILL},
    ])


def gen_bk06():
    make_wb("TC-BK-06_ManageBookingAdmin.xlsx", [
        {"tab": "Normal Flow – Create", "id": "TC-BK-06_NF_Create",
         "desc": "Manage Booking (Admin) – Create on Behalf of Guest",
         "prereqs": ["Admin signed in.", "'guest@example.com' exists.", "Dates 2026-08-01 to 2026-08-05 available."],
         "data": ["Guest: 'guest@example.com'", "Unit: 'Cozy Cottage'", "Check-in: 2026-08-01", "Check-out: 2026-08-05"],
         "scenario": "Verify admin can create a booking on behalf of a guest.",
         "steps": [
             ("Navigate to Bookings > Create Booking.", "Create form displayed."),
             ("Select guest, unit, and dates. Observe booking summary.", "Real-time availability checked. Total cost shown."),
             ("Click Confirm.", "Booking created 'awaiting payment'. Bill generated. Guest notified to pay."),
         ], "color": NORMAL_FILL},
        {"tab": "Normal Flow – Edit", "id": "TC-BK-06_NF_Edit",
         "desc": "Manage Booking (Admin) – Edit Booking",
         "prereqs": ["Booking BK-2026-005 exists. Admin signed in."],
         "data": ["New Check-out: 2026-08-07"],
         "scenario": "Verify admin can edit booking details including dates.",
         "steps": [
             ("Open BK-2026-005. Click Edit.", "Edit form with current details."),
             ("Change check-out to 2026-08-07. Click Submit.", "Availability checked for new dates."),
             ("Observe response.", "Record saved. Audit log recorded. Guest notified."),
         ], "color": NORMAL_FILL},
        {"tab": "Normal Flow – Delete&Block", "id": "TC-BK-06_NF_Delete",
         "desc": "Manage Booking (Admin) – Delete & Block Dates",
         "prereqs": ["BK-2026-006 exists. Admin signed in."],
         "data": ["Delete: BK-2026-006", "Block: 'Cozy Cottage' 2026-09-10 to 2026-09-15"],
         "scenario": "Verify admin can delete a booking and block specific dates.",
         "steps": [
             ("Select BK-2026-006. Click Delete.", "Warning: permanent, cannot undo. Confirm required."),
             ("Confirm deletion.", "Booking deleted. Dates released. Guest notified. Audit log."),
             ("Select unit, set date range 2026-09-10 to 2026-09-15. Click Block Dates.", "Blocked dates stored. Show as 'blocked' on guest calendar. Internal note stored."),
         ], "color": NORMAL_FILL},
        {"tab": "Alternate Flow", "id": "TC-BK-06_AF",
         "desc": "Manage Booking (Admin) – Alternate Flows",
         "prereqs": ["Date conflict and block conflict scenarios prepared."],
         "data": ["A1: Conflicting new dates during edit", "A2: Cancel delete dialog", "A3: Block dates with confirmed booking"],
         "scenario": "Verify conflict and cancel handling.",
         "steps": [
             ("(A1) Edit booking to already-taken dates.", "Conflict shown. Admin prompted to select different dates."),
             ("(A2) Click Delete, then Cancel on warning dialog.", "Dialog dismissed. No changes. Record intact."),
             ("(A3) Block dates that have a confirmed booking.", "System warns of conflict. Admin must cancel existing booking first."),
         ], "color": ALT_FILL},
    ])


def gen_bk07():
    make_wb("TC-BK-07_AutoCancelExpiredBooking.xlsx", [
        {"tab": "Normal Flow", "id": "TC-BK-07_NF",
         "desc": "Auto-Cancel Expired Booking – Normal Flow",
         "prereqs": ["BK-2026-010 is 'awaiting payment'. Payment deadline passed >1 hour ago."],
         "data": ["BK-2026-010: awaiting payment, deadline passed"],
         "scenario": "Verify scheduled job auto-cancels overdue 'awaiting payment' bookings.",
         "steps": [
             ("Trigger/wait for scheduled job (~hourly).", "Job executes."),
             ("System queries overdue bookings.", "Finds BK-2026-010: status 'awaiting payment' AND deadline < now."),
             ("Observe result for BK-2026-010.", "Status set to 'cancelled'. Dates released."),
             ("Verify notification.", "In-app + email cancellation sent to guest."),
             ("Verify audit log.", "Entry: 'System auto-cancelled booking BK-2026-010'."),
         ], "color": NORMAL_FILL},
    ])


# ══════════════════════════════════════════════════
#  MODULE 4 – PAYMENT
# ══════════════════════════════════════════════════
def gen_pay01():
    make_wb("TC-PAY-01_MakePayment.xlsx", [
        {"tab": "Normal Flow", "id": "TC-PAY-01_NF",
         "desc": "Make Payment – Normal Flow",
         "prereqs": ["BK-2026-001 'awaiting payment'. Bill generated. Deadline not passed."],
         "data": ["Booking: BK-2026-001", "Test card: 4242 4242 4242 4242"],
         "scenario": "Verify payment confirms booking, generates QR code, receipt, and notifications.",
         "steps": [
             ("Click 'Pay Now' from booking detail or notification.", "Payment request created. Guest redirected to secure payment page."),
             ("Complete payment with test card.", "Payment submitted to gateway."),
             ("Gateway sends confirmation.", "System receives and verifies confirmation."),
             ("Observe system actions.", "Payment = successful. Booking = 'confirmed'. QR code generated (valid check-in to check-out). Receipt generated."),
             ("Verify notifications.", "Guest: in-app + email with receipt + QR. Admin: notified of new confirmed booking."),
         ], "color": NORMAL_FILL},
        {"tab": "Alternate Flow", "id": "TC-PAY-01_AF",
         "desc": "Make Payment – Alternate Flows (Declined, Abandoned)",
         "prereqs": ["BK-2026-001 'awaiting payment'."],
         "data": ["A1: Declined card 4000 0000 0000 0002", "A2: Guest closes payment page"],
         "scenario": "Verify payment failure and abandoned payment handling.",
         "steps": [
             ("(A1) Enter declined card on payment page.", "Gateway rejects. Payment = 'failed'. Guest returned with error + Try Again option. Booking remains 'awaiting payment'."),
             ("(A2) Guest redirected to payment page but closes browser.", "No confirmation received. Booking remains 'awaiting payment' until deadline."),
         ], "color": ALT_FILL},
        {"tab": "Exception Flow", "id": "TC-PAY-01_EF",
         "desc": "Make Payment – Exception Flows (Duplicate, Unverifiable Confirmation)",
         "prereqs": ["Valid payment already processed for BK-2026-001."],
         "data": ["E1: Same confirmation sent twice", "E2: Forged confirmation payload"],
         "scenario": "Verify duplicate and unverifiable confirmations are handled securely.",
         "steps": [
             ("(E1) Simulate duplicate payment confirmation webhook.", "Duplicate detected. Ignored without re-processing. Booking stays 'confirmed' (not re-confirmed)."),
             ("(E2) Simulate forged/unverifiable webhook.", "Confirmation rejected. Security alert recorded. Booking NOT changed. No QR or receipt generated."),
         ], "color": EXCEP_FILL},
    ])


def gen_pay02():
    make_wb("TC-PAY-02_ViewPaymentBillingRecords.xlsx", [
        {"tab": "Normal Flow – Guest", "id": "TC-PAY-02_NF_Guest",
         "desc": "View Payment & Billing Records – Normal Flow (Guest)",
         "prereqs": ["Guest signed in with payment and bill records."],
         "data": ["Guest with at least 1 paid booking"],
         "scenario": "Verify guest can view payment records and download bills/receipts.",
         "steps": [
             ("Navigate to My Bookings > booking detail or Payment History.", "System retrieves payment records."),
             ("Observe records.", "Each: payment number, booking ref, date, amount, status."),
             ("Click a bill to view.", "Itemised bill: nightly rate x nights, deposit, total, deadline."),
             ("Click 'Download Bill'.", "Bill PDF downloaded."),
             ("For completed payment, click 'View Receipt'.", "Receipt displayed. Download available."),
         ], "color": NORMAL_FILL},
        {"tab": "Normal Flow – Admin", "id": "TC-PAY-02_NF_Admin",
         "desc": "View Payment & Billing Records – Normal Flow (Admin)",
         "prereqs": ["Admin signed in with payment management access."],
         "data": ["Filter: status = 'successful'", "Filter: July 2026"],
         "scenario": "Verify admin can view all billing and payment records with filters.",
         "steps": [
             ("Navigate to Payment Management.", "All billing and payment records retrieved."),
             ("Observe billing list.", "Each: reference, guest name, unit, amount, status, date."),
             ("Apply filters.", "List shows matching records only."),
             ("Click entry.", "Detail view. Admin can take action."),
         ], "color": NORMAL_FILL},
    ])


def gen_pay03():
    make_wb("TC-PAY-03_RegenerateBillReceipt.xlsx", [
        {"tab": "Normal Flow", "id": "TC-PAY-03_NF",
         "desc": "Regenerate Bill / Receipt – Normal Flow",
         "prereqs": ["BK-2026-001 has billing and payment records.", "Admin signed in."],
         "data": ["Booking: BK-2026-001"],
         "scenario": "Verify admin can regenerate bill and receipt for download and optionally resend to guest.",
         "steps": [
             ("Navigate to billing/payment detail for BK-2026-001.", "Detail page loaded."),
             ("Click 'Regenerate Bill'.", "System fetches latest data. Fresh bill generated. Available for download. Option to resend to guest."),
             ("Click 'Regenerate Receipt'.", "Fresh receipt generated. Available for download."),
         ], "color": NORMAL_FILL},
    ])


# ══════════════════════════════════════════════════
#  MODULE 5 – NOTIFICATION
# ══════════════════════════════════════════════════
def gen_notif01():
    make_wb("TC-NOTIF-01_ReceiveSystemNotification.xlsx", [
        {"tab": "Normal Flow – In-App", "id": "TC-NOTIF-01_NF_InApp",
         "desc": "Receive System Notification – In-App",
         "prereqs": ["User signed in and online.", "Booking confirmation event occurs."],
         "data": ["Triggering event: booking confirmed BK-2026-001"],
         "scenario": "Verify in-app bell updates in real time and notification navigates to relevant page.",
         "steps": [
             ("Trigger booking confirmation (payment made for BK-2026-001).", "Notification record created for guest."),
             ("Observe bell icon (user is online).", "Badge updates in real time without page refresh."),
             ("Click the bell icon.", "Notifications panel opens: list (newest first, read + unread)."),
             ("Click the booking notification.", "Marked as read. Navigated to booking detail for BK-2026-001."),
         ], "color": NORMAL_FILL},
        {"tab": "Normal Flow – Email", "id": "TC-NOTIF-01_NF_Email",
         "desc": "Receive System Notification – Email",
         "prereqs": ["Email notifications globally enabled.", "User has valid registered email."],
         "data": ["Triggering event: cancellation processed"],
         "scenario": "Verify email notification sent when event occurs and email is enabled.",
         "steps": [
             ("Trigger booking cancellation.", "System checks if email notifications enabled."),
             ("Email is enabled.", "System sends email to registered address."),
             ("Open email inbox.", "Email received with cancellation details."),
         ], "color": NORMAL_FILL},
        {"tab": "Normal Flow – Reminders", "id": "TC-NOTIF-01_NF_Reminders",
         "desc": "Receive System Notification – Automated Reminders",
         "prereqs": ["BK-2026-020: awaiting payment, deadline tomorrow.", "BK-2026-021: check-in tomorrow."],
         "data": ["BK-2026-020: payment deadline tomorrow", "BK-2026-021: check-in tomorrow"],
         "scenario": "Verify scheduled daily reminders for payment deadlines and check-in/out dates.",
         "steps": [
             ("Daily scheduled job runs.", "Finds awaiting-payment bookings with deadline approaching."),
             ("Observe payment reminder.", "Reminder sent for BK-2026-020: ref, amount, deadline, payment link."),
             ("System finds confirmed bookings with check-in tomorrow.", "Check-in reminder sent to guest and admin for BK-2026-021."),
         ], "color": NORMAL_FILL},
        {"tab": "Alternate Flow", "id": "TC-NOTIF-01_AF",
         "desc": "Receive System Notification – Email Globally Disabled",
         "prereqs": ["Email notifications globally disabled in settings."],
         "data": ["Email toggle: disabled"],
         "scenario": "Verify in-app notifications still delivered when email is disabled.",
         "steps": [
             ("Trigger booking confirmed event.", "System checks email toggle."),
             ("Email is disabled.", "System skips email. No email sent."),
             ("Observe in-app notification.", "In-app notification created and delivered normally."),
         ], "color": ALT_FILL},
        {"tab": "Exception Flow", "id": "TC-NOTIF-01_EF",
         "desc": "Receive System Notification – Email Server Unreachable",
         "prereqs": ["Email enabled. Mail server simulated unreachable."],
         "data": ["Mail server: simulated unreachable"],
         "scenario": "Verify email server failure is logged/retried and does not affect in-app notifications.",
         "steps": [
             ("Trigger event while mail server unreachable.", "System attempts to send email."),
             ("Email send fails.", "System retries. After max retries, failure recorded. In-app notification unaffected and delivered normally."),
         ], "color": EXCEP_FILL},
    ])


def gen_notif02():
    make_wb("TC-NOTIF-02_GoogleCalendarIntegration.xlsx", [
        {"tab": "Normal Flow", "id": "TC-NOTIF-02_NF",
         "desc": "Google Calendar Integration – Normal Flow",
         "prereqs": ["User connected Google account.", "BK-2026-001 confirmed."],
         "data": ["Booking: BK-2026-001", "Check-in: 2026-07-10 2PM", "Check-out: 2026-07-13 12PM"],
         "scenario": "Verify confirmed booking is added to user's Google Calendar.",
         "steps": [
             ("Payment confirmed for BK-2026-001.", "System detects booking confirmed event."),
             ("System retrieves Google account connection.", "Connection found."),
             ("System creates Calendar event.", "Event: title = 'Unit Name Stay', start = 2026-07-10 2PM, end = 2026-07-13 12PM, description = booking ref + address."),
             ("Verify event in Google Calendar.", "Event appears in user's calendar."),
         ], "color": NORMAL_FILL},
        {"tab": "Alternate Flow", "id": "TC-NOTIF-02_AF",
         "desc": "Google Calendar Integration – No Google Account Connected",
         "prereqs": ["User has NOT connected Google account."],
         "data": ["User: no Google connection stored"],
         "scenario": "Verify system skips calendar integration silently when not connected.",
         "steps": [
             ("Booking confirmed. System checks Google connection.", "No connection found."),
             ("Observe.", "Calendar step skipped. No error shown to user."),
         ], "color": ALT_FILL},
        {"tab": "Exception Flow", "id": "TC-NOTIF-02_EF",
         "desc": "Google Calendar Integration – Calendar Service Error",
         "prereqs": ["User has Google connected. Google Calendar API returns error."],
         "data": ["Simulated Google API error (token expired)"],
         "scenario": "Verify Google Calendar failure is logged without affecting booking confirmation.",
         "steps": [
             ("Booking confirmed. System attempts calendar event. API returns error.", "API call fails."),
             ("System response.", "Failure logged. User may be prompted to reconnect. Booking confirmation NOT affected."),
         ], "color": EXCEP_FILL},
    ])


# ══════════════════════════════════════════════════
#  MODULE 6 – CHAT
# ══════════════════════════════════════════════════
def gen_chat01():
    make_wb("TC-CHAT-01_SendReceiveMessages.xlsx", [
        {"tab": "Normal Flow", "id": "TC-CHAT-01_NF",
         "desc": "Send / Receive Messages – Normal Flow (Real-Time Delivery)",
         "prereqs": ["Both Guest and Admin signed in.", "Guest on Chat page."],
         "data": ["Message: 'Hello, I have a question about my booking.'"],
         "scenario": "Verify guest can send a message instantly delivered to online admin.",
         "steps": [
             ("Guest opens Chat page.", "Chat interface displayed."),
             ("Guest types message and clicks Send.", "Message submitted. Saved: sender, recipient, content, timestamp."),
             ("Admin is currently online.", "Message delivered instantly to admin chat window. Admin unread badge updated."),
             ("Message in guest window.", "Shown as 'sent'."),
         ], "color": NORMAL_FILL},
        {"tab": "Alternate Flow", "id": "TC-CHAT-01_AF",
         "desc": "Send / Receive Messages – Alternate Flows",
         "prereqs": ["Guest signed in. Admin offline."],
         "data": ["A1: Empty message", "A2: Admin is offline"],
         "scenario": "Verify empty messages blocked and offline messages stored.",
         "steps": [
             ("(A1) Leave message blank. Click Send.", "Send button disabled. Empty message cannot be sent."),
             ("(A2) Send valid message while admin offline.", "Message saved. Shown to admin on next sign-in."),
         ], "color": ALT_FILL},
        {"tab": "Exception Flow", "id": "TC-CHAT-01_EF",
         "desc": "Send / Receive Messages – Exception Flow (Connection Lost)",
         "prereqs": ["Guest on Chat page. Network interrupted."],
         "data": ["Simulated network disconnection"],
         "scenario": "Verify reconnecting indicator shown and message preserved on reconnection.",
         "steps": [
             ("Simulate network disconnection on Chat page.", "'Reconnecting...' indicator shown."),
             ("Reconnection restored.", "Message still saved. Visible on page reload."),
         ], "color": EXCEP_FILL},
    ])


def gen_chat02():
    make_wb("TC-CHAT-02_ViewChatHistory.xlsx", [
        {"tab": "Normal Flow", "id": "TC-CHAT-02_NF",
         "desc": "View Chat History – Normal Flow",
         "prereqs": ["Conversation with 5+ messages. Some unread."],
         "data": ["User opens Chat page with existing conversation"],
         "scenario": "Verify messages shown chronologically, unread marked as read, auto-scroll to latest.",
         "steps": [
             ("Open Chat page.", "System retrieves all messages oldest to newest."),
             ("Observe display.", "Sent messages on right. Received on left. Each with sender name + timestamp."),
             ("Observe unread status.", "All unread messages marked as read."),
             ("Observe scroll.", "Auto-scrolled to most recent message."),
         ], "color": NORMAL_FILL},
    ])


# ══════════════════════════════════════════════════
#  MODULE 7 – USER & ACCESS MANAGEMENT
# ══════════════════════════════════════════════════
def gen_usr01():
    make_wb("TC-USR-01_CreateUserAccount.xlsx", [
        {"tab": "Normal Flow", "id": "TC-USR-01_NF",
         "desc": "Create User Account – Normal Flow",
         "prereqs": ["Admin signed in with user management permission.", "'newstaff@example.com' not registered."],
         "data": ["Name: 'Staff Member One'", "Email: 'newstaff@example.com'", "Role: 'Admin'"],
         "scenario": "Verify admin can create new user with temp password, forced-change flag, and email notification.",
         "steps": [
             ("Navigate to User Management > Create User.", "Create user form displayed."),
             ("Enter name, email, role. Click Create.", "System generates temp password."),
             ("Observe response.", "User created. Forced change flag set. Email with temp password sent. Audit log recorded. Success message."),
         ], "color": NORMAL_FILL},
        {"tab": "Alternate Flow", "id": "TC-USR-01_AF",
         "desc": "Create User Account – Alternate Flow (Email Already Registered)",
         "prereqs": ["'existing@example.com' already registered."],
         "data": ["Email: 'existing@example.com'"],
         "scenario": "Verify user creation rejected if email already in use.",
         "steps": [
             ("Enter 'existing@example.com'. Fill other fields. Click Create.", "Validation runs."),
             ("Observe error.", "Error: 'Email already registered.' Admin can edit existing account. No new account."),
         ], "color": ALT_FILL},
    ])


def gen_usr02():
    make_wb("TC-USR-02_EditActivateDeactivateUser.xlsx", [
        {"tab": "Normal Flow – Edit", "id": "TC-USR-02_NF_Edit",
         "desc": "Edit User – Normal Flow",
         "prereqs": ["User 'staff@example.com' exists. Admin signed in."],
         "data": ["Updated Name: 'Staff Member Updated'", "Updated Role: 'Guest'"],
         "scenario": "Verify admin can edit user details and changes saved.",
         "steps": [
             ("Navigate to User Management > select 'staff@example.com' > Edit.", "Pre-filled edit form."),
             ("Update Name and Role. Click Save.", "Validation (email unique). Changes saved."),
             ("Observe response.", "Updated. If role changed, new permissions immediate. Audit log. Success message."),
         ], "color": NORMAL_FILL},
        {"tab": "Normal Flow – Deactivate", "id": "TC-USR-02_NF_Deactivate",
         "desc": "Activate / Deactivate User – Normal Flow",
         "prereqs": ["Active 'staff@example.com' exists. Admin signed in."],
         "data": ["Target: 'staff@example.com'"],
         "scenario": "Verify admin can deactivate (ending sessions) and reactivate a user.",
         "steps": [
             ("Select 'staff@example.com'. Click Deactivate.", "Confirmation dialog."),
             ("Confirm.", "Account disabled. Active sessions immediately ended. Audit log. Success."),
             ("Attempt sign in as 'staff@example.com'.", "'Your account has been deactivated.'"),
             ("Admin: Select user. Click Activate. Confirm.", "Account restored. Audit log. Success."),
             ("Attempt sign in again.", "Login succeeds."),
         ], "color": NORMAL_FILL},
    ])


def gen_usr03():
    make_wb("TC-USR-03_ResetUserPassword.xlsx", [
        {"tab": "Normal Flow", "id": "TC-USR-03_NF",
         "desc": "Reset User Password – Normal Flow",
         "prereqs": ["'staff@example.com' exists. Admin signed in. Account optionally locked."],
         "data": ["Option A: Send reset link", "Option B: Set to default temp password"],
         "scenario": "Verify admin can reset password via both methods; forced-change flag set; locked account unlocked.",
         "steps": [
             ("Navigate to User Management > select 'staff@example.com' > Reset Password.", "Reset method selection shown."),
             ("(Option A) Choose 'Send reset link'. Submit.", "Reset email sent. Forced change flag set. If locked: unlocked, counter reset. Notified. Audit log."),
             ("(Option B) Choose 'Set to default'. Submit.", "Password set to default temp. Forced change flag set. Unlocked if applicable. Notified. Audit log."),
             ("User signs in with temp/new password.", "Login succeeds. Redirected to forced change page."),
         ], "color": NORMAL_FILL},
    ])


def gen_usr04():
    make_wb("TC-USR-04_ManageRoles.xlsx", [
        {"tab": "Normal Flow – Create&Edit", "id": "TC-USR-04_NF_CreateEdit",
         "desc": "Manage Roles – Create & Edit",
         "prereqs": ["Admin signed in with role management access."],
         "data": ["Role Name: 'Property Manager'", "Updated Name: 'Senior Property Manager'"],
         "scenario": "Verify admin can create and edit roles.",
         "steps": [
             ("Navigate to Role Management > Create Role. Enter name. Click Create.", "Name uniqueness checked. Role created. Admin on role detail to assign permissions."),
             ("Select role. Click Edit. Update name. Save.", "Uniqueness checked. Role updated."),
         ], "color": NORMAL_FILL},
        {"tab": "Normal Flow – Assign&Delete", "id": "TC-USR-04_NF_AssignDelete",
         "desc": "Manage Roles – Assign Permissions & Delete",
         "prereqs": ["Role 'Temp Role' exists with no assigned users."],
         "data": ["Permissions: 'manage_bookings', 'view_reports'"],
         "scenario": "Verify admin can assign permissions and delete a role with no users.",
         "steps": [
             ("Navigate to role > Manage Permissions.", "Checklist of all permissions. Assigned ones ticked."),
             ("Tick 'manage_bookings' and 'view_reports'. Save.", "Role permissions updated. Immediate effect. Audit log."),
             ("Select 'Temp Role'. Click Delete.", "System checks: no users have this role."),
             ("Confirm deletion.", "Role deleted. Audit log."),
         ], "color": NORMAL_FILL},
        {"tab": "Alternate Flow", "id": "TC-USR-04_AF",
         "desc": "Manage Roles – Alternate Flow (Delete Assigned Role)",
         "prereqs": ["Role 'Guest' assigned to multiple users."],
         "data": ["Role: 'Guest' (assigned to 10 users)"],
         "scenario": "Verify deletion blocked when role assigned to users.",
         "steps": [
             ("Select 'Guest' role. Click Delete.", "System checks: 10 users have this role."),
             ("Observe error.", "Blocked. Message: 'This role is assigned to 10 user(s). Reassign users before deleting.'"),
         ], "color": ALT_FILL},
    ])


def gen_usr05():
    make_wb("TC-USR-05_ManagePermissions.xlsx", [
        {"tab": "Normal Flow", "id": "TC-USR-05_NF",
         "desc": "Manage Permissions – Create, Edit, Delete",
         "prereqs": ["Admin signed in.", "'temp_permission' exists and NOT attached to any role."],
         "data": ["Create: 'can_export_reports'", "Edit: 'can_export_all_reports'", "Delete: 'temp_permission'"],
         "scenario": "Verify admin can create, edit, and delete permissions.",
         "steps": [
             ("Permission Management > Create Permission. Enter 'can_export_reports'. Save.", "Name unique. Permission created."),
             ("Select. Edit. Update to 'can_export_all_reports'. Save.", "Update saved."),
             ("Select 'temp_permission'. Delete.", "System checks: not attached to any role."),
             ("Confirm.", "Deleted. Audit log."),
         ], "color": NORMAL_FILL},
        {"tab": "Alternate Flow", "id": "TC-USR-05_AF",
         "desc": "Manage Permissions – Alternate Flow (Delete Attached Permission)",
         "prereqs": ["'manage_bookings' attached to 2 roles."],
         "data": ["Permission: 'manage_bookings' (2 roles)"],
         "scenario": "Verify deletion blocked when permission still attached to roles.",
         "steps": [
             ("Select 'manage_bookings'. Click Delete.", "System checks: attached to 2 roles."),
             ("Observe.", "Blocked. 'This permission is attached to 2 role(s). Remove from roles before deleting.'"),
         ], "color": ALT_FILL},
    ])


# ══════════════════════════════════════════════════
#  MODULE 8 – SYSTEM SETTINGS
# ══════════════════════════════════════════════════
def gen_set01():
    make_wb("TC-SET-01_ConfigureSystemSettings.xlsx", [
        {"tab": "Normal Flow", "id": "TC-SET-01_NF",
         "desc": "Configure System Settings – Normal Flow",
         "prereqs": ["Admin signed in with system settings access."],
         "data": ["Max attempts: 5, Lockout: 30min", "Cancellation: >14d=100%, 7-14d=50%, <7d=0%",
                  "Extension: RM30/hr, RM150/night", "Email toggle: enabled"],
         "scenario": "Verify admin can update all settings sections and values take immediate effect.",
         "steps": [
             ("Navigate to System Settings.", "Settings page with multiple sections."),
             ("Update Security: max attempts=5, lockout=30min, session timeout=120min.", "Fields accept values."),
             ("Update Cancellation Policy tiers.", "Values entered."),
             ("Update Extension Charges and Window. Toggle Email Notifications to enabled.", "Values set."),
             ("Click Save.", "System validates all settings."),
             ("Observe response.", "Settings saved. Confirmation: 'Settings saved. All future operations use new values.'"),
         ], "color": NORMAL_FILL},
        {"tab": "Alternate Flow", "id": "TC-SET-01_AF",
         "desc": "Configure System Settings – Alternate Flows",
         "prereqs": ["Admin on Settings page. Mail server simulated unreachable."],
         "data": ["A1: Invalid mail server address", "A2: Overlapping refund tiers"],
         "scenario": "Verify email test failure and validation errors are reported.",
         "steps": [
             ("(A1) Enter invalid mail server. Click Test Connection.", "Connection fails. Error shown. Admin can correct and retry."),
             ("(A2) Enter overlapping refund tiers. Click Save.", "Validation error: overlapping/invalid tiers. Settings NOT saved."),
         ], "color": ALT_FILL},
    ])


def gen_set02():
    make_wb("TC-SET-02_ManageDefaultHomestayPolicies.xlsx", [
        {"tab": "Normal Flow", "id": "TC-SET-02_NF",
         "desc": "Manage Default Homestay Policies – Normal Flow",
         "prereqs": ["Admin signed in.", "Default policies: No Pets, No Durians, No Smoking."],
         "data": ["Add: 'No Parties'", "Edit: 'No Smoking' > 'No Smoking or Vaping'", "Remove: 'No Durians'"],
         "scenario": "Verify admin can add, edit, and remove default rules; changes only affect new units.",
         "steps": [
             ("Navigate to System Settings > Default Policies.", "Current list: No Pets, No Durians, No Smoking."),
             ("Click Add. Enter 'No Parties'. Save.", "New rule added."),
             ("Select 'No Smoking'. Edit to 'No Smoking or Vaping'. Save.", "Rule text updated."),
             ("Select 'No Durians'. Click Delete. Confirm.", "Rule removed."),
             ("Create a new unit.", "Inherits: No Pets, No Smoking or Vaping, No Parties. No Durians NOT applied."),
             ("Check existing units.", "Existing units NOT changed by this action."),
         ], "color": NORMAL_FILL},
    ])


# ══════════════════════════════════════════════════
#  MODULE 9 – AUDIT LOGS
# ══════════════════════════════════════════════════
def gen_audit01():
    make_wb("TC-AUDIT-01_ViewFilterAuditTrail.xlsx", [
        {"tab": "Normal Flow", "id": "TC-AUDIT-01_NF",
         "desc": "View / Filter Audit Trail – Normal Flow",
         "prereqs": ["Admin signed in with audit log access.", "Multiple log entries exist."],
         "data": ["Filter: date 2026-07-01 to 2026-07-15", "Filter: event type 'booking_created'"],
         "scenario": "Verify admin can view full audit trail, apply filters, and entries are read-only.",
         "steps": [
             ("Navigate to Audit Logs.", "Entries retrieved in reverse chronological order."),
             ("Observe entries.", "Each: timestamp, actor (name or 'System'), event type, what was affected."),
             ("Page through entries.", "Pagination works correctly."),
             ("Apply date filter 2026-07-01 to 2026-07-15.", "Only entries in that range shown."),
             ("Apply event type 'booking_created'.", "Only booking creation events shown."),
             ("Clear filters.", "Full unfiltered log restored."),
             ("Attempt to edit/delete a log entry.", "No edit/delete option. Entries are read-only."),
         ], "color": NORMAL_FILL},
    ])


def gen_audit02():
    make_wb("TC-AUDIT-02_AutomaticEventLogging.xlsx", [
        {"tab": "Normal Flow – User Action", "id": "TC-AUDIT-02_NF_User",
         "desc": "Automatic Event Logging – User / Admin Action",
         "prereqs": ["'guest@example.com' signed in.", "Booking creation about to be performed."],
         "data": ["Action: submit booking BK-2026-030"],
         "scenario": "Verify log entry auto-created when user performs significant action.",
         "steps": [
             ("Guest submits booking BK-2026-030.", "Booking created."),
             ("Admin navigates to Audit Logs.", "Entries listed."),
             ("Locate entry for BK-2026-030.", "Entry: timestamp = now, actor = 'guest@example.com', event = 'booking_created', record = BK-2026-030."),
             ("Verify immutability.", "No edit/delete. Cannot be modified."),
         ], "color": NORMAL_FILL},
        {"tab": "Normal Flow – System Action", "id": "TC-AUDIT-02_NF_System",
         "desc": "Automatic Event Logging – System / Automated Action",
         "prereqs": ["Auto-cancel job ran. BK-2026-031 was auto-cancelled."],
         "data": ["Auto-cancelled: BK-2026-031"],
         "scenario": "Verify automated system actions logged with 'System' as actor.",
         "steps": [
             ("Scheduled auto-cancel job runs and cancels BK-2026-031.", "System cancels booking."),
             ("Admin navigates to Audit Logs.", "Entries listed."),
             ("Locate entry for BK-2026-031.", "Entry: timestamp = job time, actor = 'System', event = 'booking_auto_cancelled', record = BK-2026-031."),
             ("Verify immutability.", "Cannot be edited or deleted."),
         ], "color": NORMAL_FILL},
    ])


# ══════════════════════════════════════════════════
#  MODULE 10 – QR CODE & ACCESS
# ══════════════════════════════════════════════════
def gen_qr01():
    make_wb("TC-QR-01_ReceiveUseQRCode.xlsx", [
        {"tab": "Normal Flow", "id": "TC-QR-01_NF",
         "desc": "Receive & Use QR Code – Normal Flow",
         "prereqs": ["BK-2026-001 confirmed. Check-in: 2026-07-10 2PM, Check-out: 2026-07-13 12PM."],
         "data": ["Booking: BK-2026-001", "Current time: 2026-07-10 3PM (within window)"],
         "scenario": "Verify QR generated on payment, delivered, and grants access within valid window.",
         "steps": [
             ("Payment confirmed for BK-2026-001.", "QR code generated: status=active, valid 2026-07-10 2PM to 2026-07-13 12PM."),
             ("Verify notifications.", "In-app + email with QR code and booking details."),
             ("Guest views QR on booking detail page.", "QR code visible."),
             ("At 2026-07-10 3PM, present QR to smart lock scanner.", "Lock queries system for validation."),
             ("System validates: active, time within window.", "Validation passes."),
             ("Observe result.", "Access granted. Door opens."),
         ], "color": NORMAL_FILL},
        {"tab": "Alternate Flow", "id": "TC-QR-01_AF",
         "desc": "Receive & Use QR Code – Alternate Flows",
         "prereqs": ["QR for BK-2026-001 expired. QR for BK-2026-002 inactive/revoked."],
         "data": ["A1: Past check-out time", "A2: QR status=inactive", "A3: No device (dead battery)"],
         "scenario": "Verify access denied for expired, inactive QR codes.",
         "steps": [
             ("(A1) Present expired QR (past 2026-07-13 12PM).", "Denied. 'QR Code Expired.'"),
             ("(A2) Present revoked/inactive QR.", "Denied. Access not granted."),
             ("(A3) Guest device unavailable.", "Cannot present QR. Must contact admin for manual assistance."),
         ], "color": ALT_FILL},
    ])


def gen_qr02():
    make_wb("TC-QR-02_ManageHousekeepingCycle.xlsx", [
        {"tab": "Normal Flow – QR Expiry", "id": "TC-QR-02_NF_Expiry",
         "desc": "Manage Housekeeping Cycle – Automatic QR Expiry",
         "prereqs": ["Active QR for BK-2026-001. Check-out 2026-07-13 12PM (past)."],
         "data": ["Check-out: 2026-07-13 12PM (past)"],
         "scenario": "Verify scheduled job expires QR codes past valid-until time and marks bookings completed.",
         "steps": [
             ("Scheduled job runs at/after 2026-07-13 12PM.", "Finds active QR codes with valid-until passed."),
             ("For BK-2026-001 QR:", "QR status = 'expired'. Booking BK-2026-001 = 'completed'."),
         ], "color": NORMAL_FILL},
        {"tab": "Normal Flow – Housekeeping", "id": "TC-QR-02_NF_Housekeeping",
         "desc": "Manage Housekeeping Cycle – Housekeeping QR & Complete",
         "prereqs": ["BK-2026-001 completed. Next booking BK-2026-002 confirmed on same unit."],
         "data": ["Housekeeping validity: 4 hours"],
         "scenario": "Verify admin generates housekeeping QR, marks complete, next guest auto-gets QR.",
         "steps": [
             ("Admin clicks Generate Housekeeping QR. Set validity: 4 hours.", "Temporary QR (type: housekeeping) with 4h validity generated."),
             ("QR displayed.", "Admin can share with cleaning staff."),
             ("Admin clicks Mark Housekeeping Complete.", "Housekeeping QR expired. Checks for next confirmed booking."),
             ("Next booking BK-2026-002 found.", "New QR for next guest. In-app + email notification sent."),
         ], "color": NORMAL_FILL},
        {"tab": "Alternate Flow", "id": "TC-QR-02_AF",
         "desc": "Manage Housekeeping Cycle – No Next Booking",
         "prereqs": ["No next confirmed booking after housekeeping complete."],
         "data": ["No upcoming confirmed booking"],
         "scenario": "Verify no QR generated when no next booking exists.",
         "steps": [
             ("Admin marks housekeeping complete.", "Housekeeping QR expired. Checks for next booking."),
             ("No next booking.", "QR generation skipped. No action needed. No notification."),
         ], "color": ALT_FILL},
    ])


def gen_qr03():
    make_wb("TC-QR-03_InitiateBookingExtension.xlsx", [
        {"tab": "Normal Flow", "id": "TC-QR-03_NF",
         "desc": "Initiate Booking Extension – Normal Flow",
         "prereqs": ["BK-2026-001 confirmed. Check-out: 2026-07-13 12PM. 2026-07-13 to 2026-07-14 free."],
         "data": ["Extension: Date (add 1 night)", "New check-out: 2026-07-14 12PM", "Rate: RM 150/night"],
         "scenario": "Verify admin initiates extension, record created, bill generated, guest notified.",
         "steps": [
             ("Admin opens BK-2026-001. Click Extend Stay.", "Extension form displayed."),
             ("Select Date extension. Enter new check-out: 2026-07-14 12PM.", "Availability checked."),
             ("Dates available.", "Charge calculated: 1 night x RM 150 = RM 150."),
             ("Observe system actions.", "Extension record (awaiting payment). Extension bill. Payment deadline set. Guest notified. Audit log."),
             ("Verify QR code.", "QR NOT extended. Still valid until original 2026-07-13 12PM."),
         ], "color": NORMAL_FILL},
        {"tab": "Alternate Flow", "id": "TC-QR-03_AF",
         "desc": "Initiate Booking Extension – Date Conflict",
         "prereqs": ["Another booking occupies 2026-07-13 to 2026-07-14 on same unit."],
         "data": ["Conflicting booking: BK-2026-005"],
         "scenario": "Verify extension blocked when requested period conflicts with another booking.",
         "steps": [
             ("Admin tries to extend BK-2026-001 to 2026-07-14.", "System checks availability."),
             ("Conflict with BK-2026-005.", "Conflict shown. 'Select different extension period.' No extension record created."),
         ], "color": ALT_FILL},
    ])


def gen_qr04():
    make_wb("TC-QR-04_PayExtensionCharge.xlsx", [
        {"tab": "Normal Flow", "id": "TC-QR-04_NF",
         "desc": "Pay Extension Charge – Normal Flow",
         "prereqs": ["EXT-001 for BK-2026-001 awaiting payment. Deadline not passed."],
         "data": ["Extension charge: RM 150", "Test card: 4242 4242 4242 4242"],
         "scenario": "Verify guest pays extension, booking updated to new check-out, QR extended.",
         "steps": [
             ("Open booking detail. Find pending extension EXT-001. Click Pay Extension Charge.", "Redirected to payment with RM 150."),
             ("Complete payment.", "Gateway sends confirmation."),
             ("System verifies and records.", "Payment = successful."),
             ("Observe updates.", "EXT-001 = 'confirmed'. Booking check-out = 2026-07-14 12PM. QR extended to 2026-07-14 12PM."),
             ("Verify notification.", "Guest notified: 'Stay extended. QR valid until 2026-07-14 12PM.' Audit log."),
         ], "color": NORMAL_FILL},
    ])


def gen_qr05():
    make_wb("TC-QR-05_AutoCancelExtension.xlsx", [
        {"tab": "Normal Flow", "id": "TC-QR-05_NF",
         "desc": "Auto-Cancel Extension – Normal Flow",
         "prereqs": ["EXT-002 for BK-2026-002 awaiting payment. Deadline passed."],
         "data": ["EXT-002: awaiting payment, deadline passed"],
         "scenario": "Verify scheduler auto-cancels overdue extensions and booking reverts to original check-out.",
         "steps": [
             ("Scheduled job runs (every few minutes).", "Finds extensions: awaiting payment AND deadline passed."),
             ("EXT-002 found.", "Extension = 'cancelled'. BK-2026-002 check-out reverted to original."),
             ("Verify QR code.", "QR NOT modified. Still reflects original check-out time."),
             ("Verify notification.", "Guest notified: 'Extension cancelled. Original check-out remains.' Audit log."),
         ], "color": NORMAL_FILL},
    ])


# ══════════════════════════════════════════════════
#  MODULE 11 – REPORTING & ANALYTICS
# ══════════════════════════════════════════════════
def gen_rpt01():
    make_wb("TC-RPT-01_ViewAnalyticsDashboard.xlsx", [
        {"tab": "Normal Flow", "id": "TC-RPT-01_NF",
         "desc": "View Analytics Dashboard – Normal Flow",
         "prereqs": ["Admin signed in with reporting access.", "Multiple bookings, payments, reviews exist."],
         "data": ["Date filter: July 2026", "Unit filter: 'Cozy Cottage'"],
         "scenario": "Verify admin can view all KPI cards and interactive charts on analytics dashboard.",
         "steps": [
             ("Navigate to Reporting & Analytics.", "Dashboard loads."),
             ("Observe KPI cards.", "Total bookings (month+YTD), total revenue, occupancy rate, cancellation rate, avg guest rating."),
             ("Observe charts.", "Booking trends (line, daily/weekly/monthly toggle), revenue summary, per-unit bar chart, feedback summary."),
             ("Interact with a chart.", "Detail/drill-down available."),
             ("Apply date filter July 2026.", "Dashboard reflects filtered date range."),
             ("Apply unit filter 'Cozy Cottage'.", "Data filtered to Cozy Cottage only."),
         ], "color": NORMAL_FILL},
    ])


def gen_rpt02():
    make_wb("TC-RPT-02_ViewRevenueReport.xlsx", [
        {"tab": "Normal Flow", "id": "TC-RPT-02_NF",
         "desc": "View Revenue Report – Normal Flow",
         "prereqs": ["Admin signed in with reporting access. Payment records exist."],
         "data": ["Date range: 2026-07-01 to 2026-07-31", "Unit: all", "Status: successful"],
         "scenario": "Verify admin can view filtered revenue report with totals, breakdown, and export.",
         "steps": [
             ("Navigate to Reporting > Revenue Report.", "Report page displayed."),
             ("Apply filters: July 2026, all units, successful.", "Filters applied."),
             ("Observe report.", "Total revenue, itemised by booking (ref, guest, unit, amount), summary per unit."),
             ("Click Export CSV.", "Export triggered (see TC-RPT-03)."),
         ], "color": NORMAL_FILL},
    ])


def gen_rpt03():
    make_wb("TC-RPT-03_ExportReport.xlsx", [
        {"tab": "Normal Flow", "id": "TC-RPT-03_NF",
         "desc": "Export Report – Normal Flow (PDF & CSV)",
         "prereqs": ["Admin viewing a report page with data loaded."],
         "data": ["Export type: PDF", "Export type: CSV"],
         "scenario": "Verify admin can export report as PDF and CSV.",
         "steps": [
             ("On report page, click 'Export PDF'.", "System collects filtered data."),
             ("Observe.", "PDF generated and automatically downloaded."),
             ("Click 'Export CSV'.", "System collects filtered data."),
             ("Observe.", "CSV generated and automatically downloaded."),
         ], "color": NORMAL_FILL},
    ])


# ══════════════════════════════════════════════════
#  MODULE 12 – GUEST FEEDBACK
# ══════════════════════════════════════════════════
def gen_fb01():
    make_wb("TC-FB-01_SubmitRatingFeedback.xlsx", [
        {"tab": "Normal Flow", "id": "TC-FB-01_NF",
         "desc": "Submit Rating & Feedback – Normal Flow",
         "prereqs": ["BK-2026-001 status 'completed'. No review yet. Guest signed in."],
         "data": ["Star rating: 4", "Comment: 'Great stay, very clean and comfortable!'"],
         "scenario": "Verify guest can submit review, it is visible, and average rating recalculated.",
         "steps": [
             ("Navigate to My Bookings > History. Find BK-2026-001. Click Leave a Review.", "System verifies: completed AND no review yet."),
             ("Feedback form displayed.", "Star rating (1-5, required) + optional comment."),
             ("Select 4 stars. Enter comment. Click Submit.", "Review submitted."),
             ("Observe response.", "Review saved (booking, unit, guest; visible). Avg rating recalculated. Thank-you message. 'Leave a Review' button hidden."),
             ("Verify review on unit listing/detail.", "Review with 4 stars and comment visible."),
         ], "color": NORMAL_FILL},
        {"tab": "Alternate Flow", "id": "TC-FB-01_AF",
         "desc": "Submit Rating & Feedback – Alternate Flows",
         "prereqs": ["BK-2026-003 already has a submitted review."],
         "data": ["A1: Submit without star rating", "A2: Already reviewed booking"],
         "scenario": "Verify missing star rating blocks submission; already-reviewed shows View Your Review.",
         "steps": [
             ("(A1) Write comment but no star rating. Click Submit.", "Star rating highlighted as required. Blocked."),
             ("(A2) Navigate to BK-2026-003 (already reviewed).", "'Leave a Review' hidden. 'View Your Review' shown. Form not accessible."),
         ], "color": ALT_FILL},
        {"tab": "Exception Flow", "id": "TC-FB-01_EF",
         "desc": "Submit Rating & Feedback – Exception Flow (Non-Completed Booking)",
         "prereqs": ["BK-2026-004 status 'confirmed' (not completed)."],
         "data": ["Direct URL access to review form for BK-2026-004"],
         "scenario": "Verify review submission blocked for non-completed bookings.",
         "steps": [
             ("Access review form for BK-2026-004 via direct URL.", "System verifies booking status."),
             ("Observe.", "Access blocked. Validation error. No review submitted."),
         ], "color": EXCEP_FILL},
    ])


def gen_fb02():
    make_wb("TC-FB-02_ViewSubmittedFeedback.xlsx", [
        {"tab": "Normal Flow", "id": "TC-FB-02_NF",
         "desc": "View Submitted Feedback – Normal Flow",
         "prereqs": ["Guest submitted at least 2 reviews. One has admin reply."],
         "data": ["Guest with 2 submitted reviews"],
         "scenario": "Verify guest views submitted reviews in read-only format including admin replies.",
         "steps": [
             ("Navigate to My Reviews or completed booking in history.", "System retrieves all guest review records."),
             ("Observe entries.", "Each (read-only): unit name, check-in/out dates, star rating, comment, submission date, admin reply."),
             ("Attempt to edit a review.", "No edit option. Read-only after submission."),
         ], "color": NORMAL_FILL},
    ])


def gen_fb03():
    make_wb("TC-FB-03_ViewManageAllFeedbackAdmin.xlsx", [
        {"tab": "Normal Flow – View", "id": "TC-FB-03_NF_View",
         "desc": "View / Manage All Feedback (Admin) – View & Filter",
         "prereqs": ["Admin signed in with feedback management access. Reviews exist (some hidden)."],
         "data": ["Filter: unit='Cozy Cottage'", "Filter: visibility='hidden'"],
         "scenario": "Verify admin can view all reviews including hidden and apply filters.",
         "steps": [
             ("Navigate to Guest Feedback management.", "All reviews retrieved including hidden."),
             ("Observe list.", "Each: guest name, unit, booking ref, rating, comment, date, visibility, reply status."),
             ("Apply unit filter 'Cozy Cottage'.", "Only Cozy Cottage reviews shown."),
             ("Apply visibility filter 'hidden'.", "Only hidden reviews shown."),
         ], "color": NORMAL_FILL},
        {"tab": "Normal Flow – Reply", "id": "TC-FB-03_NF_Reply",
         "desc": "View / Manage All Feedback (Admin) – Reply to Review",
         "prereqs": ["REV-001 exists with no admin reply."],
         "data": ["Reply: 'Thank you for your feedback! We hope to see you again.'"],
         "scenario": "Verify admin can publish reply visible on unit page.",
         "steps": [
             ("Select REV-001. Click Reply.", "Reply text area displayed."),
             ("Type reply. Click Publish Reply.", "Reply submitted."),
             ("Observe.", "Reply saved with date. Visible on unit detail page."),
             ("Verify guest sees reply.", "In UC-FB-02, guest sees admin reply."),
         ], "color": NORMAL_FILL},
        {"tab": "Normal Flow – Hide&Restore", "id": "TC-FB-03_NF_HideRestore",
         "desc": "View / Manage All Feedback (Admin) – Hide & Restore Review",
         "prereqs": ["REV-002 with inappropriate content is visible."],
         "data": ["REV-002: offensive content"],
         "scenario": "Verify admin can hide review (removes from public, recalcs rating) and restore it.",
         "steps": [
             ("Select REV-002. Click Hide.", "Confirmation: 'Hide this review from public view?'"),
             ("Confirm.", "REV-002 hidden. Avg rating recalculated (excluding REV-002). Removed from guest-facing page."),
             ("Select hidden REV-002. Click Restore.", "Visibility = 'visible'. Avg rating recalculated (including REV-002)."),
         ], "color": NORMAL_FILL},
        {"tab": "Alternate Flow", "id": "TC-FB-03_AF",
         "desc": "View / Manage All Feedback (Admin) – Alternate Flows",
         "prereqs": ["REV-001 has existing reply."],
         "data": ["A1: Cancel hide dialog", "A2: Edit reply to 'Thank you for staying!'"],
         "scenario": "Verify cancel hide makes no changes; admin can edit existing reply.",
         "steps": [
             ("(A1) Click Hide on REV-003. When dialog appears, click Cancel.", "No changes. REV-003 remains visible."),
             ("(A2) Select REV-001. Click Edit Reply. Update. Save.", "Reply updated. Reply date updated. Visible on unit page."),
         ], "color": ALT_FILL},
    ])


def gen_fb04():
    make_wb("TC-FB-04_DisplayAverageRating.xlsx", [
        {"tab": "Normal Flow", "id": "TC-FB-04_NF",
         "desc": "Display Average Rating – Normal Flow",
         "prereqs": ["'Cozy Cottage': 3 visible reviews (4,5,3 stars). 1 hidden (2 stars)."],
         "data": ["Visible reviews: [4,5,3] avg=4.0", "Hidden: [2] excluded"],
         "scenario": "Verify avg rating from visible reviews only, displayed on listing and detail.",
         "steps": [
             ("Navigate to listing page.", "System retrieves visible reviews per unit."),
             ("Observe Cozy Cottage card.", "'4.0 ★ (3 reviews)'. Hidden 2-star NOT included."),
             ("Navigate to Cozy Cottage detail.", "Same '4.0 ★ (3 reviews)' shown."),
             ("New visible review submitted (5 stars).", "Avg recalculated: (4+5+3+5)/4=4.25. '4.3 ★ (4 reviews)' on listing and detail."),
         ], "color": NORMAL_FILL},
        {"tab": "Alternate Flow", "id": "TC-FB-04_AF",
         "desc": "Display Average Rating – Alternate Flow (No Reviews)",
         "prereqs": ["Unit 'New Unit' has no reviews or all hidden."],
         "data": ["Unit: 0 visible reviews"],
         "scenario": "Verify 'No reviews yet' shown when no visible reviews.",
         "steps": [
             ("Navigate to listing.", "Retrieves visible reviews for 'New Unit'."),
             ("Observe 'New Unit' card.", "'No reviews yet' shown."),
             ("Navigate to 'New Unit' detail.", "'No reviews yet'. No star rating or count."),
         ], "color": ALT_FILL},
    ])


# ══════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════
if __name__ == "__main__":
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    print("Generating HomeLodge v2 Test Case files...\n")

    print("Module 1 – Authentication")
    gen_auth01(); gen_auth02(); gen_auth03()
    gen_auth04(); gen_auth05(); gen_auth06()

    print("Module 2 – Homestay Management")
    gen_hs01(); gen_hs02(); gen_hs03()
    gen_hs04(); gen_hs05(); gen_hs06()

    print("Module 3 – Booking")
    gen_bk01(); gen_bk02(); gen_bk03()
    gen_bk04(); gen_bk05(); gen_bk06(); gen_bk07()

    print("Module 4 – Payment")
    gen_pay01(); gen_pay02(); gen_pay03()

    print("Module 5 – Notification")
    gen_notif01(); gen_notif02()

    print("Module 6 – Chat")
    gen_chat01(); gen_chat02()

    print("Module 7 – User & Access Management")
    gen_usr01(); gen_usr02(); gen_usr03()
    gen_usr04(); gen_usr05()

    print("Module 8 – System Settings")
    gen_set01(); gen_set02()

    print("Module 9 – Audit Logs")
    gen_audit01(); gen_audit02()

    print("Module 10 – QR Code & Access")
    gen_qr01(); gen_qr02(); gen_qr03()
    gen_qr04(); gen_qr05()

    print("Module 11 – Reporting & Analytics")
    gen_rpt01(); gen_rpt02(); gen_rpt03()

    print("Module 12 – Guest Feedback")
    gen_fb01(); gen_fb02(); gen_fb03(); gen_fb04()

    print("\n✅ All 47 test case Excel files generated!")
    print(f"   Folder: {OUTPUT_DIR}")
